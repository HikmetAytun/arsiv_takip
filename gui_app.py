"""
Arşiv Takip Sistemi — GUI (PySide6)
Premium kurumsal tasarım — v3
"""

import sys
import math
from datetime import date, datetime
from pathlib import Path

import pandas as pd
from PySide6.QtCore import (
    Qt, QDate, QTimer, QSize, QPoint, QRect, QPropertyAnimation,
    QEasingCurve, QRectF, Signal,
)
from PySide6.QtGui import (
    QAction, QColor, QFont, QLinearGradient, QPainter,
    QPainterPath, QPen, QBrush, QPixmap, QGradient,
)
from PySide6.QtWidgets import (
    QApplication, QCheckBox, QComboBox, QDateEdit, QDialog, QFileDialog,
    QFormLayout, QGridLayout, QHBoxLayout, QHeaderView, QInputDialog,
    QLabel, QLineEdit, QMainWindow, QMessageBox, QPushButton,
    QSizePolicy, QTableWidget, QTableWidgetItem, QTabWidget,
    QTextEdit, QVBoxLayout, QWidget, QFrame, QScrollArea,
    QStackedWidget, QAbstractItemView, QGraphicsDropShadowEffect,
    QSpacerItem,
)

from db import (
    DB_YOLU, action_log_ekle, action_loglarini_getir,
    acik_movement_var_mi, dosya_ve_hareket_ekle, excel_verisini_yukle,
    file_arsive_al, file_gecmisi_getir, file_guncelle, file_sil,
    giris_yap, ilce_bazli_istatistik, istatistik_ozet,
    kullanici_durum_degistir, kullanici_ekle, kullanici_guncelle,
    kullanici_sifre_sifirla, login_loglarini_getir, movement_ekle,
    personel_bazli_istatistik, tablo_olustur, tum_files_ozet,
    tum_kullanicilari_getir, varsayilan_kullanicilari_olustur,
    parse_ilce_detay_from_text, migrate_legacy_dosyalar_if_needed,
    veritabani_yedekle, son_yedek_bilgisi,
    mesaj_tablolari_olustur, mesaj_gonder, mesajlari_getir,
    mesaj_oku, okunmamis_mesaj_sayisi, konusma_gecmisi,
    tum_mesajlari_oku, online_tablolari_olustur, presence_guncelle,
    online_kullanici_bilgileri, mesaj_gonder_dosya_ref,
    mesaj_sil, konusma_listesi_getir, duyuru_listesi_getir,
    toplu_mesaj_sil, konusma_sil,
    son_hareketleri_getir, trend_verisi_getir, ozet_istatistik_gelismis,
    bende_zimmetli_dosyalar,
    movement_user_id_guncelle,
    arsive_gonder, arsive_gonder_iptal, arsive_gonderilen_dosyalar,
    arsiv_gorevlisini_getir, tum_arsiv_gorevlileri,
)



# İzmir ilçeleri — tam liste
ILCE_LISTESI = [
    "ALİAĞA", "BALÇOVA", "BAYINDIR", "BAYRAKLI", "BERGAMA",
    "BEYDAĞ", "BORNOVA", "BUCA", "ÇEŞME", "ÇİĞLİ",
    "DİKİLİ", "FOÇA", "GAZİEMİR", "GÜZELBAHÇE", "KARABURUN",
    "KARABAĞLAR", "KARŞIYAKA", "KEMALPAŞA", "KINIK", "KİRAZ",
    "KONAK", "MENDERES", "MENEMEN", "NARLIDERE", "ÖDEMİŞ",
    "SEFERİHİSAR", "SELÇUK", "TİRE", "TORBALI", "URLA",
]

# Müdürlük listesi — sabit dropdown
MUDÜRLUK_LISTESI = [
    "EMLAK ŞB. MÜD.",
    "KAMULAŞTIRMA ŞB. MÜD.",
    "KİRALAMA VE TAKİP ŞB.MÜD.",
    "GAYRİ. GEL. VE YÖN. ŞB. MÜD.",
]


def _varsayilan_ods_yolu():
    for p in [Path("data/arsiv_2026.ods"), Path("arsiv_2026.ods")]:
        if p.exists():
            return p
    return Path("data/arsiv_2026.ods")


DOSYA_YOLU    = _varsayilan_ods_yolu()
APP_TITLE     = "Arşiv Takip Sistemi"
APP_SAHIP     = "Hikmet Aytun"
APP_IMZA      = "© 2026 Hikmet Aytun — Tüm hakları saklıdır."
DESTEK_TEL    = "Dahili: 3622"
APP_VERSIYON  = "v3.0"

ROL_ETIKET = {
    "admin":  "Sistem Yöneticisi",
    "arsiv":  "Arşiv Görevlisi",
    "viewer": "Görüntüleyici",
}

# ─────────────────────────────────────────────────────────────
# OTOMATİK EXCEL YEDEĞİ
# ─────────────────────────────────────────────────────────────
# Ağ klasörü yolu — IT ile koordineli ayarlayın
# Boş bırakılırsa sadece yerel klasöre kaydeder
EXCEL_AG_KLASORU = ""   # Örn: r"\\SERVER\arsiv_yedek"

def otomatik_excel_yedek(ag_klasoru: str = EXCEL_AG_KLASORU):
    """
    Tüm dosyaları Excel'e döker.
    - Yerel: backups/excel/ klasörüne
    - Ağ: EXCEL_AG_KLASORU'na (ayarlıysa)
    Her kayıtta çağrılır.
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from db import tum_files_ozet
        import datetime as _dt

        veriler = tum_files_ozet()
        simdi   = _dt.datetime.now()
        dosya_adi = f"arsiv_{simdi.strftime('%Y%m%d_%H%M%S')}.xlsx"

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Zimmet Listesi"

        # Başlıklar
        basliklar = [
            "Dosya No", "İlçe", "Şefliği", "Ada", "Parsel",
            "Durum", "Teslim Alan", "Arşiv Görevlisi",
            "Teslim Tarihi", "Bekleme (gün)", "Hareket Sayısı"
        ]
        basl_stil = Font(bold=True, color="FFFFFF")
        basl_dolu = PatternFill("solid", fgColor="0F172A")
        orta      = Alignment(horizontal="center", vertical="center")

        ws.append(basliklar)
        for cell in ws[1]:
            cell.font      = basl_stil
            cell.fill      = basl_dolu
            cell.alignment = orta

        ws.row_dimensions[1].height = 28

        # Durum renkleri
        renk_map = {
            "ARŞİVDE":           "ECFDF5",
            "ZİMMETTE":          "EFF6FF",
            "GECİKMİŞ":         "FEF2F2",
            "ARŞİVE GÖNDERİLDİ":"FFFBEB",
        }

        for ri, r in enumerate(veriler, 2):
            durum = r.get("durum", "")
            satir = [
                r.get("orijinal_dosya_no",""),
                r.get("ilce",""),
                r.get("sefligi",""),
                r.get("ada",""),
                r.get("parsel",""),
                durum,
                r.get("teslim_alan_personel",""),
                r.get("veren_arsiv_gorevlisi",""),
                r.get("teslim_tarihi",""),
                r.get("bekleme_gun", 0),
                r.get("hareket_sayisi", 0),
            ]
            ws.append(satir)
            dolu_renk = renk_map.get(durum, "FFFFFF")
            dolu = PatternFill("solid", fgColor=dolu_renk)
            for cell in ws[ri]:
                cell.fill      = dolu
                cell.alignment = orta

        # Sütun genişlikleri
        for col, w in zip("ABCDEFGHIJK", [12,15,18,8,8,20,18,18,14,12,12]):
            ws.column_dimensions[col].width = w

        # Özet sekme
        ws2 = wb.create_sheet("Özet")
        ws2.append(["Tarih", simdi.strftime("%d.%m.%Y %H:%M")])
        ws2.append(["Toplam Dosya", len(veriler)])
        ws2.append(["Arşivde", sum(1 for r in veriler if r.get("durum")=="ARŞİVDE")])
        ws2.append(["Zimmette", sum(1 for r in veriler if r.get("durum")=="ZİMMETTE")])
        ws2.append(["Gecikmiş", sum(1 for r in veriler if r.get("durum")=="GECİKMİŞ")])
        ws2.append(["Arşive Gönderildi",
                    sum(1 for r in veriler if r.get("durum")=="ARŞİVE GÖNDERİLDİ")])

        # Yerel kayıt
        yerel_klasor = Path("backups") / "excel"
        yerel_klasor.mkdir(parents=True, exist_ok=True)
        yerel_yol = yerel_klasor / dosya_adi
        wb.save(str(yerel_yol))

        # Eski yedekleri temizle (son 30 tut)
        tum_yedekler = sorted(yerel_klasor.glob("arsiv_*.xlsx"))
        for eski in tum_yedekler[:-30]:
            try: eski.unlink()
            except Exception: pass

        # Ağ klasörüne kopyala
        if ag_klasoru:
            try:
                import shutil
                ag_path = Path(ag_klasoru)
                ag_path.mkdir(parents=True, exist_ok=True)
                shutil.copy2(str(yerel_yol), str(ag_path / dosya_adi))
                # Ağda da son 30 tut
                ag_yedekler = sorted(ag_path.glob("arsiv_*.xlsx"))
                for eski in ag_yedekler[:-30]:
                    try: eski.unlink()
                    except Exception: pass
            except Exception:
                pass  # Ağ erişim hatası sessizce geç

    except Exception:
        pass  # Yedek hatası uygulamayı durdurmasın



# ─────────────────────────────────────────────────────────────
# RENK SİSTEMİ
# ─────────────────────────────────────────────────────────────
P = {
    # Temel arka plan — hafif soğuk gri
    "bg":           "#F1F5F9",
    "surface":      "#FFFFFF",
    "surface2":     "#F8FAFC",
    "border":       "#E2E8F0",
    "border2":      "#CBD5E1",

    # Sidebar — derin koyu navy
    "navy":         "#0F172A",
    "navy2":        "#1E293B",
    "navy3":        "#334155",
    "navy_text":    "#94A3B8",
    "navy_text_a":  "#F1F5F9",
    "navy_active":  "#3B82F6",

    # Yazı
    "txt":          "#0F172A",
    "txt2":         "#1E293B",
    "txt3":         "#475569",
    "txt4":         "#94A3B8",

    # Mavi — canlı
    "blue":         "#2563EB",
    "blue2":        "#1D4ED8",
    "blue_bg":      "#EFF6FF",
    "blue_t":       "#1E40AF",

    # Yeşil
    "green":        "#059669",
    "green_bg":     "#ECFDF5",
    "green_t":      "#065F46",

    # Amber
    "amber":        "#D97706",
    "amber_bg":     "#FFFBEB",
    "amber_t":      "#92400E",

    # Kırmızı
    "red":          "#DC2626",
    "red_bg":       "#FEF2F2",
    "red_t":        "#991B1B",

    # Mor
    "purple":       "#7C3AED",
    "purple_bg":    "#F5F3FF",
    "purple_t":     "#4C1D95",

    # Tablo satır renkleri
    "row_red":      "#FFF1F1",
    "row_yellow":   "#FEFCE8",
    "row_white":    "#FFFFFF",
}

# ─────────────────────────────────────────────────────────────
# ANA STİL
# ─────────────────────────────────────────────────────────────
ANA_STIL = f"""

QWidget {{
    font-family: -apple-system, 'Segoe UI', 'SF Pro Text', 'Inter', sans-serif;
    font-size: 13px;
    color: {P['txt']};
    background-color: {P['bg']};
}}
QMainWindow {{ background: {P['bg']}; }}
QDialog {{ background: {P['surface']}; }}

/* ── Scrollbar ── */
QScrollBar:vertical {{
    background: transparent; width: 6px; margin: 0;
}}
QScrollBar::handle:vertical {{
    background: {P['border2']}; border-radius: 3px; min-height: 30px;
}}
QScrollBar::handle:vertical:hover {{ background: {P['txt4']}; }}
QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical {{ height: 0; }}
QScrollBar:horizontal {{
    background: transparent; height: 6px;
}}
QScrollBar::handle:horizontal {{
    background: {P['border2']}; border-radius: 3px;
}}

/* ── Inputs ── */
QLineEdit, QTextEdit {{
    background: {P['surface']};
    border: 1.5px solid {P['border']};
    border-radius: 10px;
    padding: 9px 14px;
    color: {P['txt']};
    font-size: 13px;
    selection-background-color: {P['blue']};
    selection-color: white;
}}
QLineEdit:focus, QTextEdit:focus {{
    border: 1.5px solid {P['blue']};
    background: #FAFCFF;
}}
QLineEdit:hover {{ border-color: {P['border2']}; }}
QLineEdit::placeholder {{ color: {P['txt4']}; }}

/* ── ComboBox ── */
QComboBox {{
    background: {P['surface']};
    border: 1.5px solid {P['border']};
    border-radius: 10px;
    padding: 8px 14px;
    color: {P['txt']};
    min-width: 130px;
    font-size: 13px;
}}
QComboBox:hover {{ border-color: {P['border2']}; }}
QComboBox:focus {{ border-color: {P['blue']}; }}
QComboBox::drop-down {{ border: none; width: 28px; subcontrol-position: right center; }}
QComboBox::down-arrow {{
    image: none; width: 0; height: 0;
    border-left: 4px solid transparent;
    border-right: 4px solid transparent;
    border-top: 5px solid {P['txt3']};
    margin-right: 10px;
}}
QComboBox QAbstractItemView {{
    background: {P['surface']};
    border: 1px solid {P['border']};
    border-radius: 10px;
    selection-background-color: {P['blue_bg']};
    selection-color: {P['blue_t']};
    padding: 4px; outline: none;
}}

/* ── DateEdit ── */
QDateEdit {{
    background: {P['surface']};
    border: 1.5px solid {P['border']};
    border-radius: 10px;
    padding: 8px 14px;
    color: {P['txt']};
}}
QDateEdit:focus {{ border-color: {P['blue']}; }}
QDateEdit::drop-down {{ border: none; width: 24px; }}
QDateEdit::down-arrow {{
    image: none; width: 0; height: 0;
    border-left: 4px solid transparent;
    border-right: 4px solid transparent;
    border-top: 5px solid {P['txt3']};
    margin-right: 8px;
}}

/* ── Buttons ── */
QPushButton {{
    background: {P['blue']};
    color: white;
    border: none;
    border-radius: 10px;
    padding: 10px 20px;
    font-weight: 600;
    font-size: 13px;
}}
QPushButton:hover {{ background: {P['blue2']}; }}
QPushButton:pressed {{ background: #1E3A8A; }}
QPushButton:disabled {{ background: {P['border']}; color: {P['txt4']}; }}

QPushButton#ghost {{
    background: {P['surface']};
    color: {P['txt2']};
    border: 1.5px solid {P['border']};
}}
QPushButton#ghost:hover {{
    background: {P['bg']}; border-color: {P['border2']}; color: {P['txt']};
}}
QPushButton#success {{ background: {P['green']}; }}
QPushButton#success:hover {{ background: #047857; }}
QPushButton#danger  {{ background: {P['red']};   }}
QPushButton#danger:hover  {{ background: #B91C1C; }}
QPushButton#warning {{ background: {P['amber']}; }}
QPushButton#warning:hover {{ background: #B45309; }}
QPushButton#flat {{
    background: transparent; color: {P['blue']};
    border: none; padding: 6px 10px; font-weight: 500;
}}
QPushButton#flat:hover {{ background: {P['blue_bg']}; border-radius: 8px; }}

/* ── Table ── */
QTableWidget {{
    background: {P['surface']};
    border: 1px solid {P['border']};
    border-radius: 14px;
    gridline-color: {P['border']};
    selection-background-color: transparent;
    outline: none;
    font-size: 13px;
    alternate-background-color: #FAFBFC;
}}
QTableWidget::item {{
    padding: 11px 14px;
    border: none;
    border-bottom: 1px solid #F1F5F9;
}}
QTableWidget::item:selected {{
    background: {P['blue_bg']};
    color: {P['blue_t']};
    border-radius: 4px;
}}
QTableWidget::item:hover {{ background: #F0F7FF; }}
QHeaderView::section {{
    background: #F8FAFC;
    color: {P['txt3']};
    padding: 12px 14px;
    border: none;
    border-bottom: 2px solid {P['border']};
    font-weight: 700;
    font-size: 11px;
    letter-spacing: 0.8px;
    text-transform: uppercase;
}}
QHeaderView::section:first {{ border-radius: 14px 0 0 0; }}
QHeaderView::section:last  {{ border-radius: 0 14px 0 0; }}
QHeaderView::section:hover {{ background: {P['border']}; }}

/* ── Tabs ── */
QTabWidget::pane {{
    border: 1px solid {P['border']}; border-radius: 12px;
    background: {P['surface']}; top: -1px;
}}
QTabBar::tab {{
    background: transparent; color: {P['txt3']};
    padding: 11px 22px; margin-right: 2px;
    border-bottom: 2px solid transparent;
    font-weight: 500; font-size: 13px;
}}
QTabBar::tab:selected {{
    color: {P['blue']}; border-bottom: 2px solid {P['blue']}; font-weight: 700;
}}
QTabBar::tab:hover:!selected {{
    color: {P['txt']}; background: {P['bg']}; border-radius: 8px 8px 0 0;
}}

/* ── GroupBox ── */
QGroupBox {{
    background: {P['surface']}; border: 1px solid {P['border']};
    border-radius: 14px; margin-top: 8px;
    padding: 20px 16px 16px 16px; font-weight: 700;
}}
QGroupBox::title {{
    subcontrol-origin: margin; subcontrol-position: top left;
    left: 18px; top: 4px;
    color: {P['txt3']}; font-size: 10px;
    font-weight: 700; letter-spacing: 0.8px; text-transform: uppercase;
}}

/* ── Labels & Menu ── */
QLabel {{ color: {P['txt']}; background: transparent; }}
QMenuBar {{
    background: {P['surface']};
    border-bottom: 1px solid {P['border']};
    color: {P['txt2']}; font-size: 13px; padding: 2px 0;
}}
QMenuBar::item:selected {{
    background: {P['blue_bg']}; color: {P['blue_t']}; border-radius: 6px;
}}
QMenu {{
    background: {P['surface']}; border: 1px solid {P['border']};
    border-radius: 12px; padding: 6px;
}}
QMenu::item {{ padding: 8px 16px; border-radius: 8px; color: {P['txt2']}; }}
QMenu::item:selected {{ background: {P['blue_bg']}; color: {P['blue_t']}; }}
QMenu::separator {{ background: {P['border']}; height: 1px; margin: 4px 8px; }}

/* ── StatusBar ── */
QStatusBar {{
    background: {P['surface']};
    border-top: 1px solid {P['border']};
    color: {P['txt3']}; font-size: 11px; padding: 2px 8px;
}}
QStatusBar::item {{ border: none; }}
"""

KOYU_STIL = f"""
QWidget {{
    font-family: -apple-system, 'Segoe UI', 'SF Pro Text', 'Inter', sans-serif;
    font-size: 13px;
    color: #E2E8F0;
    background-color: #0D1117;
}}
QMainWindow {{ background: #0D1117; }}
QDialog {{ background: #161B22; }}
QFrame {{ color: #E2E8F0; }}
QLabel {{ color: #E2E8F0; background: transparent; }}

QScrollBar:vertical {{ background: transparent; width: 6px; margin: 0; }}
QScrollBar::handle:vertical {{ background: #30363D; border-radius: 3px; min-height: 30px; }}
QScrollBar::handle:vertical:hover {{ background: #484F58; }}
QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical {{ height: 0; }}
QScrollBar:horizontal {{ background: transparent; height: 6px; }}
QScrollBar::handle:horizontal {{ background: #30363D; border-radius: 3px; }}

QLineEdit, QTextEdit {{
    background: #161B22;
    border: 1.5px solid #30363D;
    border-radius: 10px;
    padding: 9px 14px;
    color: #E2E8F0;
    selection-background-color: #1D4ED8;
    selection-color: white;
}}
QLineEdit:focus, QTextEdit:focus {{ border-color: #2563EB; background: #1C2128; }}
QLineEdit:hover {{ border-color: #484F58; }}
QLineEdit::placeholder {{ color: #6E7681; }}

QComboBox {{
    background: #161B22;
    border: 1.5px solid #30363D;
    border-radius: 10px;
    padding: 8px 14px;
    color: #E2E8F0;
    min-width: 130px;
}}
QComboBox:hover {{ border-color: #484F58; }}
QComboBox:focus {{ border-color: #2563EB; }}
QComboBox::drop-down {{ border: none; width: 28px; subcontrol-position: right center; }}
QComboBox::down-arrow {{
    image: none; width: 0; height: 0;
    border-left: 4px solid transparent;
    border-right: 4px solid transparent;
    border-top: 5px solid #6E7681;
    margin-right: 10px;
}}
QComboBox QAbstractItemView {{
    background: #1C2128;
    border: 1px solid #30363D;
    border-radius: 10px;
    color: #E2E8F0;
    selection-background-color: #1F2D40;
    selection-color: #60A5FA;
    padding: 4px; outline: none;
}}

QDateEdit {{
    background: #161B22; border: 1.5px solid #30363D;
    border-radius: 10px; padding: 8px 14px; color: #E2E8F0;
}}
QDateEdit:focus {{ border-color: #2563EB; }}
QDateEdit::drop-down {{ border: none; width: 24px; }}
QDateEdit::down-arrow {{
    image: none; width: 0; height: 0;
    border-left: 4px solid transparent;
    border-right: 4px solid transparent;
    border-top: 5px solid #6E7681; margin-right: 8px;
}}

QPushButton {{
    background: #1D4ED8; color: white; border: none;
    border-radius: 10px; padding: 10px 20px;
    font-weight: 600; font-size: 13px;
}}
QPushButton:hover {{ background: #2563EB; }}
QPushButton:pressed {{ background: #1E40AF; }}
QPushButton:disabled {{ background: #21262D; color: #6E7681; }}
QPushButton#ghost {{
    background: #161B22; color: #C9D1D9;
    border: 1.5px solid #30363D;
}}
QPushButton#ghost:hover {{ background: #1C2128; border-color: #484F58; color: #E2E8F0; }}
QPushButton#success {{ background: #059669; }} QPushButton#success:hover {{ background: #047857; }}
QPushButton#danger  {{ background: #DC2626; }} QPushButton#danger:hover  {{ background: #B91C1C; }}
QPushButton#warning {{ background: #D97706; }} QPushButton#warning:hover {{ background: #B45309; }}
QPushButton#flat {{ background: transparent; color: #60A5FA; border: none; padding: 6px 10px; font-weight: 500; }}
QPushButton#flat:hover {{ background: #1F2D40; border-radius: 8px; }}

QTableWidget {{
    background: #161B22;
    border: 1px solid #30363D;
    border-radius: 14px;
    gridline-color: #21262D;
    color: #E2E8F0;
    outline: none;
}}
QTableWidget::item {{ padding: 11px 14px; border: none; border-bottom: 1px solid #21262D; color: #E2E8F0; }}
QTableWidget::item:selected {{ background: #1F2D40; color: #60A5FA; }}
QTableWidget::item:hover {{ background: #1C2128; }}
QHeaderView::section {{
    background: #0D1117; color: #8B949E;
    border: none; border-bottom: 2px solid #21262D;
    padding: 12px 14px; font-weight: 700; font-size: 11px; letter-spacing: 0.8px;
}}
QHeaderView::section:hover {{ background: #161B22; }}

QTabWidget::pane {{ border: 1px solid #30363D; border-radius: 12px; background: #161B22; top: -1px; }}
QTabBar::tab {{ background: transparent; color: #6E7681; padding: 11px 22px; margin-right: 2px; border-bottom: 2px solid transparent; font-weight: 500; font-size: 13px; }}
QTabBar::tab:selected {{ color: #60A5FA; border-bottom: 2px solid #2563EB; font-weight: 700; }}
QTabBar::tab:hover:!selected {{ color: #C9D1D9; background: #1C2128; border-radius: 8px 8px 0 0; }}

QGroupBox {{ background: #161B22; border: 1px solid #30363D; border-radius: 14px; margin-top: 8px; padding: 20px 16px 16px 16px; }}
QGroupBox::title {{ color: #6E7681; font-size: 10px; font-weight: 700; letter-spacing: 0.8px; subcontrol-origin: margin; subcontrol-position: top left; left: 18px; top: 4px; }}

QMenuBar {{ background: #161B22; border-bottom: 1px solid #21262D; color: #C9D1D9; font-size: 13px; padding: 2px 0; }}
QMenuBar::item:selected {{ background: #1F2D40; color: #60A5FA; border-radius: 6px; }}
QMenu {{ background: #161B22; border: 1px solid #30363D; border-radius: 12px; padding: 6px; }}
QMenu::item {{ padding: 8px 16px; border-radius: 8px; color: #C9D1D9; }}
QMenu::item:selected {{ background: #1F2D40; color: #60A5FA; }}
QMenu::separator {{ background: #30363D; height: 1px; margin: 4px 8px; }}

QStatusBar {{ background: #161B22; border-top: 1px solid #21262D; color: #6E7681; font-size: 11px; padding: 2px 8px; }}
QStatusBar::item {{ border: none; }}
"""


# ─────────────────────────────────────────────────────────────
# GRAFİK WİDGETLARI (saf Qt ile)
# ─────────────────────────────────────────────────────────────

class DonutChartWidget(QWidget):
    """Durum dağılımı için donut grafik."""
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setMinimumSize(220, 220)
        self._dilimler: list[tuple[float, QColor, str]] = []
        self._merkez_yazi = ""

    def set_data(self, veriler: list[tuple[int, str, str]], merkez: str = ""):
        """veriler: [(sayi, renk_hex, etiket), ...]"""
        toplam = sum(v[0] for v in veriler) or 1
        self._dilimler = [(v[0] / toplam, QColor(v[1]), v[2]) for v in veriler]
        self._merkez_yazi = merkez
        self.update()

    def paintEvent(self, event):
        p = QPainter(self)
        p.setRenderHint(QPainter.Antialiasing)
        w, h = self.width(), self.height()
        boyut = min(w, h) - 20
        cx = (w - boyut) // 2
        cy = (h - boyut) // 2
        rect = QRectF(cx, cy, boyut, boyut)
        delik_boyut = boyut * 0.56
        delik_rect = QRectF(
            cx + (boyut - delik_boyut) / 2,
            cy + (boyut - delik_boyut) / 2,
            delik_boyut, delik_boyut,
        )

        aci = -90.0
        for oran, renk, _ in self._dilimler:
            span = oran * 360.0
            p.setBrush(QBrush(renk))
            p.setPen(QPen(QColor("#FFFFFF"), 2))
            p.drawPie(rect, int(aci * 16), int(span * 16))
            aci += span

        # Delik (beyaz daire)
        p.setBrush(QBrush(QColor(self.palette().color(self.backgroundRole()))))
        p.setPen(Qt.NoPen)
        p.drawEllipse(delik_rect)

        # Merkez yazı
        if self._merkez_yazi:
            p.setPen(QColor(P["txt"]))
            f = QFont()
            f.setPointSize(12)
            f.setBold(True)
            p.setFont(f)
            p.drawText(QRectF(cx, cy, boyut, boyut),
                       Qt.AlignCenter, self._merkez_yazi)
        p.end()


class HBarChartWidget(QWidget):
    """Yatay bar chart — ilçe/personel için."""
    def __init__(self, parent=None):
        super().__init__(parent)
        self._veriler: list[tuple[str, int, int, int]] = []  # (etiket, toplam, zimmette, gecikmis)
        self._maks = 1
        self.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)

    def set_data(self, veriler: list[tuple[str, int, int, int]]):
        self._veriler = veriler[:12]
        self._maks = max((v[1] for v in self._veriler), default=1)
        h = max(300, len(self._veriler) * 44 + 60)
        self.setMinimumHeight(h)
        self.update()

    def paintEvent(self, event):
        if not self._veriler:
            return
        p = QPainter(self)
        p.setRenderHint(QPainter.Antialiasing)

        w = self.width()
        etiket_genislik = 130
        sagbosluk = 50
        grafik_w = w - etiket_genislik - sagbosluk
        satir_h = 36
        ustbosluk = 20

        f_kucuk = QFont()
        f_kucuk.setPointSize(9)
        f_kucuk.setBold(False)
        f_bold = QFont()
        f_bold.setPointSize(9)
        f_bold.setBold(True)

        for i, (etiket, toplam, zimmette, gecikmis) in enumerate(self._veriler):
            y = ustbosluk + i * (satir_h + 8)
            bar_h = 20

            # Etiket
            p.setFont(f_kucuk)
            p.setPen(QColor(P["txt2"]))
            etiket_rect = QRectF(0, y, etiket_genislik - 8, satir_h)
            p.drawText(etiket_rect, Qt.AlignRight | Qt.AlignVCenter,
                       etiket[:16] + ("…" if len(etiket) > 16 else ""))

            bx = etiket_genislik
            bar_y = y + (satir_h - bar_h) // 2

            # Arka plan
            p.setBrush(QColor("#F1F5F9"))
            p.setPen(Qt.NoPen)
            p.drawRoundedRect(QRectF(bx, bar_y, grafik_w, bar_h), 6, 6)

            # Toplam bar (mavi)
            if toplam > 0:
                toplam_w = grafik_w * toplam / self._maks
                p.setBrush(QColor("#BFDBFE"))
                p.drawRoundedRect(QRectF(bx, bar_y, toplam_w, bar_h), 6, 6)

            # Gecikmiş bar (kırmızı — üstüne)
            if gecikmis > 0:
                gec_w = grafik_w * gecikmis / self._maks
                p.setBrush(QColor(P["red"]))
                p.drawRoundedRect(QRectF(bx, bar_y, gec_w, bar_h), 6, 6)

            # Değer etiketi
            p.setFont(f_bold)
            p.setPen(QColor(P["txt2"]))
            toplam_w2 = grafik_w * toplam / self._maks if toplam > 0 else 0
            p.drawText(
                QRectF(bx + toplam_w2 + 6, bar_y, sagbosluk, bar_h),
                Qt.AlignLeft | Qt.AlignVCenter,
                str(toplam),
            )

        p.end()


class RingKarti(QFrame):
    """Dashboard için donut + legend birleşimi."""
    def __init__(self, baslik: str, parent=None):
        super().__init__(parent)
        self.setStyleSheet(f"""
            QFrame {{
                background: {P['surface']};
                border: 1px solid {P['border']};
                border-radius: 16px;
            }}
        """)
        ana = QVBoxLayout(self)
        ana.setContentsMargins(20, 16, 20, 16)
        ana.setSpacing(12)

        b = QLabel(baslik)
        b.setStyleSheet(f"font-size: 13px; font-weight: 700; color: {P['txt2']};")
        ana.addWidget(b)
        ana.addWidget(_sep())

        ic = QHBoxLayout()
        ic.setSpacing(16)
        self._donut = DonutChartWidget()
        self._donut.setFixedSize(160, 160)
        ic.addWidget(self._donut)

        self._legend_lay = QVBoxLayout()
        self._legend_lay.setSpacing(8)
        self._legend_lay.addStretch()
        ic.addLayout(self._legend_lay)
        ana.addLayout(ic)

    def set_data(self, veriler: list[tuple[int, str, str]], merkez: str = ""):
        self._donut.set_data(veriler, merkez)
        # Legend temizle
        while self._legend_lay.count() > 1:
            item = self._legend_lay.takeAt(0)
            if item.widget():
                item.widget().deleteLater()

        toplam = sum(v[0] for v in veriler) or 1
        for sayi, renk, etiket in veriler:
            satir = QHBoxLayout()
            satir.setSpacing(8)
            nokta = QLabel("●")
            nokta.setFixedWidth(14)
            nokta.setStyleSheet(f"color: {renk}; font-size: 14px;")
            lbl = QLabel(f"{etiket}  {sayi:,}".replace(",","."))
            lbl.setStyleSheet(f"font-size: 12px; color: {P['txt2']};")
            pct = QLabel(f"%{sayi*100//toplam}")
            pct.setStyleSheet(f"font-size: 11px; color: {P['txt4']}; font-weight: 600;")
            satir.addWidget(nokta)
            satir.addWidget(lbl)
            satir.addStretch()
            satir.addWidget(pct)
            w = QWidget()
            w.setStyleSheet("background: transparent;")
            w.setLayout(satir)
            self._legend_lay.insertWidget(self._legend_lay.count() - 1, w)


class BarKarti(QFrame):
    """İlçe/personel yatay bar chart kartı."""
    def __init__(self, baslik: str, parent=None):
        super().__init__(parent)
        self.setStyleSheet(f"""
            QFrame {{
                background: {P['surface']};
                border: 1px solid {P['border']};
                border-radius: 16px;
            }}
        """)
        ana = QVBoxLayout(self)
        ana.setContentsMargins(20, 16, 20, 16)
        ana.setSpacing(12)

        ust = QHBoxLayout()
        b = QLabel(baslik)
        b.setStyleSheet(f"font-size: 13px; font-weight: 700; color: {P['txt2']};")
        ust.addWidget(b)
        ust.addStretch()

        # Renk açıklaması
        for renk, txt in [("#BFDBFE", "Toplam"), (P["red"], "Gecikmiş")]:
            nokta = QLabel("■")
            nokta.setStyleSheet(f"color: {renk}; font-size: 12px;")
            lbl = QLabel(txt)
            lbl.setStyleSheet(f"font-size: 11px; color: {P['txt3']};")
            ust.addWidget(nokta)
            ust.addWidget(lbl)

        ana.addLayout(ust)
        ana.addWidget(_sep())

        scroll = QScrollArea()
        scroll.setFrameShape(QFrame.NoFrame)
        scroll.setWidgetResizable(True)
        scroll.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        self._chart = HBarChartWidget()
        scroll.setWidget(self._chart)
        scroll.setMinimumHeight(300)
        ana.addWidget(scroll)

    def set_data(self, veriler):
        self._chart.set_data(veriler)


# ─────────────────────────────────────────────────────────────
# MINI ÇUBUK GRAFİK WİDGET
# ─────────────────────────────────────────────────────────────
class MiniBarWidget(QWidget):
    """Satır içi mini yatay bar — istatistik tablolarında kullanılır."""
    def __init__(self, deger: int, maksimum: int, renk: str, parent=None):
        super().__init__(parent)
        self._deger = deger
        self._max = maksimum or 1
        self._renk = QColor(renk)
        self.setFixedHeight(20)
        self.setMinimumWidth(80)

    def paintEvent(self, event):
        p = QPainter(self)
        p.setRenderHint(QPainter.Antialiasing)
        w = self.width()
        h = self.height()
        # Arka plan
        p.setBrush(QColor("#F1F5F9"))
        p.setPen(Qt.NoPen)
        p.drawRoundedRect(0, (h - 8) // 2, w, 8, 4, 4)
        # Dolu kısım
        if self._max > 0:
            dolu = int(w * self._deger / self._max)
            if dolu > 0:
                p.setBrush(self._renk)
                p.drawRoundedRect(0, (h - 8) // 2, dolu, 8, 4, 4)
        p.end()


# ─────────────────────────────────────────────────────────────
# YARDIMCI WİDGETLAR
# ─────────────────────────────────────────────────────────────
class EtiketBadge(QLabel):
    """Renkli durum rozeti."""
    TEMA = {
        "GEC": ("#FEF2F2", "#DC2626", "#991B1B"),   # kırmızı
        "ZIM": ("#EFF6FF", "#2563EB", "#1E40AF"),   # mavi
        "ARS": ("#ECFDF5", "#059669", "#065F46"),   # yeşil
        "GON": ("#FFFBEB", "#D97706", "#92400E"),   # turuncu — arşive gönderildi
    }
    def __init__(self, metin: str, parent=None):
        super().__init__(metin, parent)
        d = (metin or "").upper()
        if "GEC" in d: key = "GEC"
        elif "GÖN" in d or "GON" in d: key = "GON"
        elif "Z" in d and "MM" in d: key = "ZIM"
        else: key = "ARS"
        bg, border, fg = self.TEMA[key]
        self.setStyleSheet(f"""
            background: {bg};
            color: {fg};
            border: 1px solid {border}30;
            border-radius: 20px;
            padding: 3px 12px;
            font-size: 11px;
            font-weight: 700;
            letter-spacing: 0.3px;
        """)
        self.setAlignment(Qt.AlignCenter)
        self.setFixedHeight(24)


class KartMetrik(QFrame):
    """Ana panel için büyük metrik kart — modern flat tasarım."""
    TEMA = {
        "gray":   ("#F1F5F9", "#94A3B8", P["txt3"],   "📦"),
        "green":  ("#ECFDF5", "#059669", P["green"],  "🏛"),
        "blue":   ("#EFF6FF", "#2563EB", P["blue"],   "📋"),
        "red":    ("#FEF2F2", "#DC2626", P["red"],    "⚠️"),
        "purple": ("#F5F3FF", "#7C3AED", P["purple"], "📊"),
    }

    def __init__(self, baslik: str, tema: str = "blue", parent=None):
        super().__init__(parent)
        bg, accent, clr, ikon = self.TEMA.get(tema, self.TEMA["blue"])
        self.setFixedHeight(106)
        self.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        self.setStyleSheet(f"""
            QFrame {{
                background: {P['surface']};
                border: 1px solid {P['border']};
                border-radius: 14px;
            }}
        """)

        # Sol renk şeridi
        serit = QFrame(self)
        serit.setGeometry(0, 0, 4, 106)
        serit.setStyleSheet(f"background:{clr}; border-radius:14px 0 0 14px;")

        lay = QHBoxLayout(self)
        lay.setContentsMargins(22, 14, 18, 14)
        lay.setSpacing(14)

        # İkon dairesi
        ikon_lbl = QLabel(ikon)
        ikon_lbl.setFixedSize(46, 46)
        ikon_lbl.setAlignment(Qt.AlignCenter)
        ikon_lbl.setStyleSheet(f"""
            background: {bg};
            border-radius: 23px;
            font-size: 20px;
        """)
        lay.addWidget(ikon_lbl)

        txt_lay = QVBoxLayout()
        txt_lay.setSpacing(3)
        txt_lay.setContentsMargins(0, 0, 0, 0)

        self._sayi = QLabel("—")
        self._sayi.setStyleSheet(f"""
            font-size: 28px; font-weight: 800;
            color: {clr}; letter-spacing: -1px;
        """)
        self._baslik = QLabel(baslik)
        self._baslik.setStyleSheet(f"""
            font-size: 11px; font-weight: 600;
            color: {P['txt4']}; letter-spacing: 0.5px;
        """)
        self._alt = QLabel("")
        self._alt.setStyleSheet(f"font-size: 10px; color: {P['txt4']};")

        txt_lay.addStretch()
        txt_lay.addWidget(self._sayi)
        txt_lay.addWidget(self._baslik)
        txt_lay.addWidget(self._alt)
        txt_lay.addStretch()
        lay.addLayout(txt_lay)
        lay.addStretch()

    def guncelle(self, sayi: int, alt: str = ""):
        self._sayi.setText(f"{sayi:,}".replace(",", "."))
        if alt:
            self._alt.setText(alt)

    def set_deger(self, deger: str, alt: str = ""):
        """String değer ile güncelle (int olmayabilir)."""
        self._sayi.setText(deger)
        if alt:
            self._alt.setText(alt)


class TrendGrafikWidget(QWidget):
    """Son 30 günün zimmet/arşivleme trend çizgi grafiği — saf Qt."""

    def __init__(self, parent=None):
        super().__init__(parent)
        self._trend: list[dict] = []
        self.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)

    def set_data(self, trend: list[dict]):
        self._trend = trend
        self.update()

    def paintEvent(self, event):
        if not self._trend:
            p = QPainter(self)
            p.setPen(QColor(P["txt4"]))
            p.drawText(self.rect(), Qt.AlignCenter, "Veri yok")
            p.end()
            return

        p = QPainter(self)
        p.setRenderHint(QPainter.Antialiasing)
        w, h = self.width(), self.height()
        pad_l, pad_r, pad_t, pad_b = 32, 12, 10, 24

        zimmet = [int(r.get("zimmet_sayisi") or 0) for r in self._trend]
        arsiv  = [int(r.get("arsiv_sayisi")  or 0) for r in self._trend]
        gunler = [r.get("gun","")[-5:] for r in self._trend]  # MM-DD

        maks = max(max(zimmet, default=0), max(arsiv, default=0), 1)
        n = len(self._trend)

        gw = w - pad_l - pad_r
        gh = h - pad_t - pad_b

        def x_pos(i):
            return pad_l + (i / max(n-1, 1)) * gw

        def y_pos(val):
            return pad_t + gh - (val / maks) * gh

        # Izgara çizgileri
        p.setPen(QPen(QColor(P["border"]), 0.5, Qt.DashLine))
        for step in [0.25, 0.5, 0.75, 1.0]:
            y = pad_t + gh * (1 - step)
            p.drawLine(int(pad_l), int(y), int(w - pad_r), int(y))
            p.setPen(QColor(P["txt4"]))
            f = QFont(); f.setPointSize(8); p.setFont(f)
            p.drawText(2, int(y) + 4, str(int(maks * step)))
            p.setPen(QPen(QColor(P["border"]), 0.5, Qt.DashLine))

        # Zimmet çizgisi (mavi)
        if n > 1:
            pen = QPen(QColor(P["blue"]), 2)
            pen.setCapStyle(Qt.RoundCap); pen.setJoinStyle(Qt.RoundJoin)
            p.setPen(pen)
            for i in range(1, n):
                p.drawLine(int(x_pos(i-1)), int(y_pos(zimmet[i-1])),
                           int(x_pos(i)),   int(y_pos(zimmet[i])))
            # Nokta
            p.setBrush(QBrush(QColor(P["blue"]))); p.setPen(Qt.NoPen)
            for i in range(n):
                p.drawEllipse(int(x_pos(i))-3, int(y_pos(zimmet[i]))-3, 6, 6)

        # Arşiv çizgisi (yeşil)
        if n > 1 and any(a > 0 for a in arsiv):
            pen2 = QPen(QColor(P["green"]), 2)
            pen2.setCapStyle(Qt.RoundCap); pen2.setJoinStyle(Qt.RoundJoin)
            p.setPen(pen2)
            for i in range(1, n):
                p.drawLine(int(x_pos(i-1)), int(y_pos(arsiv[i-1])),
                           int(x_pos(i)),   int(y_pos(arsiv[i])))

        # X ekseni etiketleri (her 5. gün)
        p.setPen(QColor(P["txt4"]))
        f2 = QFont(); f2.setPointSize(8); p.setFont(f2)
        for i in range(0, n, max(1, n // 6)):
            if i < len(gunler):
                p.drawText(int(x_pos(i)) - 16, h - 6, gunler[i])

        # Legend
        f3 = QFont(); f3.setPointSize(9); p.setFont(f3)
        p.setPen(QColor(P["blue"]))
        p.drawText(pad_l, pad_t + 14, "● Zimmet")
        p.setPen(QColor(P["green"]))
        p.drawText(pad_l + 70, pad_t + 14, "● Arşivleme")
        p.end()


class NavButon(QPushButton):
    """Sidebar nav butonu — sol kenar aktif göstergesi."""
    def __init__(self, ikon: str, metin: str, parent=None):
        super().__init__(parent)
        self.setCheckable(True)
        self.setMinimumHeight(46)
        self.setCursor(Qt.PointingHandCursor)
        self._ikon = ikon
        self._metin = metin

        lay = QHBoxLayout(self)
        lay.setContentsMargins(14, 0, 12, 0)
        lay.setSpacing(12)

        self._ikon_lbl = QLabel(ikon)
        self._ikon_lbl.setFixedSize(24, 24)
        self._ikon_lbl.setAlignment(Qt.AlignCenter)
        self._ikon_lbl.setStyleSheet("font-size: 17px; background: transparent;")

        self._metin_lbl = QLabel(metin)
        self._metin_lbl.setStyleSheet(
            f"font-size: 13px; font-weight: 500; background: transparent; color: {P['navy_text']};"
        )

        lay.addWidget(self._ikon_lbl)
        lay.addWidget(self._metin_lbl)
        lay.addStretch()

        self.setStyleSheet(f"""
            QPushButton {{
                background: transparent;
                border: none;
                border-radius: 10px;
                border-left: 3px solid transparent;
                text-align: left;
            }}
            QPushButton:hover {{
                background: rgba(255,255,255,0.07);
                border-left: 3px solid rgba(59,130,246,0.4);
            }}
            QPushButton:checked {{
                background: rgba(59,130,246,0.18);
                border-left: 3px solid {P['navy_active']};
            }}
        """)

    def setChecked(self, checked: bool):
        super().setChecked(checked)
        if checked:
            self._metin_lbl.setStyleSheet(
                f"font-size: 13px; font-weight: 700; "
                f"background: transparent; color: white;"
            )
            self._ikon_lbl.setStyleSheet(
                "font-size: 17px; background: transparent;"
            )
        else:
            self._metin_lbl.setStyleSheet(
                f"font-size: 13px; font-weight: 500; "
                f"background: transparent; color: {P['navy_text']};"
            )
            self._ikon_lbl.setStyleSheet(
                "font-size: 17px; background: transparent;"
            )


# ─────────────────────────────────────────────────────────────
# MESAJLAŞMA WİDGETLARI
# ─────────────────────────────────────────────────────────────

class MesajBalonu(QFrame):
    """Tek bir mesaj balonu — okundu tik, dosya referansı destekli."""

    # Sinyal: dosya referansına tıklandı
    dosya_tikla = None  # Dışarıdan set edilecek callable

    def __init__(self, icerik: str, gonderen: str, zaman: str,
                 benim: bool, genel: bool = False,
                 karsisi_okudu: bool = False,
                 dosya_ref_id: int = 0, dosya_ref_no: str = "",
                 dosya_git_callback=None,
                 parent=None):
        super().__init__(parent)

        if genel:
            bg, fg, brd = P["amber_bg"], P["amber_t"], "#FDE68A"
        elif benim:
            bg, fg, brd = P["blue"], "white", P["blue2"]
        else:
            bg, fg, brd = P["surface"], P["txt"], P["border"]

        radius_extra = (
            "border-bottom-right-radius: 4px;"
            if benim else "border-bottom-left-radius: 4px;"
        )
        self.setStyleSheet(f"""
            QFrame {{
                background: {bg};
                border: 1px solid {brd};
                border-radius: 14px;
                {radius_extra}
            }}
        """)
        self.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Minimum)

        lay = QVBoxLayout(self)
        lay.setContentsMargins(14, 10, 14, 8)
        lay.setSpacing(4)

        # ── Üst: gönderen + zaman ─────────────────────────────
        ust = QHBoxLayout()
        ust.setSpacing(6)

        if genel:
            duy = QLabel("📢 DUYURU")
            duy.setStyleSheet(
                f"font-size: 10px; font-weight: 700; color: {P['amber_t']};"
            )
            ust.addWidget(duy)
            ust.addSpacing(4)

        g_lbl = QLabel(gonderen if not benim else "Siz")
        g_lbl.setStyleSheet(
            f"font-size: 11px; font-weight: 700; "
            f"color: {'rgba(255,255,255,0.8)' if benim else P['txt3']};"
        )
        z_lbl = QLabel(zaman[11:16] if len(zaman) > 11 else zaman)
        z_lbl.setStyleSheet(
            f"font-size: 10px; "
            f"color: {'rgba(255,255,255,0.55)' if benim else P['txt4']};"
        )
        ust.addWidget(g_lbl)
        ust.addStretch()
        ust.addWidget(z_lbl)
        lay.addLayout(ust)

        # ── Dosya referans kutusu (varsa) ─────────────────────
        if dosya_ref_id and dosya_ref_no:
            ref_btn = QPushButton(f"📎  {dosya_ref_no}  →")
            ref_btn.setStyleSheet(f"""
                QPushButton {{
                    background: {'rgba(255,255,255,0.15)' if benim else P['blue_bg']};
                    color: {'white' if benim else P['blue_t']};
                    border: 1px solid {'rgba(255,255,255,0.25)' if benim else '#BFDBFE'};
                    border-radius: 8px;
                    padding: 5px 12px;
                    font-size: 12px;
                    font-weight: 600;
                    text-align: left;
                }}
                QPushButton:hover {{
                    background: {'rgba(255,255,255,0.22)' if benim else '#DBEAFE'};
                }}
            """)
            if dosya_git_callback:
                ref_btn.clicked.connect(
                    lambda _, fid=dosya_ref_id: dosya_git_callback(fid)
                )
            lay.addWidget(ref_btn)

        # ── İçerik ───────────────────────────────────────────
        ic_lbl = QLabel(icerik)
        ic_lbl.setStyleSheet(
            f"font-size: 13px; color: {fg}; line-height: 1.4;"
        )
        ic_lbl.setWordWrap(True)
        lay.addWidget(ic_lbl)

        # ── Alt: okundu tikleri (sadece benim mesajlarım) ────
        if benim and not genel:
            alt = QHBoxLayout()
            alt.setContentsMargins(0, 0, 0, 0)
            alt.addStretch()
            if karsisi_okudu:
                tik = QLabel("✓✓")  # çift tik = okundu
                tik.setStyleSheet(
                    "font-size: 11px; font-weight: 700; "
                    "color: rgba(255,255,255,0.9);"
                )
                tik.setToolTip("Okundu")
            else:
                tik = QLabel("✓")   # tek tik = iletildi
                tik.setStyleSheet(
                    "font-size: 11px; color: rgba(255,255,255,0.5);"
                )
                tik.setToolTip("İletildi")
            alt.addWidget(tik)
            lay.addLayout(alt)


class KonusmaListeItem(QFrame):
    """Sol paneldeki konuşma satırı — tıklanabilir, okundu rozeti, online nokta."""

    tiklandi = Signal()

    def __init__(self, isim: str = "", son_mesaj: str = "", zaman: str = "",
                 okunmamis: int = 0, secili: bool = False,
                 online: bool = False, parent=None):
        super().__init__(parent)
        self.setCursor(Qt.PointingHandCursor)
        self.setFixedHeight(64)
        self._guncelle_stil(secili)

        lay = QHBoxLayout(self)
        lay.setContentsMargins(10, 0, 12, 0)
        lay.setSpacing(10)

        # Avatar dairesi
        self._avatar = QLabel()
        self._avatar.setFixedSize(40, 40)
        self._avatar.setAlignment(Qt.AlignCenter)
        self._avatar.setAttribute(Qt.WA_TransparentForMouseEvents)
        lay.addWidget(self._avatar)

        # Orta: isim + son mesaj
        ic = QVBoxLayout()
        ic.setSpacing(3)
        ic.setContentsMargins(0, 0, 0, 0)

        self._isim_lbl = QLabel()
        self._isim_lbl.setAttribute(Qt.WA_TransparentForMouseEvents)
        self._isim_lbl.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Preferred)

        self._son_lbl = QLabel()
        self._son_lbl.setAttribute(Qt.WA_TransparentForMouseEvents)
        self._son_lbl.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Preferred)

        ic.addWidget(self._isim_lbl)
        ic.addWidget(self._son_lbl)
        lay.addLayout(ic, stretch=1)

        # Sağ: zaman + rozet
        sag = QVBoxLayout()
        sag.setSpacing(4)
        sag.setAlignment(Qt.AlignTop | Qt.AlignRight)
        sag.setContentsMargins(0, 8, 0, 0)

        self._zaman_lbl = QLabel()
        self._zaman_lbl.setAttribute(Qt.WA_TransparentForMouseEvents)
        self._zaman_lbl.setAlignment(Qt.AlignRight)
        sag.addWidget(self._zaman_lbl)

        self._rozet = QLabel()
        self._rozet.setAttribute(Qt.WA_TransparentForMouseEvents)
        self._rozet.setAlignment(Qt.AlignCenter)
        self._rozet.setFixedSize(20, 20)
        self._rozet.setStyleSheet(f"""
            background:{P['red']}; color:white;
            border-radius:10px; font-size:10px; font-weight:700;
        """)
        self._rozet.setVisible(False)
        sag.addWidget(self._rozet, alignment=Qt.AlignRight)
        lay.addLayout(sag)

        # İlk değerlerle güncelle
        self._guncelle_icerik(isim, son_mesaj, zaman, okunmamis, secili, online)

    def _guncelle_icerik(self, isim, son, zaman, unk, secili, online):
        # Avatar
        harf = next((c for c in isim if c.isalpha()), "?").upper() if isim else "?"
        self._avatar.setText(harf)
        self._avatar.setStyleSheet(f"""
            background:{P['blue_bg']}; color:{P['blue_t']};
            border-radius:20px; font-size:16px; font-weight:700;
            {'border:2px solid #22C55E;' if online else ''}
        """)
        # İsim
        self._isim_lbl.setText(isim if isim else "")
        self._isim_lbl.setStyleSheet(
            f"font-size:13px; font-weight:{'700' if unk else '500'}; "
            f"color:{P['txt']}; background:transparent;"
        )
        # Son mesaj
        kisa = (son[:38] + "…") if len(son) > 38 else son
        self._son_lbl.setText(kisa)
        self._son_lbl.setStyleSheet(
            f"font-size:11px; color:{P['txt4']}; background:transparent; "
            f"{'font-weight:600;' if unk else ''}"
        )
        # Zaman
        z_txt = zaman[11:16] if len(zaman) > 11 else (zaman[:5] if zaman else "")
        self._zaman_lbl.setText(z_txt)
        self._zaman_lbl.setStyleSheet(f"font-size:10px; color:{P['txt4']}; background:transparent;")
        # Rozet
        self._rozet.setText(str(unk) if unk < 100 else "99+")
        self._rozet.setVisible(unk > 0)
        # Stil
        self._guncelle_stil(secili)

    def mousePressEvent(self, event):
        if event.button() == Qt.LeftButton:
            self.tiklandi.emit()
        super().mousePressEvent(event)

    def _guncelle_stil(self, secili: bool):
        self.setStyleSheet(f"""
            QFrame {{
                background:{'#EFF6FF' if secili else 'transparent'};
                border-radius:10px;
                border:{'1px solid #BFDBFE' if secili else 'none'};
            }}
            QFrame:hover {{ background:{P['bg']}; border-radius:10px; }}
        """)


def _mesaj_sayfasi_olustur(kullanici: dict, stack_ref,
                             nav_btns_ref: list, badge_ref) -> tuple:
    """Chat v5 — hiç widget silinmiyor, sadece hide/show + içerik güncelleme."""
    from PySide6.QtCore import QObject, QEvent

    uid   = kullanici["id"]
    uname = kullanici["full_name"]
    urole = kullanici["role"]

    # ═══ ANA ÇERÇEVE ════════════════════════════════
    sayfa = QWidget()
    sayfa.setStyleSheet(f"background:{P['bg']};")
    root = QHBoxLayout(sayfa)
    root.setContentsMargins(0, 0, 0, 0)
    root.setSpacing(0)

    # ═══ SOL: Konuşma listesi ═══════════════════════
    sol_w = QWidget()
    sol_w.setMinimumWidth(240)
    sol_w.setMaximumWidth(300)
    sol_w.setStyleSheet(f"background:{P['surface']}; border-right:1px solid {P['border']};")
    sol_lay = QVBoxLayout(sol_w)
    sol_lay.setContentsMargins(0, 0, 0, 0)
    sol_lay.setSpacing(0)

    sol_hdr = QWidget()
    sol_hdr.setFixedHeight(64)
    sol_hdr.setStyleSheet(f"background:{P['surface']}; border-bottom:1px solid {P['border']};")
    sh_lay = QHBoxLayout(sol_hdr)
    sh_lay.setContentsMargins(16, 0, 12, 0)
    sh_lay.setSpacing(8)
    _baslik_lbl = QLabel("Mesajlar")
    _baslik_lbl.setStyleSheet(f"font-size:16px; font-weight:800; color:{P['txt']};")
    yeni_btn = QPushButton("✉  Yeni")
    yeni_btn.setFixedHeight(34)
    yeni_btn.setStyleSheet(f"""
        QPushButton {{
            background:{P['blue']}; color:white; border:none;
            border-radius:10px; font-size:12px; font-weight:700; padding:0 14px;
        }}
        QPushButton:hover {{ background:{P['blue2']}; }}
    """)
    sh_lay.addWidget(_baslik_lbl)
    sh_lay.addStretch()
    sh_lay.addWidget(yeni_btn)
    sol_lay.addWidget(sol_hdr)

    kl_scroll = QScrollArea()
    kl_scroll.setWidgetResizable(True)
    kl_scroll.setFrameShape(QFrame.NoFrame)
    kl_scroll.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
    kl_cont = QWidget()
    kl_cont.setStyleSheet("background:transparent;")
    kl_lay = QVBoxLayout(kl_cont)
    kl_lay.setContentsMargins(8, 8, 8, 8)
    kl_lay.setSpacing(2)
    kl_lay.setAlignment(Qt.AlignTop)
    kl_scroll.setWidget(kl_cont)
    sol_lay.addWidget(kl_scroll, stretch=1)
    root.addWidget(sol_w)

    # ═══ ORTA: Mesaj alanı ══════════════════════════
    orta_w = QWidget()
    orta_w.setStyleSheet(f"background:{P['bg']};")
    orta_lay = QVBoxLayout(orta_w)
    orta_lay.setContentsMargins(0, 0, 0, 0)
    orta_lay.setSpacing(0)

    orta_hdr = QWidget()
    orta_hdr.setFixedHeight(64)
    orta_hdr.setStyleSheet(f"background:{P['surface']}; border-bottom:1px solid {P['border']};")
    oh_lay = QHBoxLayout(orta_hdr)
    oh_lay.setContentsMargins(20, 0, 16, 0)
    oh_lay.setSpacing(10)
    konusan_av = QLabel("?")
    konusan_av.setFixedSize(36, 36)
    konusan_av.setAlignment(Qt.AlignCenter)
    konusan_av.setStyleSheet(f"background:{P['blue_bg']}; color:{P['blue_t']}; border-radius:18px; font-size:15px; font-weight:700;")
    konusan_av.setVisible(False)
    orta_txt = QVBoxLayout(); orta_txt.setSpacing(1)
    konusan_isim = QLabel("Bir konuşma seçin")
    konusan_isim.setStyleSheet(f"font-size:14px; font-weight:700; color:{P['txt']};")
    konusan_alt = QLabel("")
    konusan_alt.setStyleSheet(f"font-size:11px; color:{P['txt4']};")
    orta_txt.addWidget(konusan_isim); orta_txt.addWidget(konusan_alt)
    oh_lay.addWidget(konusan_av); oh_lay.addLayout(orta_txt); oh_lay.addStretch()

    # Seçim modu butonu
    sec_mod_btn = QPushButton("☑ Seç")
    sec_mod_btn.setFixedHeight(32)
    sec_mod_btn.setVisible(False)
    sec_mod_btn.setStyleSheet(f"""
        QPushButton {{
            background:{P['blue_bg']}; color:{P['blue_t']};
            border:1px solid #BFDBFE; border-radius:8px;
            font-size:12px; font-weight:600; padding:0 12px;
        }}
        QPushButton:hover {{ background:#DBEAFE; }}
    """)

    # Sohbeti sil butonu
    sohbet_sil_btn = QPushButton("🗑 Sohbeti Sil")
    sohbet_sil_btn.setFixedHeight(32)
    sohbet_sil_btn.setVisible(False)
    sohbet_sil_btn.setStyleSheet(f"""
        QPushButton {{
            background:{P['red_bg']}; color:{P['red']};
            border:1px solid #FECACA; border-radius:8px;
            font-size:12px; font-weight:600; padding:0 12px;
        }}
        QPushButton:hover {{ background:#FEE2E2; color:#991B1B; }}
    """)

    oh_lay.addWidget(sec_mod_btn)
    oh_lay.addWidget(sohbet_sil_btn)
    orta_lay.addWidget(orta_hdr)

    msg_scroll = QScrollArea()
    msg_scroll.setWidgetResizable(True)
    msg_scroll.setFrameShape(QFrame.NoFrame)
    msg_scroll.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
    msg_cont = QWidget()
    msg_cont.setStyleSheet(f"background:{P['bg']};")
    msg_lay = QVBoxLayout(msg_cont)
    msg_lay.setContentsMargins(20, 16, 20, 16)
    msg_lay.setSpacing(6)
    msg_lay.setAlignment(Qt.AlignTop)
    msg_scroll.setWidget(msg_cont)
    orta_lay.addWidget(msg_scroll, stretch=1)

    # Seçim modu bar — seçim modundayken görünür
    sec_bar = QWidget()
    sec_bar.setStyleSheet(f"""
        QWidget {{
            background:{P['blue_bg']};
            border-top:1px solid #BFDBFE;
        }}
    """)
    sec_bar.setFixedHeight(52)
    sec_bar.setVisible(False)
    sec_bar_lay = QHBoxLayout(sec_bar)
    sec_bar_lay.setContentsMargins(16, 0, 16, 0)
    sec_bar_lay.setSpacing(10)
    sec_secilen_lbl = QLabel("0 mesaj seçildi")
    sec_secilen_lbl.setStyleSheet(f"font-size:13px; font-weight:600; color:{P['blue_t']};")
    sec_tumunu_btn = QPushButton("Tümünü Seç")
    sec_tumunu_btn.setFixedHeight(34)
    sec_tumunu_btn.setStyleSheet(f"""
        QPushButton {{
            background:white; color:{P['blue_t']};
            border:1px solid #BFDBFE; border-radius:8px;
            font-size:12px; padding:0 12px;
        }}
        QPushButton:hover {{ background:#DBEAFE; }}
    """)
    sec_sil_btn = QPushButton("🗑 Seçilenleri Sil")
    sec_sil_btn.setFixedHeight(34)
    sec_sil_btn.setStyleSheet(f"""
        QPushButton {{
            background:{P['red']}; color:white;
            border:none; border-radius:8px;
            font-size:12px; font-weight:700; padding:0 16px;
        }}
        QPushButton:hover {{ background:#DC2626; }}
        QPushButton:disabled {{ background:{P['border']}; color:{P['txt4']}; }}
    """)
    sec_sil_btn.setEnabled(False)
    sec_iptal_btn = QPushButton("İptal")
    sec_iptal_btn.setFixedHeight(34)
    sec_iptal_btn.setObjectName("ghost")
    sec_bar_lay.addWidget(sec_secilen_lbl)
    sec_bar_lay.addStretch()
    sec_bar_lay.addWidget(sec_tumunu_btn)
    sec_bar_lay.addWidget(sec_sil_btn)
    sec_bar_lay.addWidget(sec_iptal_btn)
    orta_lay.addWidget(sec_bar)

    yaz_w = QFrame()
    yaz_w.setStyleSheet(f"QFrame {{ background:{P['surface']}; border-top:1.5px solid {P['border']}; }}")
    yaz_lay_h = QHBoxLayout(yaz_w)
    yaz_lay_h.setContentsMargins(16, 12, 16, 12)
    yaz_lay_h.setSpacing(10)
    dosya_btn = QPushButton("📎")
    dosya_btn.setFixedSize(40, 40)
    dosya_btn.setEnabled(False)
    dosya_btn.setStyleSheet(f"""
        QPushButton {{ background:{P['bg']}; color:{P['txt4']}; border:1.5px solid {P['border']}; border-radius:10px; font-size:16px; }}
        QPushButton:enabled:hover {{ background:{P['blue_bg']}; color:{P['blue']}; border-color:#BFDBFE; }}
    """)
    yaz_input = QTextEdit()
    yaz_input.setPlaceholderText("Mesaj yazın... (Ctrl+Enter)")
    yaz_input.setMaximumHeight(80); yaz_input.setMinimumHeight(48)
    yaz_input.setEnabled(False)
    yaz_input.setStyleSheet(f"""
        QTextEdit {{ background:{P['bg']}; border:1.5px solid {P['border']}; border-radius:14px; padding:12px 16px; font-size:13px; color:{P['txt']}; }}
        QTextEdit:enabled:focus {{ border-color:{P['blue']}; }}
    """)
    gonder_btn = QPushButton("Gönder")
    gonder_btn.setFixedHeight(48); gonder_btn.setMinimumWidth(90)
    gonder_btn.setEnabled(False)
    gonder_btn.setStyleSheet(f"""
        QPushButton {{ background:{P['blue']}; color:white; border:none; border-radius:14px; font-size:13px; font-weight:700; padding:0 20px; }}
        QPushButton:enabled:hover {{ background:{P['blue2']}; }}
        QPushButton:disabled {{ background:{P['border']}; color:{P['txt4']}; }}
    """)
    yaz_lay_h.addWidget(dosya_btn); yaz_lay_h.addWidget(yaz_input, stretch=1); yaz_lay_h.addWidget(gonder_btn)
    orta_lay.addWidget(yaz_w)
    root.addWidget(orta_w, stretch=1)

    # ═══ SAĞ: Online ════════════════════════════════
    sag_w = QWidget()
    sag_w.setFixedWidth(185)
    sag_w.setStyleSheet(f"background:{P['surface']}; border-left:1px solid {P['border']};")
    sag_main = QVBoxLayout(sag_w)
    sag_main.setContentsMargins(0, 0, 0, 0); sag_main.setSpacing(0)
    ol_hdr = QWidget(); ol_hdr.setFixedHeight(64)
    ol_hdr.setStyleSheet(f"background:{P['surface']}; border-bottom:1px solid {P['border']};")
    ol_h = QHBoxLayout(ol_hdr); ol_h.setContentsMargins(14, 0, 14, 0)
    ol_baslik = QLabel("🟢  Online")
    ol_baslik.setStyleSheet(f"font-size:13px; font-weight:700; color:{P['txt']};")
    ol_sayi = QLabel("0")
    ol_sayi.setStyleSheet(f"background:{P['green_bg']}; color:{P['green_t']}; border-radius:8px; padding:2px 8px; font-size:11px; font-weight:700;")
    ol_h.addWidget(ol_baslik); ol_h.addStretch(); ol_h.addWidget(ol_sayi)
    sag_main.addWidget(ol_hdr)
    ol_scroll = QScrollArea(); ol_scroll.setWidgetResizable(True); ol_scroll.setFrameShape(QFrame.NoFrame)
    ol_cont = QWidget(); ol_cont.setStyleSheet("background:transparent;")
    ol_lay = QVBoxLayout(ol_cont); ol_lay.setContentsMargins(10,10,10,10); ol_lay.setSpacing(4); ol_lay.setAlignment(Qt.AlignTop)
    ol_scroll.setWidget(ol_cont); sag_main.addWidget(ol_scroll, stretch=1)
    root.addWidget(sag_w)

    # ── STATE ────────────────────────────────────────
    state = {
        "secili_id":    None,
        "secili_isim":  "",
        "secili_genel": False,
        "sec_modu":     False,    # seçim modu aktif mi
        "secili_ids":   set(),    # seçilen mesaj id'leri
    }

    # ── BİLDİRİM ─────────────────────────────────────
    _bq = []; _ba = [False]
    def bildirim_goster(b, i):
        _bq.append((b,i))
        if not _ba[0]: _bc()
    def _bc():
        if not _bq: _ba[0]=False; return
        _ba[0]=True; b,i=_bq.pop(0)
        p=QFrame(sayfa)
        p.setStyleSheet(f"QFrame{{background:{P['navy']};border:1px solid {P['navy3']};border-radius:14px;}}")
        p.setFixedWidth(290)
        pl=QVBoxLayout(p); pl.setContentsMargins(16,12,16,12); pl.setSpacing(4)
        QLabel(f"🔔  {b}",p).setStyleSheet("color:white;font-size:12px;font-weight:700;")
        pb=QLabel((i[:55]+"…") if len(i)>55 else i,p)
        pb.setStyleSheet(f"color:{P['navy_text']};font-size:12px;"); pb.setWordWrap(True)
        pl.addWidget(pl.parentWidget().findChild(QLabel)); pl.addWidget(pb)
        p.adjustSize(); p.move(sayfa.width()-310, sayfa.height()-p.height()-16)
        p.show(); p.raise_()
        QTimer.singleShot(3500, lambda: (p.hide(), p.deleteLater(), QTimer.singleShot(300,_bc)))

    # ── ONLINE GÜNCELLE ──────────────────────────────
    # Sabit widget havuzu — hiç silinmez, sadece hide/show
    _ol_widgets = []

    def online_guncelle():
        online = []
        try: online = online_kullanici_bilgileri(dakika=3)
        except Exception: pass
        ids = {u["id"] for u in online}
        cnt = sum(1 for u in online if u["id"] != uid)
        ol_sayi.setText(str(cnt))
        filtre = [u for u in online if u["id"] != uid]

        # Gerektiğinde yeni widget ekle havuza
        while len(_ol_widgets) < len(filtre):
            row = QPushButton()
            row.setFlat(True)
            row.setStyleSheet(f"""
                QPushButton {{
                    background:transparent; border:none;
                    border-radius:8px; padding:6px 8px; text-align:left;
                    font-size:12px; color:{P['txt2']};
                }}
                QPushButton:hover {{ background:#F1F5F9; }}
            """)
            ol_lay.addWidget(row)
            _ol_widgets.append(row)

        # Havuzdan güncelle
        for i, btn in enumerate(_ol_widgets):
            if i < len(filtre):
                u = filtre[i]
                btn.setText(f"🟢  {u['full_name'][:15]}\n     {ROL_ETIKET.get(u['role'],u['role'])}")
                _i = u["id"]; _n = u["full_name"]
                try: btn.clicked.disconnect()
                except Exception: pass  # İlk bağlanmada normal
                btn.clicked.connect(lambda checked=False, i=_i, n=_n: konusmayi_sec(i, n, False))
                btn.setVisible(True)
            else:
                btn.setVisible(False)
        return ids

    # ── KONUŞMA LİSTESİ ──────────────────────────────
    # Havuz: sadece KonusmaListeItem — QLabel ASLA eklenmez
    _kl_widgets = []
    # Boş durum için AYRI sabit label (havuza karışmaz)
    _kl_bos = QLabel("Henüz mesaj yok.\n✉ Yeni'ye tıklayın.")
    _kl_bos.setAlignment(Qt.AlignCenter)
    _kl_bos.setStyleSheet(f"font-size:12px; color:{P['txt4']}; padding:40px 20px;")
    kl_lay.addWidget(_kl_bos)

    def _kl_havuz_al(idx):
        """idx. KonusmaListeItem'ı havuzdan al, yoksa oluştur."""
        while len(_kl_widgets) <= idx:
            item = KonusmaListeItem("", "", "", 0, False, False)
            kl_lay.addWidget(item)
            _kl_widgets.append(item)
        return _kl_widgets[idx]

    def konusma_listesini_yukle(online_ids=None):
        if online_ids is None:
            try: online_ids = {u["id"] for u in online_kullanici_bilgileri(dakika=3)}
            except Exception: online_ids = set()

        yazismalar = []
        duyurular  = []
        try:
            yazismalar = konusma_listesi_getir(uid)
            duyurular  = duyuru_listesi_getir(uid)
        except Exception:
            pass

        tum_okunmamis = 0
        idx = 0

        if duyurular:
            d = duyurular[0]
            unk = int(d.get("okunmamis") or 0)
            tum_okunmamis += unk
            w = _kl_havuz_al(idx)
            _kl_guncelle(w, "📢 Duyurular", str(d.get("son_mesaj") or ""),
                         str(d.get("son_zaman") or ""), unk,
                         state["secili_id"] == -1, False)
            try: w.tiklandi.disconnect()
            except Exception: pass
            w.tiklandi.connect(lambda: konusmayi_sec(-1, "Duyurular", True))
            w.setVisible(True)
            idx += 1

        for y in yazismalar:
            unk = int(y.get("okunmamis") or 0)
            tum_okunmamis += unk
            w = _kl_havuz_al(idx)
            _i = y["diger_id"]; _n = y["diger_isim"]
            _kl_guncelle(w, _n, str(y.get("son_mesaj") or ""),
                         str(y.get("son_zaman") or ""), unk,
                         state["secili_id"] == _i and not state["secili_genel"],
                         _i in online_ids)
            try: w.tiklandi.disconnect()
            except Exception: pass
            w.tiklandi.connect(lambda c=False, i=_i, n=_n: konusmayi_sec(i, n, False))
            w.setVisible(True)
            idx += 1

        # Kullanılmayan havuz widget'larını gizle
        for i in range(idx, len(_kl_widgets)):
            _kl_widgets[i].setVisible(False)

        # Boş durum — ayrı label, havuzdan bağımsız
        _kl_bos.setVisible(idx == 0)

        # Badge
        badge_ref["sayac"] = tum_okunmamis
        if badge_ref.get("widget"):
            try:
                bw = badge_ref["widget"]
                bw.setText(str(tum_okunmamis) if tum_okunmamis else "")
                bw.setVisible(tum_okunmamis > 0)
            except RuntimeError: pass

    def _kl_guncelle(w, isim, son, zaman, unk, secili, online):
        """KonusmaListeItem içeriğini güncelle — hata varsa logla."""
        try:
            w._guncelle_icerik(isim, son, zaman, unk, secili, online)
        except Exception:
            pass

    # ── KONUŞMA SEÇ ──────────────────────────────────
    def konusmayi_sec(diger_id, diger_isim, genel=False):
        state["secili_id"]    = diger_id
        state["secili_isim"]  = diger_isim
        state["secili_genel"] = genel
        state["sec_modu"]     = False
        state["secili_ids"]   = set()

        harf = next((c for c in diger_isim if c.isalpha()), "?").upper()
        konusan_av.setText(harf); konusan_av.setVisible(True)
        konusan_isim.setText(diger_isim)
        konusan_alt.setText("Duyurular" if genel else "Özel mesaj")
        yaz_input.setEnabled(not genel)
        gonder_btn.setEnabled(not genel)
        dosya_btn.setEnabled(not genel)

        # Header butonları — duyuruda silme yok
        sec_mod_btn.setVisible(not genel)
        sohbet_sil_btn.setVisible(not genel)
        sec_bar.setVisible(False)

        mesajlari_yukle(); konusma_listesini_yukle()

    # ── MESAJLARI YUKLE ──────────────────────────────
    # Sabit widget havuzu
    # Mesaj container referansı — her yüklemede yeni container

    _msg_container = [None]

    def mesajlari_yukle():
        # Eski container'ı temizle — msg_scroll içeriğini sıfırla
        old = _msg_container[0]
        if old is not None:
            try:
                old.setVisible(False)
                # msg_lay'den kaldır — layout'a yığılmayı önle
                msg_lay.removeWidget(old)
            except Exception:
                pass

        if state["secili_id"] is None:
            # Boş ekran göster
            _msg_bos_goster("💬  Bir konuşma seçin\nveya ✉ Yeni'ye tıklayın.")
            return

        msgs = []
        try:
            if state["secili_genel"]:
                tum = mesajlari_getir(uid)
                msgs = list(reversed([m for m in tum if m["genel"] == 1]))
            else:
                msgs = konusma_gecmisi(uid, state["secili_id"])
        except Exception:
            return

        if not msgs:
            _msg_bos_goster("Henüz mesaj yok.\nİlk mesajı siz gönderin!")
            return

        _msg_bos_gizle()

        # Yeni container oluştur
        cont = QWidget()
        cont.setStyleSheet("background:transparent;")
        cont_lay = QVBoxLayout(cont)
        cont_lay.setContentsMargins(0, 0, 0, 0)
        cont_lay.setSpacing(6)
        cont_lay.setAlignment(Qt.AlignTop)

        onceki_tarih = ""
        for m in msgs:
            tarih = m["olusturma"][:10]
            if tarih != onceki_tarih:
                onceki_tarih = tarih
                tl = QLabel(_tarih_fmt(tarih))
                tl.setAlignment(Qt.AlignCenter)
                tl.setStyleSheet(f"color:{P['txt4']}; font-size:11px; padding:4px;")
                cont_lay.addWidget(tl)

            benim = m["gonderen_id"] == uid
            balon = MesajBalonu(
                icerik=m["icerik"], gonderen=m["gonderen"],
                zaman=m["olusturma"], benim=benim,
                genel=state["secili_genel"],
                karsisi_okudu=bool(m.get("karsisi_okudu", 0)),
                dosya_ref_id=int(m.get("dosya_ref_id") or 0),
                dosya_ref_no=m.get("dosya_ref_no") or "",
                dosya_git_callback=dosya_git if not state["secili_genel"] else None,
            )

            # ── Hover üzerinde gösterilecek sil butonu ──
            class _MesajSatir(QWidget):
                """Hover'da sil ikonu gösteren mesaj satırı."""
                def __init__(self, benim_, balon_, mid_, ic_, sec_modu_, secili_ids_,
                             secili_genel_, sil_onay_fn, sec_lbl, sec_btn):
                    super().__init__()
                    self.setStyleSheet("background:transparent;")
                    lay = QHBoxLayout(self)
                    lay.setContentsMargins(4, 2, 4, 2)
                    lay.setSpacing(8)

                    if secili_genel_:
                        lay.addSpacing(4); lay.addWidget(balon_); lay.addStretch()

                    elif sec_modu_:
                        # Seçim modu — checkbox + balon
                        cb = QCheckBox()
                        cb.setChecked(mid_ in secili_ids_)
                        cb.setStyleSheet("""
                            QCheckBox { background:transparent; }
                            QCheckBox::indicator {
                                width:20px; height:20px; border-radius:5px;
                            }
                            QCheckBox::indicator:unchecked {
                                border:2px solid #94A3B8; background:white; border-radius:5px;
                            }
                            QCheckBox::indicator:checked {
                                background:#2563EB; border:2px solid #1D4ED8;
                                border-radius:5px; image:none;
                            }
                        """)
                        def _tog(checked, mid=mid_, sl=sec_lbl, sb=sec_btn):
                            if checked: secili_ids_.add(mid)
                            else: secili_ids_.discard(mid)
                            n = len(secili_ids_)
                            sl.setText(f"{n} mesaj seçildi")
                            sb.setEnabled(n > 0)
                        cb.toggled.connect(_tog)
                        if benim_:
                            lay.addStretch()
                            lay.addWidget(balon_)
                            lay.addWidget(cb, alignment=Qt.AlignVCenter)
                        else:
                            lay.addWidget(cb, alignment=Qt.AlignVCenter)
                            lay.addWidget(balon_)
                            lay.addStretch()
                        self._sil_btn = None

                    else:
                        # Normal mod — sil butonu hover'da görünür
                        self._sil_btn = QPushButton("🗑")
                        self._sil_btn.setFixedSize(28, 28)
                        self._sil_btn.setToolTip("Mesajı sil")
                        self._sil_btn.setVisible(False)   # başta gizli
                        self._sil_btn.setStyleSheet(f"""
                            QPushButton {{
                                background:{P["red_bg"]}; color:{P["red"]};
                                border:1px solid #FECACA; border-radius:8px;
                                font-size:13px;
                            }}
                            QPushButton:hover {{ background:#FEE2E2; color:#991B1B; }}
                        """)
                        self._sil_btn.clicked.connect(
                            lambda _, mid=mid_, ic=ic_: sil_onay_fn(mid, ic)
                        )
                        if benim_:
                            lay.addStretch()
                            lay.addWidget(self._sil_btn, alignment=Qt.AlignVCenter)
                            lay.addWidget(balon_)
                        else:
                            lay.addWidget(balon_)
                            lay.addWidget(self._sil_btn, alignment=Qt.AlignVCenter)
                            lay.addStretch()

                def enterEvent(self, e):
                    if hasattr(self, '_sil_btn') and self._sil_btn:
                        self._sil_btn.setVisible(True)
                    super().enterEvent(e)

                def leaveEvent(self, e):
                    if hasattr(self, '_sil_btn') and self._sil_btn:
                        self._sil_btn.setVisible(False)
                    super().leaveEvent(e)

            wrap = _MesajSatir(
                benim_=benim, balon_=balon,
                mid_=m["id"], ic_=m["icerik"],
                sec_modu_=state["sec_modu"],
                secili_ids_=state["secili_ids"],
                secili_genel_=state["secili_genel"],
                sil_onay_fn=_sil_onay,
                sec_lbl=sec_secilen_lbl,
                sec_btn=sec_sil_btn,
            )
            cont_lay.addWidget(wrap)

            if not benim and not m.get("okundu"):
                try: mesaj_oku(m["id"], uid)
                except Exception: pass

        # Container'ı scroll içine ekle ve göster
        msg_lay.addWidget(cont)
        _msg_container[0] = cont

        QTimer.singleShot(80, lambda: msg_scroll.verticalScrollBar().setValue(
            msg_scroll.verticalScrollBar().maximum()
        ))

    _msg_bos_lbl = None

    def _msg_bos_goster(txt):
        nonlocal _msg_bos_lbl
        if _msg_bos_lbl is None:
            _msg_bos_lbl = QLabel()
            _msg_bos_lbl.setAlignment(Qt.AlignCenter)
            _msg_bos_lbl.setStyleSheet(
                f"font-size:14px; color:{P['txt4']}; padding:60px 20px;"
            )
            msg_lay.addWidget(_msg_bos_lbl)
        _msg_bos_lbl.setText(txt)
        _msg_bos_lbl.setVisible(True)

    def _msg_bos_gizle():
        if _msg_bos_lbl is not None:
            _msg_bos_lbl.setVisible(False)

    def _tarih_fmt(t):
        from datetime import date as _d, datetime as _dt
        try:
            d = _dt.strptime(t, "%Y-%m-%d").date()
            f = (_d.today() - d).days
            if f == 0: return "── Bugün ──"
            if f == 1: return "── Dün ──"
            if f < 7:  return f"── {d.strftime('%A')} ──"
            return f"── {d.strftime('%d %B %Y')} ──"
        except Exception: return t



    # ── SİL, DOSYA, GÖNDER ───────────────────────────
    # ── SEÇİM MODU ───────────────────────────────────
    def sec_modu_ac():
        state["sec_modu"] = True
        state["secili_ids"] = set()
        sec_bar.setVisible(True)
        sec_secilen_lbl.setText("0 mesaj seçildi")
        sec_sil_btn.setEnabled(False)
        yaz_w.setVisible(False)
        mesajlari_yukle()

    def sec_modu_kapat():
        state["sec_modu"] = False
        state["secili_ids"] = set()
        sec_bar.setVisible(False)
        yaz_w.setVisible(True)
        mesajlari_yukle()

    def sec_tumunu():
        """Tüm görünen mesajları seç."""
        if state["secili_id"] is None: return
        try:
            msgs = konusma_gecmisi(uid, state["secili_id"])
            state["secili_ids"] = {m["id"] for m in msgs}
        except Exception:
            pass
        sec_secilen_lbl.setText(f"{len(state['secili_ids'])} mesaj seçildi")
        sec_sil_btn.setEnabled(bool(state["secili_ids"]))
        mesajlari_yukle()

    def sec_toplu_sil():
        if not state["secili_ids"]: return
        d = QDialog(sayfa)
        d.setWindowTitle("Mesajları Sil")
        d.setFixedWidth(400)
        d.setStyleSheet(f"background:{P['surface']};")
        dl = QVBoxLayout(d)
        dl.setContentsMargins(28, 28, 28, 24)
        dl.setSpacing(16)
        ikon = QLabel("🗑")
        ikon.setAlignment(Qt.AlignCenter)
        ikon.setStyleSheet("font-size:36px;")
        baslik = QLabel(f"{len(state['secili_ids'])} mesaj silinecek")
        baslik.setAlignment(Qt.AlignCenter)
        baslik.setStyleSheet(f"font-size:16px; font-weight:700; color:{P['txt']};")
        aciklama = QLabel(
            "Seçilen mesajlar sizin görünümünüzden\n"
            "kalıcı olarak silinecek."
        )
        aciklama.setAlignment(Qt.AlignCenter)
        aciklama.setStyleSheet(f"font-size:13px; color:{P['txt3']}; line-height:1.5;")
        btn_lay = QHBoxLayout()
        btn_lay.setSpacing(10)
        iptal = QPushButton("İptal")
        iptal.setFixedHeight(42)
        iptal.setStyleSheet(f"""
            QPushButton {{
                background:{P['bg']}; color:{P['txt']};
                border:1.5px solid {P['border']}; border-radius:10px;
                font-size:14px; font-weight:600;
            }}
            QPushButton:hover {{ background:{P['surface']}; }}
        """)
        iptal.clicked.connect(d.reject)
        sil_onayla = QPushButton("Evet, Sil")
        sil_onayla.setFixedHeight(42)
        sil_onayla.setStyleSheet(f"""
            QPushButton {{
                background:{P['red']}; color:white;
                border:none; border-radius:10px;
                font-size:14px; font-weight:700;
            }}
            QPushButton:hover {{ background:#DC2626; }}
        """)
        sil_onayla.clicked.connect(d.accept)
        btn_lay.addWidget(iptal); btn_lay.addWidget(sil_onayla)
        dl.addWidget(ikon); dl.addWidget(baslik); dl.addWidget(aciklama); dl.addLayout(btn_lay)
        if d.exec() == QDialog.Accepted:
            try:
                toplu_mesaj_sil(list(state["secili_ids"]), uid, uname, urole)
            except Exception: pass
            sec_modu_kapat()
            konusma_listesini_yukle()

    def _sohbet_sil():
        if state["secili_id"] is None or state["secili_genel"]: return
        d = QDialog(sayfa)
        d.setWindowTitle("Sohbeti Sil")
        d.setFixedWidth(400)
        d.setStyleSheet(f"background:{P['surface']};")
        dl = QVBoxLayout(d)
        dl.setContentsMargins(28, 28, 28, 24)
        dl.setSpacing(16)
        ikon = QLabel("🗑")
        ikon.setAlignment(Qt.AlignCenter)
        ikon.setStyleSheet("font-size:36px;")
        isim_txt = state["secili_isim"]
        baslik = QLabel(f'"{isim_txt}" sohbeti silinecek')
        baslik.setAlignment(Qt.AlignCenter)
        baslik.setStyleSheet(f"font-size:15px; font-weight:700; color:{P['txt']};")
        baslik.setWordWrap(True)
        aciklama = QLabel(
            "Bu sohbetteki tüm mesajlar sizin görünümünüzden\n"
            "kalıcı olarak silinecek.\n"
            "Karşı taraf mesajları görmeye devam edebilir."
        )
        aciklama.setAlignment(Qt.AlignCenter)
        aciklama.setStyleSheet(f"font-size:13px; color:{P['txt3']}; line-height:1.5;")
        btn_lay = QHBoxLayout()
        btn_lay.setSpacing(10)
        iptal = QPushButton("İptal")
        iptal.setFixedHeight(42)
        iptal.setStyleSheet(f"""
            QPushButton {{
                background:{P['bg']}; color:{P['txt']};
                border:1.5px solid {P['border']}; border-radius:10px;
                font-size:14px; font-weight:600;
            }}
            QPushButton:hover {{ background:{P['surface']}; }}
        """)
        iptal.clicked.connect(d.reject)
        sil_onayla = QPushButton("Evet, Tüm Sohbeti Sil")
        sil_onayla.setFixedHeight(42)
        sil_onayla.setStyleSheet(f"""
            QPushButton {{
                background:{P['red']}; color:white;
                border:none; border-radius:10px;
                font-size:14px; font-weight:700;
            }}
            QPushButton:hover {{ background:#DC2626; }}
        """)
        sil_onayla.clicked.connect(d.accept)
        btn_lay.addWidget(iptal); btn_lay.addWidget(sil_onayla)
        dl.addWidget(ikon); dl.addWidget(baslik); dl.addWidget(aciklama); dl.addLayout(btn_lay)
        if d.exec() == QDialog.Accepted:
            try:
                konusma_sil(uid, state["secili_id"], uname, urole)
            except Exception: pass
            state["secili_id"] = None
            state["secili_isim"] = ""
            konusan_av.setVisible(False)
            konusan_isim.setText("Bir konuşma seçin")
            konusan_alt.setText("")
            sec_mod_btn.setVisible(False)
            sohbet_sil_btn.setVisible(False)
            yaz_input.setEnabled(False)
            gonder_btn.setEnabled(False)
            dosya_btn.setEnabled(False)
            mesajlari_yukle()
            konusma_listesini_yukle()

    def _sil_onay(mid, ic):
        d = QDialog(sayfa)
        d.setWindowTitle("Mesajı Sil")
        d.setFixedWidth(380)
        d.setStyleSheet(f"background:{P['surface']};")
        dl = QVBoxLayout(d)
        dl.setContentsMargins(28, 24, 28, 20)
        dl.setSpacing(14)
        ikon = QLabel("🗑")
        ikon.setAlignment(Qt.AlignCenter)
        ikon.setStyleSheet("font-size:32px;")
        baslik = QLabel("Mesajı Sil")
        baslik.setAlignment(Qt.AlignCenter)
        baslik.setStyleSheet(f"font-size:16px; font-weight:700; color:{P['txt']};")
        aciklama = QLabel(
            "Bu mesaj sizin görünümünüzden\nkalıcı olarak silinecek."
        )
        aciklama.setAlignment(Qt.AlignCenter)
        aciklama.setStyleSheet(f"font-size:13px; color:{P['txt3']};")
        btn_lay = QHBoxLayout()
        btn_lay.setSpacing(10)
        iptal = QPushButton("İptal")
        iptal.setFixedHeight(40)
        iptal.setStyleSheet(f"""
            QPushButton {{
                background:{P['bg']}; color:{P['txt']};
                border:1.5px solid {P['border']}; border-radius:10px;
                font-size:13px; font-weight:600;
            }}
            QPushButton:hover {{ background:{P['surface']}; }}
        """)
        iptal.clicked.connect(d.reject)
        sil_ok = QPushButton("Evet, Sil")
        sil_ok.setFixedHeight(40)
        sil_ok.setStyleSheet(f"""
            QPushButton {{
                background:{P['red']}; color:white;
                border:none; border-radius:10px;
                font-size:13px; font-weight:700;
            }}
            QPushButton:hover {{ background:#DC2626; }}
        """)
        sil_ok.clicked.connect(d.accept)
        btn_lay.addWidget(iptal); btn_lay.addWidget(sil_ok)
        dl.addWidget(ikon); dl.addWidget(baslik); dl.addWidget(aciklama); dl.addLayout(btn_lay)
        if d.exec() == QDialog.Accepted:
            try: mesaj_sil(mid, uid, uname, urole, ic)
            except Exception: pass
            mesajlari_yukle()
            konusma_listesini_yukle()
            # Eğer görüntülenen konuşmada mesaj kalmadıysa listeyi temizle
            try:
                kalan = konusma_gecmisi(uid, state["secili_id"])
                if not kalan and state["secili_id"] is not None:
                    state["secili_id"] = None
                    state["secili_isim"] = ""
                    konusma_listesini_yukle()
            except Exception:
                pass

    # Buton bağlantıları
    sec_mod_btn.clicked.connect(sec_modu_ac)
    sohbet_sil_btn.clicked.connect(_sohbet_sil)
    sec_iptal_btn.clicked.connect(sec_modu_kapat)
    sec_tumunu_btn.clicked.connect(sec_tumunu)
    sec_sil_btn.clicked.connect(sec_toplu_sil)

    def dosya_git(fid):
        for btn in nav_btns_ref:
            s=getattr(btn,'_stack_idx',None); btn.setChecked(s==1)
        stack_ref.setCurrentIndex(1)
        QMessageBox.information(sayfa,"Dosya Ref",f"Dosya ID:{fid}\nDosya Kayıtları'na geçildi.")

    def dosya_ref_sec():
        if state["secili_id"] is None or state["secili_genel"]: return
        d=QDialog(sayfa); d.setWindowTitle("Dosya Seç"); d.setFixedWidth(460)
        d.setStyleSheet(f"background:{P['surface']};")
        dl=QVBoxLayout(d); dl.setContentsMargins(20,20,20,20); dl.setSpacing(10)
        lbl=QLabel("DOSYA ARA"); lbl.setStyleSheet(f"font-size:11px;font-weight:700;color:{P['txt3']};")
        ai=QLineEdit(); ai.setPlaceholderText("Dosya no veya ID...")
        tbl=QTableWidget(); tbl.setColumnCount(3); tbl.setHorizontalHeaderLabels(["ID","Dosya No","Şefliği"])
        tbl.setEditTriggers(QTableWidget.NoEditTriggers); tbl.setSelectionBehavior(QTableWidget.SelectRows)
        tbl.horizontalHeader().setStretchLastSection(True); tbl.setMaximumHeight(180)
        ni=QLineEdit(); ni.setPlaceholderText("Not (isteğe bağlı)")
        def ara():
            ar=ai.text().strip().lower()
            if not ar: return
            from db import tum_files_ozet
            res=[r for r in tum_files_ozet() if ar in (r.get("orijinal_dosya_no") or "").lower() or ar==str(r.get("file_id",""))][:20]
            tbl.setRowCount(len(res))
            for ri,r in enumerate(res):
                tbl.setItem(ri,0,QTableWidgetItem(str(r["file_id"]))); tbl.setItem(ri,1,QTableWidgetItem(r["orijinal_dosya_no"])); tbl.setItem(ri,2,QTableWidgetItem(r.get("sefligi") or ""))
        ai.returnPressed.connect(ara); ab=QPushButton("Ara"); ab.clicked.connect(ara)
        al=QHBoxLayout(); al.addWidget(ai); al.addWidget(ab)
        bl=QHBoxLayout(); ip=QPushButton("İptal"); ip.setObjectName("ghost"); ip.clicked.connect(d.reject)
        gn=QPushButton("📎 Gönder"); gn.clicked.connect(d.accept); bl.addWidget(ip); bl.addWidget(gn)
        dl.addWidget(lbl); dl.addLayout(al); dl.addWidget(tbl); dl.addWidget(ni); dl.addLayout(bl)
        if d.exec()==QDialog.Accepted:
            row=tbl.currentRow()
            if row<0: return
            fid=int(tbl.item(row,0).text()); fno=tbl.item(row,1).text()
            nt=ni.text().strip() or f"Dosya: {fno}"
            tum=tum_kullanicilari_getir()
            af=next((k["full_name"] for k in tum if k["id"]==state["secili_id"]),state["secili_isim"])
            mesaj_gonder_dosya_ref(gonderen_id=uid,gonderen=uname,icerik=nt,alici_id=state["secili_id"],alici=af,dosya_ref_id=fid,dosya_ref_no=fno)
            mesajlari_yukle()
            konusma_listesini_yukle()
            # Eğer görüntülenen konuşmada mesaj kalmadıysa listeyi temizle
            try:
                kalan = konusma_gecmisi(uid, state["secili_id"])
                if not kalan and state["secili_id"] is not None:
                    state["secili_id"] = None
                    state["secili_isim"] = ""
                    konusma_listesini_yukle()
            except Exception:
                pass
    dosya_btn.clicked.connect(dosya_ref_sec)

    def gonder():
        ic=yaz_input.toPlainText().strip()
        if not ic or state["secili_id"] is None or state["secili_genel"]: return
        tum=tum_kullanicilari_getir()
        af=next((k["full_name"] for k in tum if k["id"]==state["secili_id"]),state["secili_isim"])
        mesaj_gonder(gonderen_id=uid,gonderen=uname,icerik=ic,alici_id=state["secili_id"],alici=af)
        yaz_input.clear(); mesajlari_yukle(); konusma_listesini_yukle()
    gonder_btn.clicked.connect(gonder)

    class _CEF(QObject):
        def eventFilter(self,obj,event):
            if obj is yaz_input and event.type()==QEvent.KeyPress:
                if event.key()==Qt.Key_Return and event.modifiers()&Qt.ControlModifier:
                    gonder(); return True
            return super().eventFilter(obj,event)
    _cef=_CEF(); yaz_input.installEventFilter(_cef); sayfa._cef=_cef

    # ── YENİ KONUŞMA ─────────────────────────────────
    def yeni_konusma():
        d=QDialog(sayfa); d.setWindowTitle("Yeni Mesaj"); d.setFixedWidth(420)
        d.setStyleSheet(f"background:{P['surface']};")
        dl=QVBoxLayout(d); dl.setContentsMargins(24,24,24,24); dl.setSpacing(12)
        la=QLabel("ALICI SEÇ"); la.setStyleSheet(f"font-size:11px;font-weight:700;color:{P['txt3']};letter-spacing:0.5px;")
        ks=QScrollArea(); ks.setWidgetResizable(True); ks.setFrameShape(QFrame.NoFrame); ks.setFixedHeight(210)
        ki=QWidget(); ki.setStyleSheet(f"background:{P['bg']};border-radius:10px;")
        kil=QVBoxLayout(ki); kil.setContentsMargins(8,8,8,8); kil.setSpacing(6)
        sk=[None]; bl2=[]
        try: oids={u["id"] for u in online_kullanici_bilgileri(dakika=3)}
        except: oids=set()
        for k in tum_kullanicilari_getir():
            if k["id"]==uid or not k["active"]: continue
            io=k["id"] in oids
            btn=QPushButton(f"  {'🟢' if io else '⚫'}  {k['full_name']} — {ROL_ETIKET.get(k['role'],k['role'])}")
            btn.setCheckable(True)
            btn.setStyleSheet(f"""
                QPushButton{{background:{P['surface']};color:{P['txt']};border:1.5px solid {P['border']};border-radius:10px;padding:10px 14px;text-align:left;font-size:13px;}}
                QPushButton:checked{{background:{P['blue_bg']};border-color:{P['blue']};color:{P['blue_t']};font-weight:700;}}
                QPushButton:hover:!checked{{background:{P['bg']};}}
            """)
            _kid=k["id"]; _kn=k["full_name"]
            def _tog(c,kid=_kid,kn=_kn,b=btn):
                if c:
                    sk[0]=(kid,kn)
                    for b2 in bl2:
                        if b2 is not b: b2.setChecked(False)
                elif sk[0] and sk[0][0]==kid: sk[0]=None
            btn.toggled.connect(_tog); kil.addWidget(btn); bl2.append(btn)
        kil.addStretch(); ks.setWidget(ki)
        lm=QLabel("MESAJ"); lm.setStyleSheet(la.styleSheet())
        mi=QTextEdit(); mi.setPlaceholderText("Mesajınızı yazın..."); mi.setFixedHeight(80)
        bl=QHBoxLayout(); bl.setSpacing(10)
        ip=QPushButton("İptal"); ip.setObjectName("ghost"); ip.setFixedHeight(44); ip.clicked.connect(d.reject)
        gnd=QPushButton("Gönder ↑"); gnd.setFixedHeight(44)
        gnd.setStyleSheet(f"QPushButton{{background:{P['blue']};color:white;border:none;border-radius:12px;font-size:14px;font-weight:700;padding:0 24px;}}QPushButton:hover{{background:{P['blue2']};}}")
        gnd.clicked.connect(d.accept); bl.addWidget(ip); bl.addWidget(gnd)
        dl.addWidget(la); dl.addWidget(ks); dl.addWidget(lm); dl.addWidget(mi); dl.addSpacing(8); dl.addLayout(bl)
        if d.exec()==QDialog.Accepted:
            if not sk[0]: return
            kid,kn=sk[0]; ic=mi.toPlainText().strip()
            if not ic: return
            mesaj_gonder(gonderen_id=uid,gonderen=uname,icerik=ic,alici_id=kid,alici=kn)
            konusma_listesini_yukle(); konusmayi_sec(kid,kn,False)
    yeni_btn.clicked.connect(yeni_konusma)

    # Admin duyuru
    if urole=="admin":
        db2=QPushButton("📢"); db2.setFixedSize(34,34); db2.setToolTip("Duyuru gönder")
        db2.setStyleSheet(f"QPushButton{{background:{P['amber_bg']};color:{P['amber_t']};border:none;border-radius:10px;font-size:14px;}}QPushButton:hover{{background:#FDE68A;}}")
        sh_lay.addWidget(db2)
        def duyuru():
            d=QDialog(sayfa); d.setWindowTitle("Duyuru Gönder"); d.setFixedWidth(440)
            d.setStyleSheet(f"background:{P['surface']};")
            dl=QVBoxLayout(d); dl.setContentsMargins(24,24,24,24); dl.setSpacing(10)
            kl2=QLabel("KONU"); kl2.setStyleSheet(f"font-size:11px;font-weight:700;color:{P['txt3']};")
            ki2=QLineEdit(); ki2.setPlaceholderText("Konu...")
            il=QLabel("İÇERİK"); il.setStyleSheet(kl2.styleSheet())
            ii=QTextEdit(); ii.setFixedHeight(100); ii.setPlaceholderText("Duyuru içeriği...")
            bl3=QHBoxLayout()
            ip2=QPushButton("İptal"); ip2.setObjectName("ghost"); ip2.clicked.connect(d.reject)
            gd=QPushButton("📢 Duyur"); gd.setObjectName("warning"); gd.clicked.connect(d.accept)
            bl3.addWidget(ip2); bl3.addWidget(gd)
            dl.addWidget(kl2); dl.addWidget(ki2); dl.addWidget(il); dl.addWidget(ii); dl.addLayout(bl3)
            if d.exec()==QDialog.Accepted:
                ic=ii.toPlainText().strip(); ko=ki2.text().strip()
                if ic:
                    mesaj_gonder(gonderen_id=uid,gonderen=uname,icerik=ic,konu=ko,genel=True)
                    konusma_listesini_yukle()
                    if state["secili_genel"]: mesajlari_yukle()
        db2.clicked.connect(duyuru)

    # ── OTOMATİK YENİLEME ────────────────────────────
    _oids=[set()]
    def otomatik_yenile():
        try: presence_guncelle(uid)
        except Exception: pass
        try:
            ids=online_guncelle(); _oids[0]=ids
            konusma_listesini_yukle(ids)
            if state["secili_id"] is not None: mesajlari_yukle()
        except Exception: pass
        try:
            yeni=mesajlari_getir(uid)
            for m in yeni:
                if m["id"] not in _oids[0] and m.get("yon")=="gelen" and not m.get("okundu"):
                    bildirim_goster(f"Yeni mesaj — {m['gonderen']}",m["icerik"])
        except Exception: pass

    timer=QTimer(); timer.setInterval(5000); timer.timeout.connect(otomatik_yenile); timer.start()
    sayfa._mesaj_timer=timer

    try:
        ids0=online_guncelle(); konusma_listesini_yukle(ids0)
    except Exception: pass

    return sayfa, konusma_listesini_yukle



def _tablo_olustur() -> QTableWidget:
    t = QTableWidget()
    t.setSelectionBehavior(QTableWidget.SelectRows)
    t.setEditTriggers(QTableWidget.NoEditTriggers)
    t.setAlternatingRowColors(False)
    t.setShowGrid(True)
    t.setWordWrap(False)
    # Responsive sütun modu: son sütun esner, diğerleri içeriğe göre
    t.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeToContents)
    t.horizontalHeader().setStretchLastSection(True)
    t.horizontalHeader().setDefaultAlignment(Qt.AlignLeft | Qt.AlignVCenter)
    t.verticalHeader().setVisible(False)
    t.verticalHeader().setDefaultSectionSize(42)
    t.setVerticalScrollMode(QAbstractItemView.ScrollPerPixel)
    t.setHorizontalScrollMode(QAbstractItemView.ScrollPerPixel)
    t.setFocusPolicy(Qt.NoFocus)
    t.setMinimumHeight(120)
    return t


def _bolum_baslik(metin: str, alt: str = "") -> QVBoxLayout:
    lay = QVBoxLayout()
    lay.setSpacing(2)
    lbl = QLabel(metin)
    lbl.setStyleSheet(
        f"font-size: 22px; font-weight: 800; color: {P['txt']}; letter-spacing: -0.5px;"
    )
    lay.addWidget(lbl)
    if alt:
        sub = QLabel(alt)
        sub.setStyleSheet(f"font-size: 12px; color: {P['txt4']}; font-weight: 400;")
        lay.addWidget(sub)
    return lay


def _sep() -> QFrame:
    f = QFrame()
    f.setFrameShape(QFrame.HLine)
    f.setFixedHeight(1)
    f.setStyleSheet(f"background: {P['border']}; border: none;")
    return f


# ─────────────────────────────────────────────────────────────
# LOGİN DİYALOĞU
# ─────────────────────────────────────────────────────────────
class LoginDialog(QDialog):
    # Motivasyon sözleri — her açılışta farklı
    SOZLER = [
        ("Düzen, başarının temelidir.", "Konfüçyüs"),
        ("Detaylara gösterilen özen, mükemmelliğin özüdür.", "Michelangelo"),
        ("İyi organize edilmiş bir sistem, en değerli varlıktır.", ""),
        ("Bir şeyi kaydetmek, onu asla kaybetmemektir.", ""),
        ("Düzenli çalışmak, zamanın en iyi yatırımıdır.", ""),
        ("Kayıt altına alınmayan bir iş, yapılmamış sayılır.", ""),
        ("Sistematik düşünce, büyük işlerin anahtarıdır.", ""),
        ("Verimliliğin sırrı, doğru sistemi kurmaktır.", ""),
        ("Küçük adımlar, büyük düzeni inşa eder.", ""),
        ("Düzen olmadan özgürlük olmaz.", "Edmund Burke"),
        ("Her şeyin bir yeri olmalı ve her şey yerinde olmalıdır.", ""),
        ("Organize olmak, başarının yarısıdır.", ""),
        ("İyi bir sistem, en yetenekli insandan daha güvenilirdir.", ""),
        ("Belge olmayan iş, olmayan iştir.", ""),
        ("Düzen bir erdem değil, her erdemin ön koşuludur.", ""),
        ("Takip edemediğin şeyi yönetemezsin.", ""),
        ("Küçük kayıtlar, büyük kayıpları önler.", ""),
        ("Disiplin, hedeflere ulaşmanın en kısa yoludur.", ""),
        ("Bir sistemin gücü, onu kullananların disiplinindedir.", ""),
        ("Doğru arşivleme, kurumun hafızasıdır.", ""),
    ]

    def __init__(self):
        super().__init__()
        self.kullanici = None
        self.setWindowTitle(APP_TITLE)
        self.setFixedSize(960, 600)
        self.setWindowFlags(Qt.Dialog | Qt.FramelessWindowHint)

        root = QHBoxLayout(self)
        root.setContentsMargins(0, 0, 0, 0)
        root.setSpacing(0)

        root.addWidget(self._sol_panel())
        root.addWidget(self._sag_panel())

        self._drag_pos = None

    # ── SOL PANEL ─────────────────────────────────────────────
    def _sol_panel(self) -> QWidget:
        import random
        soz, yazar = random.choice(self.SOZLER)

        saat = datetime.now().hour
        if saat < 6:
            sel_icon, sel_txt, sel_alt = "🌙", "İyi geceler", "Geç saatte de göreve devam."
        elif saat < 12:
            sel_icon, sel_txt, sel_alt = "☀️", "Günaydın", "Güzel bir gün olsun."
        elif saat < 14:
            sel_icon, sel_txt, sel_alt = "🌤", "İyi öğleler", "Verimli bir öğle geçiriyorsunuz."
        elif saat < 18:
            sel_icon, sel_txt, sel_alt = "🌇", "İyi günler", "Günün geri kalanı için başarılar."
        else:
            sel_icon, sel_txt, sel_alt = "🌆", "İyi akşamlar", "Bugünkü görevler tamamlanıyor mu?"

        w = QWidget()
        w.setFixedWidth(440)
        w.setStyleSheet(f"background: {P['navy']};")
        lay = QVBoxLayout(w)
        lay.setContentsMargins(52, 44, 52, 36)
        lay.setSpacing(0)

        # Logo + versiyon rozeti
        logo_satir = QHBoxLayout()
        logo_lbl = QLabel("🗂")
        logo_lbl.setFixedSize(48, 48)
        logo_lbl.setAlignment(Qt.AlignCenter)
        logo_lbl.setStyleSheet("""
            background: rgba(37,99,235,0.25);
            border: 1.5px solid rgba(37,99,235,0.5);
            border-radius: 13px; font-size: 22px;
        """)
        ver_badge = QLabel(APP_VERSIYON)
        ver_badge.setStyleSheet(f"""
            color: {P['navy_text']};
            background: rgba(255,255,255,0.07);
            border: 1px solid rgba(255,255,255,0.12);
            border-radius: 8px; padding: 3px 10px;
            font-size: 11px; font-weight: 600;
        """)
        logo_satir.addWidget(logo_lbl)
        logo_satir.addStretch()
        logo_satir.addWidget(ver_badge)
        lay.addLayout(logo_satir)
        lay.addSpacing(24)

        # Saate göre selamlama
        sel_satir = QHBoxLayout()
        sel_satir.setSpacing(10)
        sel_ikon_lbl = QLabel(sel_icon)
        sel_ikon_lbl.setFixedSize(34, 34)
        sel_ikon_lbl.setAlignment(Qt.AlignCenter)
        sel_ikon_lbl.setStyleSheet("""
            background: rgba(255,255,255,0.08);
            border-radius: 9px; font-size: 16px;
        """)
        sel_txt_lay = QVBoxLayout()
        sel_txt_lay.setSpacing(1)
        sel_baslik = QLabel(sel_txt)
        sel_baslik.setStyleSheet(
            "color: white; font-size: 20px; font-weight: 800; letter-spacing: -0.3px;"
        )
        sel_alt_lbl = QLabel(sel_alt)
        sel_alt_lbl.setStyleSheet(f"color: {P['navy_text']}; font-size: 11px;")
        sel_txt_lay.addWidget(sel_baslik)
        sel_txt_lay.addWidget(sel_alt_lbl)
        sel_satir.addWidget(sel_ikon_lbl)
        sel_satir.addLayout(sel_txt_lay)
        sel_satir.addStretch()
        lay.addLayout(sel_satir)
        lay.addSpacing(4)

        # Tarih
        tarih_lbl = QLabel(datetime.now().strftime("%d %B %Y, %A"))
        tarih_lbl.setStyleSheet(
            "color: rgba(148,163,184,0.55); font-size: 11px; padding-left: 44px;"
        )
        lay.addWidget(tarih_lbl)
        lay.addSpacing(20)

        # İnce çizgi
        sep = QFrame()
        sep.setFixedHeight(1)
        sep.setStyleSheet("background: rgba(255,255,255,0.08); border: none;")
        lay.addWidget(sep)
        lay.addSpacing(20)

        # Uygulama adı + kurum
        app_lbl = QLabel(APP_TITLE)
        app_lbl.setStyleSheet(
            "color: white; font-size: 18px; font-weight: 800; letter-spacing: -0.3px;"
        )
        lay.addWidget(app_lbl)
        lay.addSpacing(3)
        kurum_lbl = QLabel("T.C. İzmir Büyükşehir Belediyesi")
        kurum_lbl.setStyleSheet(f"color: {P['navy_text']}; font-size: 12px;")
        lay.addWidget(kurum_lbl)
        lay.addStretch()

        # Motivasyon söz kutusu — sol kenarda mavi şerit
        soz_frame = QFrame()
        soz_frame.setStyleSheet(f"""
            QFrame {{
                background: rgba(37,99,235,0.10);
                border: 1px solid rgba(37,99,235,0.22);
                border-left: 3px solid {P['blue']};
                border-radius: 12px;
            }}
        """)
        soz_lay = QVBoxLayout(soz_frame)
        soz_lay.setContentsMargins(18, 14, 18, 14)
        soz_lay.setSpacing(6)
        soz_lbl = QLabel(f"\u201c{soz}\u201d")
        soz_lbl.setStyleSheet(
            "color: #E2E8F0; font-size: 13px; font-style: italic; line-height: 1.5;"
        )
        soz_lbl.setWordWrap(True)
        soz_lay.addWidget(soz_lbl)
        if yazar:
            yazar_lbl = QLabel(f"— {yazar}")
            yazar_lbl.setStyleSheet(
                f"color: {P['blue']}; font-size: 11px; font-weight: 600;"
            )
            soz_lay.addWidget(yazar_lbl)
        lay.addWidget(soz_frame)
        lay.addSpacing(20)

        # Alt bilgi
        alt = QHBoxLayout()
        dev = QLabel(f"© 2026 {APP_SAHIP}")
        dev.setStyleSheet("color: rgba(148,163,184,0.45); font-size: 10px;")
        des = QLabel(f"📞 {DESTEK_TEL}")
        des.setStyleSheet("color: rgba(148,163,184,0.45); font-size: 10px;")
        alt.addWidget(dev)
        alt.addStretch()
        alt.addWidget(des)
        lay.addLayout(alt)

        return w

    # ── SAĞ PANEL ─────────────────────────────────────────────
    def _sag_panel(self) -> QWidget:
        w = QWidget()
        w.setStyleSheet(f"background: {P['surface']};")

        lay = QVBoxLayout(w)
        lay.setContentsMargins(0, 0, 0, 0)
        lay.setSpacing(0)

        # Kapatma butonu — sağ üst
        kapat_wrap = QHBoxLayout()
        kapat_wrap.setContentsMargins(0, 16, 20, 0)
        kapat_wrap.addStretch()
        kapat = QPushButton("✕")
        kapat.setFixedSize(30, 30)
        kapat.setStyleSheet(f"""
            QPushButton {{
                background: transparent;
                border: none;
                color: {P['txt4']};
                font-size: 15px;
                border-radius: 8px;
            }}
            QPushButton:hover {{
                background: {P['bg']};
                color: {P['txt']};
            }}
        """)
        kapat.clicked.connect(lambda: __import__('sys').exit(0))
        kapat_wrap.addWidget(kapat)
        lay.addLayout(kapat_wrap)

        # Form içeriği — dikey ortala
        form_wrap = QVBoxLayout()
        form_wrap.setContentsMargins(56, 0, 56, 0)
        form_wrap.setSpacing(0)

        # Hoş geldiniz
        hos = QLabel("Hoş Geldiniz")
        hos.setStyleSheet(
            f"font-size: 28px; font-weight: 800; color: {P['txt']}; letter-spacing: -0.8px;"
        )
        form_wrap.addWidget(hos)
        form_wrap.addSpacing(6)

        alt_hos = QLabel("Sisteme erişmek için giriş yapın.")
        alt_hos.setStyleSheet(f"font-size: 14px; color: {P['txt3']};")
        form_wrap.addWidget(alt_hos)
        form_wrap.addSpacing(36)

        # Kullanıcı adı
        self._ku_lbl = QLabel("KULLANICI ADI")
        self._ku_lbl.setStyleSheet(
            f"font-size: 11px; font-weight: 700; color: {P['txt3']}; letter-spacing: 0.6px;"
        )
        form_wrap.addWidget(self._ku_lbl)
        form_wrap.addSpacing(6)

        self.username_input = QLineEdit()
        self.username_input.setPlaceholderText("kullaniciadi")
        self.username_input.setFixedHeight(50)
        self.username_input.setStyleSheet(self._input_stil())
        form_wrap.addWidget(self.username_input)
        form_wrap.addSpacing(20)

        # Şifre
        self._sp_lbl = QLabel("ŞİFRE")
        self._sp_lbl.setStyleSheet(self._ku_lbl.styleSheet())
        form_wrap.addWidget(self._sp_lbl)
        form_wrap.addSpacing(6)

        self.password_input = QLineEdit()
        self.password_input.setEchoMode(QLineEdit.Password)
        self.password_input.setPlaceholderText("••••••••")
        self.password_input.setFixedHeight(50)
        self.password_input.setStyleSheet(self._input_stil())
        self.password_input.returnPressed.connect(self.login)
        form_wrap.addWidget(self.password_input)
        form_wrap.addSpacing(28)

        # Giriş butonu
        self.login_btn = QPushButton("Giriş Yap  →")
        self.login_btn.setFixedHeight(52)
        self.login_btn.setStyleSheet(f"""
            QPushButton {{
                background: {P['navy']};
                color: white;
                border: none;
                border-radius: 13px;
                font-size: 15px;
                font-weight: 700;
                letter-spacing: 0.3px;
            }}
            QPushButton:hover {{
                background: {P['navy3']};
            }}
            QPushButton:pressed {{
                background: {P['navy2']};
            }}
        """)
        self.login_btn.clicked.connect(self.login)
        form_wrap.addWidget(self.login_btn)
        form_wrap.addSpacing(16)

        # Hata mesajı
        self.hata_lbl = QLabel("")
        self.hata_lbl.setAlignment(Qt.AlignCenter)
        self.hata_lbl.setFixedHeight(0)
        self.hata_lbl.setStyleSheet(f"""
            background: {P['red_bg']};
            color: {P['red_t']};
            border: 1px solid #FECACA;
            border-radius: 10px;
            padding: 10px;
            font-size: 13px;
            font-weight: 500;
        """)
        self.hata_lbl.setVisible(False)
        form_wrap.addWidget(self.hata_lbl)

        lay.addStretch()
        lay.addLayout(form_wrap)
        lay.addStretch()

        # Destek bilgisi
        destek_lbl = QLabel(f"📞 Destek için {DESTEK_TEL}")
        destek_lbl.setAlignment(Qt.AlignCenter)
        destek_lbl.setStyleSheet(
            f"font-size: 12px; color: {P['txt4']}; padding: 0 0 20px 0;"
        )
        lay.addWidget(destek_lbl)

        return w

    def _input_stil(self) -> str:
        return f"""
            QLineEdit {{
                background: {P['bg']};
                border: 1.5px solid {P['border']};
                border-radius: 13px;
                padding: 0 18px;
                font-size: 14px;
                color: {P['txt']};
            }}
            QLineEdit:focus {{
                background: {P['surface']};
                border: 1.5px solid {P['navy']};
            }}
            QLineEdit:hover {{
                border-color: {P['border2']};
            }}
        """

    def mousePressEvent(self, e):
        if e.button() == Qt.LeftButton:
            self._drag_pos = e.globalPosition().toPoint() - self.frameGeometry().topLeft()

    def mouseMoveEvent(self, e):
        if self._drag_pos and e.buttons() == Qt.LeftButton:
            self.move(e.globalPosition().toPoint() - self._drag_pos)

    def mouseReleaseEvent(self, e):
        self._drag_pos = None

    def login(self):
        self.hata_lbl.setVisible(False)
        self.hata_lbl.setFixedHeight(0)
        self.login_btn.setText("Doğrulanıyor...")
        self.login_btn.setEnabled(False)
        QApplication.processEvents()

        k = giris_yap(
            self.username_input.text().strip(),
            self.password_input.text().strip(),
        )

        self.login_btn.setText("Giriş Yap  →")
        self.login_btn.setEnabled(True)

        if k:
            self.kullanici = k
            self.accept()
        else:
            self.hata_lbl.setText("⚠  Kullanıcı adı veya şifreniz hatalı. Lütfen tekrar deneyin.")
            self.hata_lbl.setFixedHeight(46)
            self.hata_lbl.setVisible(True)
            self.password_input.clear()
            self.password_input.setFocus()
            # Şifre kutusu kırmızı border — görsel uyarı
            self.password_input.setStyleSheet(
                self.password_input.styleSheet() +
                "border:1.5px solid #EF4444 !important;"
            )
            QTimer.singleShot(2000, lambda: self.password_input.setStyleSheet(""))


# ─────────────────────────────────────────────────────────────
# FORM DİYALOGLARI (ortak şablon)
# ─────────────────────────────────────────────────────────────
FORM_STIL = f"""
QDialog {{ background: {P['surface']}; }}
QLabel {{ color: {P['txt2']}; }}
"""

def _dialog_kur(dialog: QDialog, genislik: int, ikon: str, baslik: str, alt: str = ""):
    dialog.setStyleSheet(FORM_STIL)
    dialog.setFixedWidth(genislik)
    lay = QVBoxLayout(dialog)
    lay.setContentsMargins(28, 28, 28, 28)
    lay.setSpacing(0)

    # Başlık bloğu
    ust = QHBoxLayout()
    ust.setSpacing(14)

    ikon_f = QFrame()
    ikon_f.setFixedSize(44, 44)
    ikon_f.setStyleSheet(f"""
        background: {P['blue_bg']};
        border-radius: 12px;
    """)
    ikon_lay = QHBoxLayout(ikon_f)
    ikon_lay.setContentsMargins(0, 0, 0, 0)
    ikon_ic = QLabel(ikon)
    ikon_ic.setAlignment(Qt.AlignCenter)
    ikon_ic.setStyleSheet("font-size: 20px; background: transparent;")
    ikon_lay.addWidget(ikon_ic)

    txt_lay = QVBoxLayout()
    txt_lay.setSpacing(2)
    b = QLabel(baslik)
    b.setStyleSheet(f"font-size: 17px; font-weight: 700; color: {P['txt']};")
    txt_lay.addWidget(b)
    if alt:
        a = QLabel(alt)
        a.setStyleSheet(f"font-size: 12px; color: {P['txt3']};")
        txt_lay.addWidget(a)

    ust.addWidget(ikon_f)
    ust.addLayout(txt_lay)
    ust.addStretch()
    lay.addLayout(ust)
    lay.addSpacing(20)
    lay.addWidget(_sep())
    lay.addSpacing(20)

    return lay


def _form_satir(lay: QVBoxLayout, etiket: str, widget: QWidget):
    lbl = QLabel(etiket)
    lbl.setStyleSheet(
        f"font-size: 11px; font-weight: 700; color: {P['txt3']}; letter-spacing: 0.5px;"
    )
    lay.addWidget(lbl)
    lay.addSpacing(4)
    lay.addWidget(widget)
    lay.addSpacing(14)


def _btn_satir(lay: QVBoxLayout, iptal_slot, tamam_btn: QPushButton):
    lay.addSpacing(6)
    lay.addWidget(_sep())
    lay.addSpacing(16)
    satir = QHBoxLayout()
    iptal = QPushButton("İptal")
    iptal.setObjectName("ghost")
    iptal.setFixedHeight(42)
    iptal.clicked.connect(iptal_slot)
    tamam_btn.setFixedHeight(42)
    satir.addWidget(iptal)
    satir.addWidget(tamam_btn)
    lay.addLayout(satir)


class YeniDosyaDialog(QDialog):
    """
    Yeni dosya kaydı — elle giriş:
    Dosya No (sayısal) → İlçe (dropdown) → Şefliği → Ada → Parsel
    → Teslim Alan → Arşiv Görevlisi → Tarih → Not
    """
    def __init__(self, kullanici: dict):
        super().__init__()
        self.kullanici = kullanici
        self._secili_user_id = None
        lay = _dialog_kur(self, 520, "📁", "Yeni Dosya Kaydı", "Tüm alanları eksiksiz doldurun")

        # ── Dosya No — sadece sayısal ──
        self.dosya_no = QLineEdit()
        self.dosya_no.setPlaceholderText("Örn: 1234")
        self.dosya_no.setValidator(None)  # Sayısal doğrulama kaydet'te
        self.dosya_no.setStyleSheet(f"""
            QLineEdit {{
                font-size: 18px; font-weight: 700;
                color: {P['blue']}; letter-spacing: 1px;
            }}
        """)

        # ── İlçe — sabit dropdown ──
        self.ilce_cb = QComboBox()
        self.ilce_cb.setFixedHeight(42)
        self.ilce_cb.addItem("— İlçe seçin —", "")
        for ilce in ILCE_LISTESI:
            self.ilce_cb.addItem(ilce, ilce)

        # ── Müdürlük — sabit dropdown ──
        self.sefligi = QComboBox()
        self.sefligi.setFixedHeight(42)
        self.sefligi.addItem("— Müdürlük seçin —", "")
        for mud in MUDÜRLUK_LISTESI:
            self.sefligi.addItem(mud, mud)

        # ── Ada / Parsel yan yana ──
        self.ada    = QLineEdit(); self.ada.setPlaceholderText("Ada no")
        self.parsel = QLineEdit(); self.parsel.setPlaceholderText("Parsel no")
        ada_row = QHBoxLayout(); ada_row.setSpacing(10)
        ada_ic = QVBoxLayout(); ada_ic.setSpacing(3)
        ada_ic.addWidget(QLabel("Ada", styleSheet=f"font-size:11px;color:{P['txt4']};font-weight:600;"))
        ada_ic.addWidget(self.ada)
        pars_ic = QVBoxLayout(); pars_ic.setSpacing(3)
        pars_ic.addWidget(QLabel("Parsel", styleSheet=f"font-size:11px;color:{P['txt4']};font-weight:600;"))
        pars_ic.addWidget(self.parsel)
        ada_row.addLayout(ada_ic); ada_row.addLayout(pars_ic)
        ada_w = QWidget(); ada_w.setLayout(ada_row); ada_w.setStyleSheet("background:transparent;")

        # ── Teslim Alan — dropdown + serbest ──
        self._teslim_cb = QComboBox()
        self._teslim_cb.setEditable(True)
        self._teslim_cb.setInsertPolicy(QComboBox.NoInsert)
        self._teslim_cb.setFixedHeight(42)
        self._teslim_cb.lineEdit().setPlaceholderText("Kullanıcı seç veya isim yaz...")
        self._teslim_cb.addItem("— Kullanıcı seç —", None)
        for k in tum_kullanicilari_getir():
            if k["active"] and k["role"] != "admin":
                self._teslim_cb.addItem(
                    f"{k['full_name']}  ({ROL_ETIKET.get(k['role'],k['role'])})", k["id"])
        self._teslim_cb.addItem("── Serbest isim gir ──", -1)
        self._teslim_cb.currentIndexChanged.connect(
            lambda i: setattr(self, '_secili_user_id',
                self._teslim_cb.currentData()
                if self._teslim_cb.currentData() not in (None,-1) else None))

        # ── Arşiv Görevlisi ──
        self.arsiv_gor = QLineEdit(kullanici["full_name"])

        # ── Tarih ──
        self.teslim_tarihi = QDateEdit(QDate.currentDate())
        self.teslim_tarihi.setCalendarPopup(True)
        self.teslim_tarihi.setDisplayFormat("dd.MM.yyyy")
        self.teslim_tarihi.setFixedHeight(42)

        # ── Not ──
        self.notlar = QTextEdit()
        self.notlar.setFixedHeight(65)
        self.notlar.setPlaceholderText("İsteğe bağlı açıklama...")

        # ── Form ──
        for etiket, widget in [
            ("DOSYA NO *",              self.dosya_no),
            ("İLÇE *",                  self.ilce_cb),
            ("MÜDÜRLÜK *",               self.sefligi),
            ("ADA / PARSEL",            ada_w),
            ("TESLİM ALAN PERSONEL *",  self._teslim_cb),
            ("ARŞİV GÖREVLİSİ *",      self.arsiv_gor),
            ("TESLİM TARİHİ",           self.teslim_tarihi),
            ("NOT",                     self.notlar),
        ]:
            _form_satir(lay, etiket, widget)

        kaydet = QPushButton("💾  Kaydet & Zimmetle")
        kaydet.setFixedHeight(48)
        _btn_satir(lay, self.reject, kaydet)
        kaydet.clicked.connect(self.kaydet)

    def kaydet(self):
        try:
            d      = self.dosya_no.text().strip()
            ilce   = self.ilce_cb.currentData() or ""
            s      = (self.sefligi.currentData()
                       if hasattr(self.sefligi, 'currentData')
                       else self.sefligi.text().strip()) or ""
            ada    = self.ada.text().strip()
            parsel = self.parsel.text().strip()
            a      = self.arsiv_gor.text().strip()
            t      = self._teslim_cb.currentText().split("  (")[0].strip()
            if t in ("— Kullanıcı seç —","── Serbest isim gir ──"):
                t = self._teslim_cb.lineEdit().text().strip()

            # Doğrulama
            if not d:
                raise ValueError("Dosya numarası boş olamaz.")
            if not d.isdigit():
                raise ValueError(f"Dosya numarası yalnızca rakam olmalı.\nGirilen: '{d}'")
            if not ilce:
                raise ValueError("İlçe seçmediniz.")
            if not t:
                raise ValueError("Teslim alan personeli seçin veya yazın.")
            if not a:
                raise ValueError("Arşiv görevlisi boş olamaz.")

            # Dosya no = sadece sayı, ilçe ayrı kolona gidecek
            fid = dosya_ve_hareket_ekle(
                orijinal_dosya_no=d, sefligi=s,
                teslim_alan_personel=t, veren_arsiv_gorevlisi=a,
                teslim_tarihi=self.teslim_tarihi.date().toString("yyyy-MM-dd"),
                notlar=self.notlar.toPlainText(),
                ada=ada, parsel=parsel, ilce=ilce,
            )
            if self._secili_user_id:
                movement_user_id_guncelle(fid, self._secili_user_id)

            # Otomatik Excel yedeği
            otomatik_excel_yedek()

            QMessageBox.information(
                self, "Kaydedildi ✓",
                f"Dosya No: {d}\n"
                f"İlçe: {ilce}\n"
                f"Teslim Alan: {t}\n\n"
                f"Dosya kaydedildi ve zimmetlendi."
            )
            self.accept()
        except Exception as e:
            QMessageBox.critical(self, "Hata", str(e))


class ZimmetEkleDialog(QDialog):
    def __init__(self, file_id: int, dosya_no: str, kullanici: dict):
        super().__init__()
        self.file_id = file_id
        self._secili_user_id = None
        lay = _dialog_kur(self, 500, "➕", "Zimmet Ekle", dosya_no)

        # Teslim alan — kullanıcı dropdown + serbest giriş
        self._teslim_cb = QComboBox()
        self._teslim_cb.setEditable(True)
        self._teslim_cb.setPlaceholderText("Kullanıcı seç veya yaz...")
        self._teslim_cb.addItem("— Kullanıcı seç —", None)
        for k in tum_kullanicilari_getir():
            if k["active"] and k["role"] != "admin":
                self._teslim_cb.addItem(
                    f"{k['full_name']}  ({ROL_ETIKET.get(k['role'],k['role'])})", k["id"]
                )
        self._teslim_cb.addItem("── Serbest giriş ──", -1)
        self._teslim_cb.currentIndexChanged.connect(self._on_secim)

        self.arsiv_gor = QLineEdit(kullanici["full_name"])
        self.notlar    = QTextEdit(); self.notlar.setFixedHeight(64)
        self.notlar.setPlaceholderText("Not...")
        self.teslim_tarihi = QDateEdit(QDate.currentDate())
        self.teslim_tarihi.setCalendarPopup(True)
        self.teslim_tarihi.setDisplayFormat("dd.MM.yyyy")

        for etiket, widget in [
            ("TESLİM ALAN *",      self._teslim_cb),
            ("ARŞİV GÖREVLİSİ",   self.arsiv_gor),
            ("TESLİM TARİHİ",      self.teslim_tarihi),
            ("NOTLAR",             self.notlar),
        ]:
            _form_satir(lay, etiket, widget)

        ekle = QPushButton("➕  Zimmet Ekle")
        _btn_satir(lay, self.reject, ekle)
        ekle.clicked.connect(self.kaydet)

    def _on_secim(self, idx):
        uid = self._teslim_cb.currentData()
        self._secili_user_id = uid if (uid and uid != -1) else None

    def kaydet(self):
        try:
            t = self._teslim_cb.currentText().split("  (")[0].strip()
            if t in ("— Kullanıcı seç —", "── Serbest giriş ──"): t = ""
            a = self.arsiv_gor.text().strip()
            if not t: raise ValueError("Teslim alan seçin veya yazın.")
            if acik_movement_var_mi(self.file_id):
                raise ValueError("Bu dosya zaten zimmette. Önce arşive alın.")
            movement_ekle(
                file_id=self.file_id,
                teslim_tarihi=self.teslim_tarihi.date().toString("yyyy-MM-dd"),
                teslim_alan_personel=t, veren_arsiv_gorevlisi=a,
                notlar=self.notlar.toPlainText(),
                teslim_alan_user_id=self._secili_user_id,
            )
            QMessageBox.information(self, "Başarılı", "Zimmet eklendi.")
            self.accept()
        except Exception as e:
            QMessageBox.critical(self, "Hata", str(e))


class ArsiveAlDialog(QDialog):
    def __init__(self, file_id: int, dosya_no: str, kullanici: dict):
        super().__init__()
        self.file_id = file_id
        lay = _dialog_kur(self, 440, "✅", "Arşive Al", dosya_no)

        self.iade_alan = QLineEdit(kullanici["full_name"])
        self.iade_tarihi = QDateEdit(QDate.currentDate())
        self.iade_tarihi.setCalendarPopup(True)
        self.iade_tarihi.setDisplayFormat("dd.MM.yyyy")

        for etiket, widget in [
            ("İADE ALAN GÖREVLİ", self.iade_alan),
            ("İADE TARİHİ", self.iade_tarihi),
        ]:
            _form_satir(lay, etiket, widget)

        al = QPushButton("✅  Arşive Al")
        al.setObjectName("success")
        _btn_satir(lay, self.reject, al)
        al.clicked.connect(self.kaydet)

    def kaydet(self):
        try:
            i = self.iade_alan.text().strip()
            if not i: raise ValueError("İade alan görevli boş olamaz.")
            if not acik_movement_var_mi(self.file_id):
                raise ValueError("Bu dosya zaten arşivde.")
            file_arsive_al(self.file_id,
                           self.iade_tarihi.date().toString("yyyy-MM-dd"), i)
            QMessageBox.information(self, "Başarılı", "Dosya arşive alındı.")
            self.accept()
        except Exception as e:
            QMessageBox.critical(self, "Hata", str(e))


class SifreDegistirDialog(QDialog):
    def __init__(self, user_id: int):
        super().__init__()
        self.user_id = user_id
        lay = _dialog_kur(self, 420, "🔑", "Şifre Değiştir", "Yeni şifrenizi girin")
        self.eski   = QLineEdit(); self.eski.setEchoMode(QLineEdit.Password); self.eski.setPlaceholderText("Mevcut şifre")
        self.yeni1  = QLineEdit(); self.yeni1.setEchoMode(QLineEdit.Password); self.yeni1.setPlaceholderText("Yeni şifre")
        self.yeni2  = QLineEdit(); self.yeni2.setEchoMode(QLineEdit.Password); self.yeni2.setPlaceholderText("Yeni şifre tekrar")
        for lbl, w in [("Mevcut Şifre", self.eski), ("Yeni Şifre", self.yeni1), ("Tekrar", self.yeni2)]:
            _form_satir(lay, lbl, w)
        kaydet = QPushButton("💾  Kaydet")
        _btn_satir(lay, self.reject, kaydet)
        kaydet.clicked.connect(self._kaydet)

    def _kaydet(self):
        from db import giris_yap, kullanici_sifre_sifirla, tum_kullanicilari_getir
        tum = tum_kullanicilari_getir()
        kullanici = next((k for k in tum if k["id"] == self.user_id), None)
        if not kullanici:
            QMessageBox.critical(self, "Hata", "Kullanıcı bulunamadı."); return
        dogrulama = giris_yap(kullanici["username"], self.eski.text())
        if not dogrulama:
            QMessageBox.critical(self, "Hata", "Mevcut şifre hatalı."); return
        if self.yeni1.text() != self.yeni2.text():
            QMessageBox.critical(self, "Hata", "Yeni şifreler eşleşmiyor."); return
        if len(self.yeni1.text()) < 4:
            QMessageBox.critical(self, "Hata", "Şifre en az 4 karakter olmalı."); return
        kullanici_sifre_sifirla(self.user_id, self.yeni1.text())
        QMessageBox.information(self, "Başarılı", "Şifre değiştirildi.")
        self.accept()


class KullaniciEkleDialog(QDialog):
    def __init__(self):
        super().__init__()
        lay = _dialog_kur(self, 460, "👤", "Yeni Kullanıcı Ekle", "Kullanıcı bilgilerini girin")
        self.ad       = QLineEdit(); self.ad.setPlaceholderText("Ad Soyad")
        self.username = QLineEdit(); self.username.setPlaceholderText("kullaniciadi")
        self.sifre    = QLineEdit(); self.sifre.setEchoMode(QLineEdit.Password); self.sifre.setPlaceholderText("Şifre")
        self.rol_cb   = QComboBox()
        self.rol_cb.addItems(["arsiv — Arşiv Görevlisi", "viewer — Görüntüleyici", "admin — Sistem Yöneticisi"])
        for lbl, w in [("Ad Soyad", self.ad), ("Kullanıcı Adı", self.username),
                        ("Şifre", self.sifre), ("Rol", self.rol_cb)]:
            _form_satir(lay, lbl, w)
        ekle = QPushButton("➕  Ekle")
        _btn_satir(lay, self.reject, ekle)
        ekle.clicked.connect(self._ekle)

    def _ekle(self):
        ad = self.ad.text().strip()
        un = self.username.text().strip()
        sp = self.sifre.text().strip()
        rol = self.rol_cb.currentText().split(" — ")[0]
        if not ad or not un or not sp:
            QMessageBox.warning(self, "Uyarı", "Tüm alanları doldurun."); return
        try:
            kullanici_ekle(un, sp, ad, rol)
            QMessageBox.information(self, "Başarılı", f"{ad} eklendi.")
            self.accept()
        except Exception as e:
            QMessageBox.critical(self, "Hata", str(e))


class KullaniciDuzenleDialog(QDialog):
    def __init__(self, kullanici: dict):
        super().__init__()
        self.k = kullanici
        lay = _dialog_kur(self, 460, "✏️", "Kullanıcı Düzenle", kullanici["full_name"])

        self.ad  = QLineEdit(kullanici["full_name"])
        self.rol_cb = QComboBox()
        roller = ["arsiv — Arşiv Görevlisi", "viewer — Görüntüleyici", "admin — Sistem Yöneticisi"]
        self.rol_cb.addItems(roller)
        mevcut_rol = kullanici["role"]
        for i, r in enumerate(roller):
            if r.startswith(mevcut_rol): self.rol_cb.setCurrentIndex(i); break

        _form_satir(lay, "Ad Soyad", self.ad)
        _form_satir(lay, "Rol",      self.rol_cb)

        # Aksiyon butonları
        btn_row = QHBoxLayout(); btn_row.setSpacing(8)

        aktif_metin = "🔴 Pasife Al" if kullanici["active"] else "🟢 Aktife Al"
        aktif_btn = QPushButton(aktif_metin)
        aktif_btn.setObjectName("ghost")
        aktif_btn.clicked.connect(self._aktif_pasif)

        sifre_btn = QPushButton("🔑 Şifre Sıfırla")
        sifre_btn.setObjectName("ghost")
        sifre_btn.clicked.connect(self._sifre_sifirla)

        kaydet = QPushButton("💾  Kaydet")
        btn_row.addWidget(aktif_btn); btn_row.addWidget(sifre_btn)
        btn_row.addStretch()
        lay.addLayout(btn_row)
        _btn_satir(lay, self.reject, kaydet)
        kaydet.clicked.connect(self._kaydet)

    def _kaydet(self):
        ad  = self.ad.text().strip()
        rol = self.rol_cb.currentText().split(" — ")[0]
        if not ad: QMessageBox.warning(self, "Uyarı", "Ad boş olamaz."); return
        try:
            kullanici_guncelle(self.k["id"], ad, rol)
            QMessageBox.information(self, "Başarılı", "Güncellendi.")
            self.accept()
        except Exception as e:
            QMessageBox.critical(self, "Hata", str(e))

    def _aktif_pasif(self):
        yeni_durum = not self.k["active"]
        kullanici_durum_degistir(self.k["id"], yeni_durum)
        durum_txt = "aktif" if yeni_durum else "pasif"
        QMessageBox.information(self, "Başarılı", f"Kullanıcı {durum_txt} yapıldı.")
        self.accept()

    def _sifre_sifirla(self):
        yeni, ok = QInputDialog.getText(self, "Şifre Sıfırla",
                                         f"{self.k['full_name']} için yeni şifre:",
                                         QLineEdit.Password)
        if ok and yeni.strip():
            kullanici_sifre_sifirla(self.k["id"], yeni.strip())
            QMessageBox.information(self, "Başarılı", "Şifre sıfırlandı.")


class GecmisDialog(QDialog):
    def __init__(self, file_id: int, dosya_no: str):
        super().__init__()
        self.resize(960, 500)
        lay = _dialog_kur(self, 960, "📜", "Hareket Geçmişi", dosya_no)

        gecmis = file_gecmisi_getir(file_id)
        table = _tablo_olustur()
        kolonlar  = ["id","teslim_tarihi","teslim_alan_personel",
                     "veren_arsiv_gorevlisi","iade_tarihi","iade_alan_gorevli","notlar"]
        basliklar = ["#","Teslim Tarihi","Teslim Alan",
                     "Arşiv Görevlisi","İade Tarihi","İade Alan","Notlar"]
        table.setColumnCount(len(kolonlar))
        table.setHorizontalHeaderLabels(basliklar)
        table.setRowCount(len(gecmis))

        for ri, satir in enumerate(gecmis):
            acik = satir.get("iade_tarihi") is None
            bg = QColor(P["amber_bg"]) if acik else QColor(P["surface"])
            for ci, kol in enumerate(kolonlar):
                v = satir.get(kol) or ""
                item = QTableWidgetItem(str(v))
                item.setBackground(bg)
                item.setTextAlignment(Qt.AlignCenter)
                if ci == 2 and acik:   # Teslim alan — bold
                    f = QFont(); f.setBold(True); item.setFont(f)
                table.setItem(ri, ci, item)
        table.resizeRowsToContents()
        lay.addWidget(table)

        if gecmis:
            acik_sayisi = sum(1 for r in gecmis if r.get("iade_tarihi") is None)
            durum_satir = QHBoxLayout()
            info = QLabel(
                f"Toplam {len(gecmis)} hareket  •  "
                f"{'🟡 ' + str(acik_sayisi) + ' açık zimmet' if acik_sayisi else '✅ Tüm hareketler tamamlanmış'}"
            )
            info.setStyleSheet(f"color: {P['txt3']}; font-size: 12px; padding: 8px 0;")
            durum_satir.addWidget(info)
            durum_satir.addStretch()
            kapat = QPushButton("Kapat")
            kapat.setObjectName("ghost")
            kapat.setFixedHeight(38)
            kapat.clicked.connect(self.accept)
            durum_satir.addWidget(kapat)
            lay.addLayout(durum_satir)


class DosyaDuzenleDialog(QDialog):
    """Mevcut dosyanın bilgilerini düzenle."""
    def __init__(self, file_id: int, dosya_no: str, sefligi: str,
                 ada: str = "", parsel: str = ""):
        super().__init__()
        self.file_id = file_id
        lay = _dialog_kur(self, 480, "✏️", "Dosya Bilgisini Düzenle",
                          f"ID: {file_id}")

        self.dosya_no = QLineEdit(dosya_no)
        self.sefligi  = QLineEdit(sefligi)
        self.ada      = QLineEdit(ada)
        self.ada.setPlaceholderText("örn. 245")
        self.parsel   = QLineEdit(parsel)
        self.parsel.setPlaceholderText("örn. 18")

        # Ada-parsel yan yana
        ada_row = QHBoxLayout(); ada_row.setSpacing(8)
        ada_row.addWidget(self.ada); ada_row.addWidget(QLabel("/"))
        ada_row.addWidget(self.parsel)
        ada_w = QWidget(); ada_w.setLayout(ada_row); ada_w.setStyleSheet("background:transparent;")

        for etiket, widget in [
            ("DOSYA NO", self.dosya_no),
            ("ŞEFLİĞİ",  self.sefligi),
            ("ADA / PARSEL", ada_w),
        ]:
            _form_satir(lay, etiket, widget)

        kaydet = QPushButton("💾  Kaydet")
        _btn_satir(lay, self.reject, kaydet)
        kaydet.clicked.connect(self.kaydet)

    def kaydet(self):
        try:
            d = self.dosya_no.text().strip()
            s = self.sefligi.text().strip()
            if not d: raise ValueError("Dosya no boş olamaz.")
            if not s: raise ValueError("Şefliği boş olamaz.")
            file_guncelle(self.file_id, d, s,
                          ada=self.ada.text().strip(),
                          parsel=self.parsel.text().strip())
            QMessageBox.information(self, "Başarılı", "Dosya bilgisi güncellendi.")
            self.accept()
        except Exception as e:
            QMessageBox.critical(self, "Hata", str(e))




class YeniKullaniciDialog(QDialog):
    def __init__(self):
        super().__init__()
        lay = _dialog_kur(self, 460, "👤", "Yeni Kullanıcı Ekle")

        self.username  = QLineEdit()
        self.full_name = QLineEdit()
        self.password  = QLineEdit(); self.password.setEchoMode(QLineEdit.Password)
        self.role = QComboBox()
        self.role.addItems(["viewer", "arsiv", "admin"])

        for etiket, w in [
            ("KULLANICI ADI", self.username),
            ("AD SOYAD", self.full_name),
            ("ŞİFRE", self.password),
            ("ROL", self.role),
        ]:
            _form_satir(lay, etiket, w)

        olustur = QPushButton("👤  Oluştur")
        _btn_satir(lay, self.reject, olustur)
        olustur.clicked.connect(self.kaydet)

    def kaydet(self):
        try:
            u = self.username.text().strip()
            f = self.full_name.text().strip()
            p = self.password.text().strip()
            r = self.role.currentText()
            if not u: raise ValueError("Kullanıcı adı boş olamaz.")
            if not f: raise ValueError("Ad soyad boş olamaz.")
            if not p: raise ValueError("Şifre boş olamaz.")
            kullanici_ekle(u, p, f, r)
            QMessageBox.information(self, "Başarılı", "Kullanıcı oluşturuldu.")
            self.accept()
        except Exception as e:
            QMessageBox.critical(self, "Hata", str(e))


# ─────────────────────────────────────────────────────────────
# ZİMMET PDF
# ─────────────────────────────────────────────────────────────
def zimmet_pdf_olustur(satir: dict, kaydet_yolu: str):
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.lib.units import cm
        from reportlab.platypus import (
            SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, HRFlowable
        )
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.enums import TA_CENTER, TA_LEFT
    except ImportError:
        raise ImportError("pip install reportlab")

    doc = SimpleDocTemplate(kaydet_yolu, pagesize=A4,
                            rightMargin=2.5*cm, leftMargin=2.5*cm,
                            topMargin=2*cm, bottomMargin=2*cm)
    styles = getSampleStyleSheet()

    def _s(name, **kw):
        return ParagraphStyle(name, parent=styles["Normal"], **kw)

    baslik_s = _s("B", fontSize=20, alignment=TA_CENTER, spaceAfter=4,
                  textColor=colors.HexColor("#0A1628"), fontName="Helvetica-Bold")
    alt_s    = _s("A", fontSize=10, alignment=TA_CENTER, spaceAfter=24,
                  textColor=colors.HexColor("#667085"))

    story = [
        Paragraph("T.C. İZMİR BÜYÜKŞEHİR BELEDİYESİ", _s("K", fontSize=11,
                  alignment=TA_CENTER, textColor=colors.HexColor("#667085"),
                  fontName="Helvetica-Bold")),
        Paragraph("ARŞİV ZİMMET FORMU", baslik_s),
        HRFlowable(width="100%", thickness=2, color=colors.HexColor("#2563EB"),
                   spaceAfter=8),
        Paragraph(
            f"Form No: #{satir.get('file_id','?')}  •  "
            f"Düzenlenme: {date.today().strftime('%d.%m.%Y')}  •  {APP_SAHIP}",
            alt_s,
        ),
    ]

    veri = [
        ["ALAN", "BİLGİ"],
        ["Dosya No",        satir.get("orijinal_dosya_no","")],
        ["İlçe",            satir.get("ilce","")],
        ["Şefliği",         satir.get("sefligi","")],
        ["Teslim Alan",     satir.get("teslim_alan_personel","")],
        ["Arşiv Görevlisi", satir.get("veren_arsiv_gorevlisi","")],
        ["Teslim Tarihi",   satir.get("teslim_tarihi","")],
        ["Durum",           satir.get("durum","")],
        ["Bekleme (gün)",   str(satir.get("bekleme_gun",0))],
    ]

    tablo = Table(veri, colWidths=[5*cm, 12*cm])
    tablo.setStyle(TableStyle([
        ("BACKGROUND",    (0,0),(-1,0), colors.HexColor("#0A1628")),
        ("TEXTCOLOR",     (0,0),(-1,0), colors.white),
        ("FONTNAME",      (0,0),(-1,0), "Helvetica-Bold"),
        ("FONTSIZE",      (0,0),(-1,0), 11),
        ("ALIGN",         (0,0),(-1,0), "CENTER"),
        ("BACKGROUND",    (0,1),(0,-1), colors.HexColor("#F0F2F5")),
        ("FONTNAME",      (0,1),(0,-1), "Helvetica-Bold"),
        ("FONTSIZE",      (0,1),(-1,-1), 10),
        ("GRID",          (0,0),(-1,-1), 0.5, colors.HexColor("#E4E7EC")),
        ("ROWBACKGROUNDS",(0,1),(-1,-1),
         [colors.white, colors.HexColor("#F8FAFC")]),
        ("VALIGN",        (0,0),(-1,-1), "MIDDLE"),
        ("TOPPADDING",    (0,0),(-1,-1), 11),
        ("BOTTOMPADDING", (0,0),(-1,-1), 11),
        ("LEFTPADDING",   (0,0),(-1,-1), 14),
    ]))
    story.append(tablo)
    story.append(Spacer(1, 2.5*cm))

    imza_veri = [
        ["TESLIM EDEN — ARŞİV GÖREVLİSİ", "TESLİM ALAN — PERSONEL"],
        ["\n\n\n\n________________________", "\n\n\n\n________________________"],
        ["Ad Soyad / İmza / Tarih", "Ad Soyad / İmza / Tarih"],
    ]
    imza = Table(imza_veri, colWidths=[8.5*cm, 8.5*cm])
    imza.setStyle(TableStyle([
        ("FONTNAME",      (0,0),(-1,0), "Helvetica-Bold"),
        ("FONTSIZE",      (0,0),(-1,-1), 10),
        ("ALIGN",         (0,0),(-1,-1), "CENTER"),
        ("VALIGN",        (0,0),(-1,-1), "MIDDLE"),
        ("TOPPADDING",    (0,0),(-1,-1), 10),
        ("BOTTOMPADDING", (0,0),(-1,-1), 10),
        ("LINEABOVE",     (0,0),(-1,0), 1, colors.HexColor("#E4E7EC")),
        ("TEXTCOLOR",     (0,0),(-1,0), colors.HexColor("#344054")),
        ("TEXTCOLOR",     (0,2),(-1,2), colors.HexColor("#98A2B3")),
        ("FONTSIZE",      (0,2),(-1,2), 9),
    ]))
    story.append(imza)
    story.append(Spacer(1, 1*cm))
    story.append(HRFlowable(width="100%", thickness=0.5,
                            color=colors.HexColor("#E4E7EC"), spaceAfter=8))
    story.append(Paragraph(
        f"{APP_IMZA}  •  {DESTEK_TEL}",
        _s("F", fontSize=9, alignment=TA_CENTER, textColor=colors.HexColor("#98A2B3")),
    ))
    doc.build(story)


# ─────────────────────────────────────────────────────────────
# ANA PENCERE
# ─────────────────────────────────────────────────────────────
class MainWindow(QMainWindow):
    KOL = [
        "file_id","ilce","orijinal_dosya_no","ada","parsel",
        "teslim_alan_personel","veren_arsiv_gorevlisi",
        "teslim_tarihi","durum","bekleme_gun","hareket_sayisi",
    ]
    BSL = [
        "ID","İlçe","Dosya No","Ada","Parsel",
        "Teslim Alan","Arşiv Görevlisi",
        "Teslim Tarihi","Durum","Bekleme (g)","Hareket",
    ]

    def __init__(self, kullanici: dict):
        super().__init__()
        self.kullanici = kullanici
        self._data: list[dict] = []
        self._filtreli: list[dict] = []
        self._koyu = False

        self.setWindowTitle(
            f"{APP_TITLE}  |  {kullanici['full_name']}  ·  "
            f"{ROL_ETIKET.get(kullanici['role'], kullanici['role'])}"
        )
        # Minimum boyut — MacBook 13" için
        self.setMinimumSize(1024, 680)
        # Ekran boyutuna göre ayarla
        ekran = QApplication.primaryScreen().availableGeometry()
        genislik = min(1720, int(ekran.width() * 0.92))
        yukseklik = min(980, int(ekran.height() * 0.90))
        self.resize(genislik, yukseklik)
        # Ekrana ortala
        self.move(
            (ekran.width()  - genislik)  // 2,
            (ekran.height() - yukseklik) // 2,
        )
        self._kur()
        self._menubar()
        self._status_bar()
        self.veriyi_yukle()
        self._admin_sekmeleri_yukle()
        self._sayfa_degistir(0)

    # ── YAPILANMA ────────────────────────────────────────────
    def _kur(self):
        merkez = QWidget()
        self.setCentralWidget(merkez)
        root = QHBoxLayout(merkez)
        root.setContentsMargins(0, 0, 0, 0)
        root.setSpacing(0)

        # Badge referansı — sidebar ve mesaj sayfası arasında paylaşılır
        self._badge_ref = {"sayac": 0, "widget": None}

        root.addWidget(self._sidebar_olustur())
        self._stack = QStackedWidget()
        root.addWidget(self._stack)
        self._stack.addWidget(self._sayfa_panel())
        self._stack.addWidget(self._sayfa_kayitlar())
        self._stack.addWidget(self._sayfa_istatistik())
        self._stack.addWidget(self._sayfa_kullanicilar())
        self._stack.addWidget(self._sayfa_loglar())
        self._uzerimdeki_sayfa = self._sayfa_uzerimdeki()
        self._stack.addWidget(self._uzerimdeki_sayfa)

        # Mesaj sayfası
        self._mesaj_sayfa, self._mesaj_guncelle = _mesaj_sayfasi_olustur(
            self.kullanici, self._stack, self._nav_btns, self._badge_ref
        )
        self._stack.addWidget(self._mesaj_sayfa)

    # ── SİDEBAR ──────────────────────────────────────────────
    def _sidebar_olustur(self) -> QWidget:
        sb = QWidget()
        sb.setMinimumWidth(200)
        sb.setMaximumWidth(240)
        sb.setStyleSheet(f"background: {P['navy']};")

        lay = QVBoxLayout(sb)
        lay.setContentsMargins(12, 0, 12, 16)
        lay.setSpacing(2)

        # Logo bloğu
        logo_frame = QFrame()
        logo_frame.setFixedHeight(72)
        logo_frame.setStyleSheet("background: transparent;")
        logo_lay = QHBoxLayout(logo_frame)
        logo_lay.setContentsMargins(6, 0, 6, 0)
        logo_lay.setSpacing(10)

        logo_ikon = QLabel("🗂")
        logo_ikon.setFixedSize(36, 36)
        logo_ikon.setAlignment(Qt.AlignCenter)
        logo_ikon.setStyleSheet(f"""
            background: rgba(37,99,235,0.25);
            border: 1.5px solid rgba(37,99,235,0.4);
            border-radius: 10px;
            font-size: 18px;
        """)
        logo_txt = QVBoxLayout()
        logo_txt.setSpacing(0)
        app_name = QLabel("Arşiv Takip")
        app_name.setStyleSheet("color: white; font-size: 14px; font-weight: 800;")
        ver_lbl = QLabel(APP_VERSIYON)
        ver_lbl.setStyleSheet(f"color: {P['navy_text']}; font-size: 10px;")
        logo_txt.addWidget(app_name)
        logo_txt.addWidget(ver_lbl)

        logo_lay.addWidget(logo_ikon)
        logo_lay.addLayout(logo_txt)
        logo_lay.addStretch()
        lay.addWidget(logo_frame)

        # Ayırıcı
        sep = QFrame()
        sep.setFixedHeight(1)
        sep.setStyleSheet(f"background: rgba(255,255,255,0.08); border: none;")
        lay.addWidget(sep)
        lay.addSpacing(8)

        # Nav etiket
        nav_lbl = QLabel("MENÜ")
        nav_lbl.setStyleSheet(
            f"color: {P['navy_text']}; font-size: 10px; font-weight: 700; "
            "letter-spacing: 1px; padding: 0 8px;"
        )
        lay.addWidget(nav_lbl)
        lay.addSpacing(4)

        self._nav_btns: list[NavButon] = []
        navlar = [
            ("🏠", "Ana Panel",        0),
            ("📂", "Dosya Kayıtları",  1),
            ("📊", "İstatistikler",    2),
        ]
        if self.kullanici["role"] == "admin":
            navlar += [
                ("👥", "Kullanıcılar", 3),
                ("📋", "Loglar",       4),
            ]

        # Mesajlar stack index 6, Üzerimdekiler stack index 5
        self._mesaj_sayfa_idx = 6
        # Üzerimdekiler (index 5 — stack sırasıyla eşleşiyor)
        self._uzerimdeki_idx = 5
        navlar_final = list(navlar)
        navlar_final.append(("📌", "Üzerimdekiler", 5))
        navlar_final.append(("💬", "Mesajlar", 6))

        for ikon, metin, stack_idx in navlar_final:
            btn = NavButon(ikon, metin)
            btn._stack_idx = stack_idx
            btn.clicked.connect(lambda _, idx=stack_idx: self._sayfa_degistir(idx))
            self._nav_btns.append(btn)

            # Mesajlar butonu için rozet ekle
            if metin == "Mesajlar":
                btn_wrap = QWidget()
                btn_wrap.setStyleSheet("background: transparent;")
                btn_wrap_lay = QHBoxLayout(btn_wrap)
                btn_wrap_lay.setContentsMargins(0, 0, 0, 0)
                btn_wrap_lay.setSpacing(0)
                btn_wrap_lay.addWidget(btn, stretch=1)

                badge = QLabel("")
                badge.setFixedSize(20, 20)
                badge.setAlignment(Qt.AlignCenter)
                badge.setStyleSheet(f"""
                    background: {P['red']};
                    color: white;
                    border-radius: 10px;
                    font-size: 10px;
                    font-weight: 700;
                """)
                badge.setVisible(False)
                badge.move(-8, 4)
                btn_wrap_lay.addWidget(badge)
                btn_wrap_lay.setAlignment(badge, Qt.AlignTop)

                self._badge_ref["widget"] = badge
                lay.addWidget(btn_wrap)
            else:
                lay.addWidget(btn)

        lay.addStretch()

        # Alt kullanıcı kartı
        sep2 = QFrame()
        sep2.setFixedHeight(1)
        sep2.setStyleSheet(f"background: rgba(255,255,255,0.08); border: none;")
        lay.addWidget(sep2)
        lay.addSpacing(12)

        user_frame = QFrame()
        user_frame.setStyleSheet(f"""
            QFrame {{
                background: rgba(255,255,255,0.05);
                border: 1px solid rgba(255,255,255,0.08);
                border-radius: 12px;
            }}
        """)
        user_lay = QHBoxLayout(user_frame)
        user_lay.setContentsMargins(12, 10, 12, 10)
        user_lay.setSpacing(10)

        # Avatar
        avatar = QLabel(self.kullanici["full_name"][0].upper())
        avatar.setFixedSize(36, 36)
        avatar.setAlignment(Qt.AlignCenter)
        avatar.setStyleSheet(f"""
            background: {P['navy_active']};
            color: white;
            border-radius: 10px;
            font-size: 15px;
            font-weight: 700;
        """)

        user_txt = QVBoxLayout()
        user_txt.setSpacing(1)
        name_lbl = QLabel(self.kullanici["full_name"])
        name_lbl.setStyleSheet("color: white; font-size: 12px; font-weight: 600;")
        name_lbl.setMaximumWidth(130)

        rol = self.kullanici["role"]
        rol_renk = {"admin": "#F59E0B", "arsiv": "#34D399", "viewer": "#94A3B8"}
        rol_lbl = QLabel(ROL_ETIKET.get(rol, rol))
        rol_lbl.setStyleSheet(
            f"color: {rol_renk.get(rol,'#94A3B8')}; font-size: 10px; font-weight: 600;"
        )
        user_txt.addWidget(name_lbl)
        user_txt.addWidget(rol_lbl)

        user_lay.addWidget(avatar)
        user_lay.addLayout(user_txt)
        user_lay.addStretch()

        lay.addWidget(user_frame)

        # Telif
        telif = QLabel(f"© 2026 {APP_SAHIP}")
        telif.setAlignment(Qt.AlignCenter)
        telif.setStyleSheet(f"color: rgba(148,163,184,0.4); font-size: 10px; padding-top: 8px;")
        lay.addWidget(telif)

        return sb

    def _sayfa_degistir(self, idx: int):
        self._stack.setCurrentIndex(idx)
        for btn in self._nav_btns:
            stack_idx = getattr(btn, '_stack_idx', None)
            if stack_idx is not None:
                btn.setChecked(stack_idx == idx)
            else:
                btn.setChecked(False)

    def _selamlama_guncelle(self):
        """Saate göre selamlama ve tarihi güncelle."""
        saat = datetime.now().hour
        gun  = date.today()
        if saat < 6:    sel, ikon = "İyi geceler",  "🌙"
        elif saat < 12: sel, ikon = "Günaydın",     "☀️"
        elif saat < 14: sel, ikon = "İyi öğleler",  "🌤"
        elif saat < 18: sel, ikon = "İyi günler",   "🌇"
        else:           sel, ikon = "İyi akşamlar", "🌆"
        ad = self.kullanici["full_name"]
        if hasattr(self, '_hos_lbl'):
            self._hos_lbl.setText(f"{ikon}  {sel}, {ad}")
        if hasattr(self, '_tarih_lbl'):
            self._tarih_lbl.setText(
                f"{gun.strftime('%d %B %Y, %A')}  ·  "
                f"{ROL_ETIKET.get(self.kullanici['role'], self.kullanici['role'])}  ·  "
                f"{DESTEK_TEL}"
            )

    def resizeEvent(self, event):
        super().resizeEvent(event)
        if not hasattr(self, '_sidebar'):
            return
        w = self.width()
        if w < 1400:   self._sidebar.setMaximumWidth(200)
        elif w < 1600: self._sidebar.setMaximumWidth(220)
        else:          self._sidebar.setMaximumWidth(240)

    def veriyi_yukle(self):
        try:
            tablo_olustur()
            self._data = tum_files_ozet()
            self._ilce_listesi_yukle()
            self._ozetleri_guncelle()
            self._istatistikleri_guncelle()
            self._dashboard_guncelle()
            self._tablo_goster(self._data)
            self._uzerimdeki_guncelle()
            if hasattr(self, '_hos_lbl'):
                self._selamlama_guncelle()
            if hasattr(self, '_mesaj_log_table'):
                self._mesaj_loglarini_yukle()
        except Exception as e:
            QMessageBox.critical(self, "Hata", f"Veri yüklenemedi:\n{e}")

    # ── ANA PANEL (Dashboard) ─────────────────────────────────

    def _sayfa_panel(self) -> QWidget:
        sayfa = QWidget()
        sayfa.setStyleSheet(f"background:{P['bg']};")
        ana = QVBoxLayout(sayfa)
        ana.setContentsMargins(28, 24, 28, 20)
        ana.setSpacing(0)

        # ── BAŞLIK ───────────────────────────────────────────────
        hdr = QFrame()
        hdr.setStyleSheet("background:transparent;")
        hdr_lay = QHBoxLayout(hdr)
        hdr_lay.setContentsMargins(0, 0, 0, 0)
        sol = QVBoxLayout(); sol.setSpacing(2)
        self._hos_lbl = QLabel()
        self._hos_lbl.setStyleSheet(
            f"font-size:22px; font-weight:800; color:{P['txt']}; letter-spacing:-0.5px;"
        )
        self._tarih_lbl = QLabel()
        self._tarih_lbl.setStyleSheet(f"font-size:12px; color:{P['txt4']};")
        sol.addWidget(self._hos_lbl); sol.addWidget(self._tarih_lbl)
        self._selamlama_guncelle()
        hdr_lay.addLayout(sol); hdr_lay.addStretch()
        yenile = QPushButton("↻  Yenile")
        yenile.setObjectName("ghost"); yenile.setFixedHeight(36)
        yenile.clicked.connect(self.veriyi_yukle)
        hdr_lay.addWidget(yenile)
        ana.addWidget(hdr)
        ana.addSpacing(16)

        # ── UYARI BANDI ──────────────────────────────────────────
        self._banner = QLabel("")
        self._banner.setVisible(False)
        self._banner.setAlignment(Qt.AlignCenter)
        self._banner.setStyleSheet(f"""
            background:{P['red_bg']}; color:{P['red_t']};
            border:1.5px solid #FECACA; border-radius:10px;
            padding:12px 20px; font-size:13px; font-weight:600;
        """)
        ana.addWidget(self._banner)
        ana.addSpacing(4)

        # ── ÜSTTERE SATIR: 6 metrik kart ─────────────────────────
        kart_lay = QHBoxLayout(); kart_lay.setSpacing(10)
        self._km_toplam   = KartMetrik("TOPLAM DOSYA",      "gray")
        self._km_arsivde  = KartMetrik("ARŞİVDE",           "green")
        self._km_zimmette = KartMetrik("ZİMMETTE",          "blue")
        self._km_gecikmis = KartMetrik("10+ GÜN GECİKMİŞ", "red")
        self._km_bugun    = KartMetrik("BUGÜN HAREKETLENDİ","blue")
        self._km_bu_hafta = KartMetrik("BU HAFTA",          "gray")
        for k in [self._km_toplam, self._km_arsivde, self._km_zimmette,
                  self._km_gecikmis, self._km_bugun, self._km_bu_hafta]:
            kart_lay.addWidget(k)
        ana.addLayout(kart_lay)
        ana.addSpacing(12)

        # ── ANA İÇERİK: 3 kolon ──────────────────────────────────
        icerik = QHBoxLayout(); icerik.setSpacing(12)

        # SOL: Hızlı eylemler + Aktivite akışı
        sol_lay = QVBoxLayout(); sol_lay.setSpacing(10)
        sol_lay.addWidget(self._hizli_eylem_karti())
        sol_lay.addWidget(self._aktivite_akisi_karti(), stretch=1)
        sol_w = QWidget(); sol_w.setStyleSheet("background:transparent;")
        sol_w.setLayout(sol_lay)
        sol_w.setMinimumWidth(210); sol_w.setMaximumWidth(280)
        icerik.addWidget(sol_w, stretch=2)

        # ORTA: Trend grafiği + Gecikmiş tablo
        orta_lay = QVBoxLayout(); orta_lay.setSpacing(10)
        orta_lay.addWidget(self._trend_grafiği_karti())
        orta_lay.addWidget(self._gecmis_tablo_karti(), stretch=1)
        icerik.addWidget(self._orta_widget(orta_lay), stretch=5)

        # SAĞ: Donut + Personel + Sistem
        sag_lay = QVBoxLayout(); sag_lay.setSpacing(10)
        sag_lay.addWidget(self._mini_donut_karti())
        sag_lay.addWidget(self._personel_ozet_karti(), stretch=1)
        sag_lay.addWidget(self._sistem_bilgi_karti())
        sag_w = QWidget(); sag_w.setStyleSheet("background:transparent;")
        sag_w.setLayout(sag_lay)
        sag_w.setMinimumWidth(200); sag_w.setMaximumWidth(290)
        icerik.addWidget(sag_w, stretch=2)

        ana.addLayout(icerik, stretch=1)
        return sayfa

    def _orta_widget(self, lay):
        w = QWidget(); w.setStyleSheet("background:transparent;")
        w.setLayout(lay); return w

    def _kart_cerceve(self, baslik="", ikon=""):
        frame = QFrame()
        frame.setStyleSheet(f"""
            QFrame {{
                background:{P['surface']}; border:1px solid {P['border']};
                border-radius:14px;
            }}
        """)
        lay = QVBoxLayout(frame)
        lay.setContentsMargins(16, 14, 16, 14); lay.setSpacing(8)
        if baslik:
            ust = QHBoxLayout(); ust.setSpacing(6)
            if ikon:
                il = QLabel(ikon); il.setStyleSheet("font-size:13px;background:transparent;")
                ust.addWidget(il)
            b = QLabel(baslik)
            b.setStyleSheet(f"font-size:12px; font-weight:700; color:{P['txt2']}; letter-spacing:0.2px;")
            ust.addWidget(b); ust.addStretch()
            lay.addLayout(ust)
            sep = QFrame(); sep.setFixedHeight(1)
            sep.setStyleSheet(f"background:{P['border']}; border:none;")
            lay.addWidget(sep)
        return frame, lay

    def _hizli_eylem_karti(self):
        frame, lay = self._kart_cerceve("Hızlı Eylemler", "⚡")
        rol = self.kullanici["role"]
        eylemler = []
        if rol in ["arsiv","admin"]:
            eylemler.append(("📁  Yeni Dosya Oluştur", self._yeni_dosya, True))
        eylemler += [
            ("📂  Dosya Kayıtları",     lambda: self._sayfa_degistir(1), False),
            ("📤  Zimmettekiler",        self._zimmettekiler_goster,      False),
            ("📊  İstatistikler",        lambda: self._sayfa_degistir(2), False),
            ("⚠️  Gecikenler",           self._gecikenlere_git,           False),
        ]
        if rol in ["arsiv","admin"]:
            eylemler.append(("📥  Arşive Gönderilmiş", self._gonderilenleri_goster, False))
        for metin, slot, primary in eylemler:
            btn = QPushButton(metin)
            if not primary: btn.setObjectName("ghost")
            btn.setFixedHeight(36); btn.clicked.connect(slot)
            lay.addWidget(btn)
        lay.addStretch()
        return frame

    def _zimmettekiler_goster(self):
        self._sayfa_degistir(1)
        try:
            v = sorted(
                [r for r in self._data
                 if r.get("durum","") in ("ZİMMETTE","GECİKMİŞ","ARŞİVE GÖNDERİLDİ")],
                key=lambda x: x["bekleme_gun"], reverse=True,
            )
            self._tablo_goster(v)
        except Exception: pass

    def _gonderilenleri_goster(self):
        self._sayfa_degistir(1)
        try:
            v = [r for r in self._data if r.get("durum") == "ARŞİVE GÖNDERİLDİ"]
            self._tablo_goster(v)
        except Exception: pass

    def _gecikenlere_git(self):
        self._sayfa_degistir(1)
        try: self._gecikenleri_goster()
        except Exception: pass

    def _aktivite_akisi_karti(self):
        frame, lay = self._kart_cerceve("Son Hareketler", "🔄")
        self._aktivite_lay = lay
        self._aktivite_itemler = []
        lay.addStretch()
        return frame

    def _trend_grafiği_karti(self):
        frame, lay = self._kart_cerceve("Son 30 Gün Trend", "📈")
        self._trend_widget = TrendGrafikWidget()
        self._trend_widget.setMinimumHeight(120)
        self._trend_widget.setMaximumHeight(140)
        lay.addWidget(self._trend_widget)
        return frame

    def _gecmis_tablo_karti(self):
        frame = QFrame()
        frame.setStyleSheet(f"""
            QFrame {{
                background:{P['surface']}; border:1px solid {P['border']};
                border-radius:14px;
            }}
        """)
        frame_lay = QVBoxLayout(frame)
        frame_lay.setContentsMargins(16, 14, 16, 14); frame_lay.setSpacing(8)
        ust = QHBoxLayout(); ust.setSpacing(8)
        QLabel("⚠️", styleSheet="font-size:13px;background:transparent;",
               parent=frame)
        baslik = QLabel("⚠️  En Çok Bekleyen Dosyalar")
        baslik.setStyleSheet(f"font-size:13px; font-weight:700; color:{P['txt']};")
        hepsi = QPushButton("Tümünü Gör →")
        hepsi.setObjectName("flat")
        hepsi.clicked.connect(lambda: (self._sayfa_degistir(1), self._gecikenleri_goster()))
        ust.addWidget(baslik); ust.addStretch(); ust.addWidget(hepsi)
        sep = QFrame(); sep.setFixedHeight(1)
        sep.setStyleSheet(f"background:{P['border']}; border:none;")
        frame_lay.addLayout(ust); frame_lay.addWidget(sep)
        self._dashboard_tablo = _tablo_olustur()
        self._dashboard_tablo.setColumnCount(5)
        self._dashboard_tablo.setHorizontalHeaderLabels(
            ["Dosya No", "İlçe", "Teslim Alan", "Bekleme (g)", "Durum"]
        )
        self._dashboard_tablo.verticalHeader().setDefaultSectionSize(38)
        self._dashboard_tablo.doubleClicked.connect(self._dashboard_cift_tiklama)
        frame_lay.addWidget(self._dashboard_tablo, stretch=1)
        ipucu = QLabel("💡 Çift tıklayarak hareket geçmişini görüntüleyin.")
        ipucu.setStyleSheet(f"font-size:11px; color:{P['txt4']}; padding:2px 0;")
        frame_lay.addWidget(ipucu)
        return frame

    def _mini_donut_karti(self):
        frame = QFrame()
        frame.setStyleSheet(f"""
            QFrame {{
                background:{P['surface']}; border:1px solid {P['border']};
                border-radius:14px;
            }}
        """)
        lay = QVBoxLayout(frame)
        lay.setContentsMargins(16, 14, 16, 14); lay.setSpacing(8)
        b = QLabel("📊  Durum Dağılımı")
        b.setStyleSheet(f"font-size:12px; font-weight:700; color:{P['txt2']};")
        sep = QFrame(); sep.setFixedHeight(1)
        sep.setStyleSheet(f"background:{P['border']}; border:none;")
        lay.addWidget(b); lay.addWidget(sep)
        self._mini_ring = RingKarti("")
        self._mini_ring.setStyleSheet("background:transparent; border:none;")
        self._mini_ring.setMinimumHeight(160)
        lay.addWidget(self._mini_ring)
        return frame

    def _personel_ozet_karti(self):
        frame, lay = self._kart_cerceve("👤  En Yoğun Personel", "")
        self._personel_ozet_lay = lay
        self._personel_ozet_itemler = []
        lay.addStretch()
        return frame

    def _sistem_bilgi_karti(self):
        frame, lay = self._kart_cerceve("ℹ️  Sistem", "")
        for etiket, deger in [
            ("Uygulama", APP_TITLE),
            ("Versiyon",  APP_VERSIYON),
            ("Geliştirici", APP_SAHIP),
            ("Destek",    DESTEK_TEL),
        ]:
            satir = QHBoxLayout(); satir.setSpacing(4)
            e = QLabel(etiket); e.setStyleSheet(f"font-size:11px; color:{P['txt4']};"); e.setFixedWidth(70)
            d = QLabel(deger);  d.setStyleSheet(f"font-size:11px; color:{P['txt2']}; font-weight:600;"); d.setWordWrap(True)
            satir.addWidget(e); satir.addWidget(d)
            lay.addLayout(satir)
        lay.addStretch()
        return frame

    def _dashboard_cift_tiklama(self):
        row = self._dashboard_tablo.currentRow()
        if row < 0: return
        geciken = sorted(
            [r for r in self._data if "GEC" in (r.get("durum") or "").upper()],
            key=lambda x: x["bekleme_gun"], reverse=True,
        )
        if row < len(geciken):
            s = geciken[row]
            GecmisDialog(s["file_id"], s["orijinal_dosya_no"]).exec()

    def _dashboard_guncelle(self):
        # ── Gecikmiş tablo
        geciken = sorted(
            [r for r in self._data if "GEC" in (r.get("durum") or "").upper()],
            key=lambda x: x["bekleme_gun"], reverse=True,
        )[:12]
        t = self._dashboard_tablo
        t.setRowCount(len(geciken))
        for ri, satir in enumerate(geciken):
            gun = satir.get("bekleme_gun", 0)
            vals = [
                satir.get("orijinal_dosya_no", ""),
                satir.get("ilce", ""),
                satir.get("teslim_alan_personel", ""),
                str(gun),
                satir.get("durum", ""),
            ]
            for ci, v in enumerate(vals):
                item = QTableWidgetItem(v)
                item.setTextAlignment(Qt.AlignCenter)
                item.setBackground(QColor(P["row_red"]))
                if ci == 3:
                    f = QFont(); f.setBold(True); item.setFont(f)
                    item.setForeground(QColor(P["red"]))
                if ci == 4:
                    f2 = QFont(); f2.setBold(True); item.setFont(f2)
                    item.setForeground(QColor(P["red_t"]))
                t.setItem(ri, ci, item)
        t.resizeRowsToContents()

        # ── Mini donut
        oz = istatistik_ozet()
        self._mini_ring.set_data([
            (oz["arsivde"],  "#059669", "Arşivde"),
            (oz["zimmette"], "#2563EB", "Zimmette"),
            (oz["gecikmis"], "#DC2626", "Gecikmiş"),
        ], merkez=f"{oz['toplam']}\ntoplam")

        # ── Özet kartları (gelişmiş)
        try:
            gelismis = ozet_istatistik_gelismis()
            self._km_bugun.set_deger(str(gelismis["bugun"]))
            self._km_bu_hafta.set_deger(str(gelismis["bu_hafta"]))
        except Exception:
            pass

        # ── Trend grafiği
        try:
            trend = trend_verisi_getir(30)
            self._trend_widget.set_data(trend)
        except Exception:
            pass

        # ── Aktivite akışı
        try:
            hareketler = son_hareketleri_getir(8)
            self._aktivite_guncelle(hareketler)
        except Exception:
            pass

        # ── Personel özeti
        for w in self._personel_ozet_itemler:
            try: w.setParent(None)
            except RuntimeError: pass
        self._personel_ozet_itemler.clear()
        lay = self._personel_ozet_lay
        if lay.count() > 0:
            last = lay.itemAt(lay.count()-1)
            if last and last.spacerItem():
                lay.takeAt(lay.count()-1)
        per  = personel_bazli_istatistik()[:5]
        maks = max((r["zimmette"] for r in per), default=1)
        for r in per:
            sw = QWidget(); sw.setStyleSheet("background:transparent;")
            sl = QVBoxLayout(sw); sl.setContentsMargins(0,2,0,2); sl.setSpacing(3)
            ust = QHBoxLayout(); ust.setSpacing(4)
            nm = QLabel(r["personel"][:18]); nm.setStyleSheet(f"font-size:12px; font-weight:600; color:{P['txt2']};")
            sz = QLabel(f"{r['zimmette']} dosya"); sz.setStyleSheet(f"font-size:11px; color:{P['txt4']};")
            ust.addWidget(nm); ust.addStretch(); ust.addWidget(sz)
            sl.addLayout(ust)
            bar = MiniBarWidget(r["zimmette"], maks, P["red"] if r["gecikmis"]>0 else P["blue"])
            bar.setFixedHeight(6); sl.addWidget(bar)
            lay.addWidget(sw)
            self._personel_ozet_itemler.append(sw)
        lay.addStretch()

    def _aktivite_guncelle(self, hareketler):
        # Eski item'ları temizle
        for w in self._aktivite_itemler:
            try: w.setParent(None)
            except RuntimeError: pass
        self._aktivite_itemler.clear()
        lay = self._aktivite_lay
        if lay.count() > 0:
            last = lay.itemAt(lay.count()-1)
            if last and last.spacerItem():
                lay.takeAt(lay.count()-1)

        for h in hareketler:
            islem = h.get("islem", "ZİMMET")
            renk  = P["blue"] if islem == "ZİMMET" else P["green"]
            bg    = P["blue_bg"] if islem == "ZİMMET" else "#EAF3DE"

            sw = QWidget(); sw.setStyleSheet("background:transparent;")
            sl = QHBoxLayout(sw); sl.setContentsMargins(0,3,0,3); sl.setSpacing(8)

            # İşlem rozeti
            badge = QLabel(islem[:3])
            badge.setFixedWidth(36)
            badge.setAlignment(Qt.AlignCenter)
            badge.setStyleSheet(f"""
                background:{bg}; color:{renk};
                border-radius:5px; font-size:9px; font-weight:700; padding:2px 0;
            """)

            # Bilgi
            ic = QVBoxLayout(); ic.setSpacing(0)
            dosya_lbl = QLabel(h.get("dosya_no","")[:14])
            dosya_lbl.setStyleSheet(f"font-size:11px; font-weight:700; color:{P['txt2']};")
            per_lbl = QLabel(f"{h.get('teslim_alan_personel','')[:12]}  ·  {h.get('ilce','')[:8]}")
            per_lbl.setStyleSheet(f"font-size:10px; color:{P['txt4']};")
            ic.addWidget(dosya_lbl); ic.addWidget(per_lbl)

            sl.addWidget(badge); sl.addLayout(ic); sl.addStretch()
            lay.addWidget(sw)
            self._aktivite_itemler.append(sw)

        if not hareketler:
            bos = QLabel("Henüz hareket yok")
            bos.setStyleSheet(f"font-size:12px; color:{P['txt4']}; padding:10px 0;")
            lay.addWidget(bos)
            self._aktivite_itemler.append(bos)
        lay.addStretch()


    def _menubar(self):
        menu = self.menuBar()
        dosya = menu.addMenu("Dosya")
        dosya.addAction(QAction("↻  Yenile", self, triggered=self.veriyi_yukle, shortcut="F5"))
        dosya.addSeparator()
        dosya.addAction(QAction("Çıkış", self, triggered=self.close))
        gorunum = menu.addMenu("Görünüm")
        gorunum.addAction(QAction("🌙  Koyu Tema", self, triggered=self.koyu_tema_ac))
        gorunum.addAction(QAction("☀️  Açık Tema",  self, triggered=self.acik_tema_ac))
        hesap = menu.addMenu("Hesabım")
        hesap.addAction(QAction(
            "🔑  Şifremi Değiştir", self,
            triggered=lambda: SifreDegistirDialog(self.kullanici["id"]).exec()
        ))

    def koyu_tema_ac(self):
        QApplication.instance().setStyleSheet(KOYU_STIL)

    def acik_tema_ac(self):
        QApplication.instance().setStyleSheet(ANA_STIL)

    def _status_bar(self):
        sb = self.statusBar()
        sb.setStyleSheet(f"""
            QStatusBar {{
                background:{P['surface']}; border-top:1px solid {P['border']};
                font-size:12px; color:{P['txt4']};
            }}
        """)
        self._status_kullanici = QLabel(
            f"  👤 {self.kullanici['full_name']}  ·  "
            f"{ROL_ETIKET.get(self.kullanici['role'], self.kullanici['role'])}"
        )
        self._status_imza = QLabel(f"{APP_IMZA}  📞 {DESTEK_TEL}  ")
        sb.addWidget(self._status_kullanici)
        sb.addPermanentWidget(self._status_imza)

    def _sayfa_kayitlar(self) -> QWidget:
        sayfa = QWidget()
        ana = QVBoxLayout(sayfa)
        ana.setContentsMargins(32, 28, 32, 16)
        ana.setSpacing(16)

        # Başlık
        ust = QHBoxLayout()
        ust.addLayout(_bolum_baslik("Dosya Kayıtları", f"Geliştirici: {APP_SAHIP}  •  {APP_IMZA}"))
        ust.addStretch()
        yenile = QPushButton("↻  Yenile")
        yenile.setObjectName("ghost"); yenile.setFixedHeight(38)
        yenile.clicked.connect(self.veriyi_yukle)
        ust.addWidget(yenile)
        ana.addLayout(ust)

        # Uyarı bandı
        self._banner2 = QLabel("")
        self._banner2.setVisible(False)
        self._banner2.setAlignment(Qt.AlignCenter)
        self._banner2.setStyleSheet(f"""
            background:{P['red_bg']}; color:{P['red_t']};
            border:1.5px solid #FECACA; border-radius:12px;
            padding:14px 20px; font-size:14px; font-weight:600;
        """)
        ana.addWidget(self._banner2)

        # Filtre
        filtre = QHBoxLayout(); filtre.setSpacing(8)
        self._arama = QLineEdit()
        self._arama.setPlaceholderText("🔍  Dosya no, ilçe, ada, parsel veya personel ara...")
        self._arama.setMinimumWidth(300); self._arama.setFixedHeight(40)
        self._arama.returnPressed.connect(self._ara)

        self._ilce_cb = QComboBox(); self._ilce_cb.setFixedHeight(40); self._ilce_cb.setMinimumWidth(140)
        self._ilce_cb.currentTextChanged.connect(lambda: self._ara())

        self._durum_cb = QComboBox(); self._durum_cb.setFixedHeight(40)
        self._durum_cb.addItems(["Tüm Durumlar", "ARŞİVDE", "ZİMMETTE", "GECİKMİŞ", "ARŞİVE GÖNDERİLDİ"])
        self._durum_cb.currentTextChanged.connect(lambda: self._ara())

        ara_btn = QPushButton("🔍 Ara"); ara_btn.setFixedHeight(40); ara_btn.clicked.connect(self._ara)
        tum_btn = QPushButton("Tümü"); tum_btn.setObjectName("ghost"); tum_btn.setFixedHeight(40)
        tum_btn.clicked.connect(self._tum_kayitlar)

        filtre.addWidget(self._arama, stretch=1)
        filtre.addWidget(self._ilce_cb)
        filtre.addWidget(self._durum_cb)
        filtre.addWidget(ara_btn)
        filtre.addWidget(tum_btn)

        rol = self.kullanici["role"]
        if rol in ["arsiv","admin"]:
            yeni_btn = QPushButton("📁  Yeni Dosya"); yeni_btn.setFixedHeight(40)
            yeni_btn.clicked.connect(self._yeni_dosya)
            filtre.addWidget(yeni_btn)

        ana.addLayout(filtre)

        # Tablo
        self._table = _tablo_olustur()
        self._table.doubleClicked.connect(self._tablo_cift_tiklama)
        self._table.horizontalHeader().sectionClicked.connect(self._sutun_sirala)
        self._table.setContextMenuPolicy(Qt.CustomContextMenu)
        self._table.customContextMenuRequested.connect(self._sag_tik_menu)
        self._siralama_sutun = -1
        self._siralama_artan = True
        ana.addWidget(self._table)
        return sayfa

    def _sayfa_istatistik(self) -> QWidget:
        sayfa = QWidget()
        ana = QVBoxLayout(sayfa)
        ana.setContentsMargins(32, 28, 32, 20)
        ana.setSpacing(20)
        ana.addLayout(_bolum_baslik("İstatistikler", "İlçe ve personel bazlı zimmet dağılımı"))

        scroll = QScrollArea(); scroll.setWidgetResizable(True); scroll.setFrameShape(QFrame.NoFrame)
        ic = QWidget(); ic_lay = QVBoxLayout(ic); ic_lay.setSpacing(20)

        # Donut kartları
        ust = QHBoxLayout(); ust.setSpacing(16)
        self._ring_durum = RingKarti("📊 Durum Dağılımı")
        self._ring_durum.setMinimumWidth(280); self._ring_durum.setMaximumWidth(400)
        ust.addWidget(self._ring_durum, stretch=1)
        self._ring_ilce = RingKarti("📍 İlçe Dağılımı (Top 5)")
        self._ring_ilce.setMinimumWidth(280); self._ring_ilce.setMaximumWidth(400)
        ust.addWidget(self._ring_ilce, stretch=1)

        # Özet kutusu
        ozet_frame = QFrame()
        ozet_frame.setStyleSheet(f"QFrame{{background:{P['surface']};border:1px solid {P['border']};border-radius:16px;}}")
        ozet_lay = QVBoxLayout(ozet_frame); ozet_lay.setContentsMargins(20,16,20,16); ozet_lay.setSpacing(10)
        ob = QLabel("🔢 Hızlı Özet"); ob.setStyleSheet(f"font-size:13px;font-weight:700;color:{P['txt2']};")
        ozet_lay.addWidget(ob); ozet_lay.addWidget(_sep())
        self._ozet_satirlar = {}
        for key,lbl in [("toplam","Toplam Dosya"),("arsivde","Arşivde"),("zimmette","Zimmette"),
                         ("gecikmis","10+ Gün Gecikmiş"),("ort_gun","Ort. Bekleme"),("max_gun","Max. Bekleme")]:
            satir = QHBoxLayout()
            l = QLabel(lbl); l.setStyleSheet(f"font-size:12px;color:{P['txt3']};")
            v = QLabel("—"); v.setStyleSheet(f"font-size:13px;font-weight:700;color:{P['txt']};"); v.setAlignment(Qt.AlignRight)
            satir.addWidget(l); satir.addStretch(); satir.addWidget(v)
            self._ozet_satirlar[key] = v
            w = QWidget(); w.setStyleSheet("background:transparent;"); w.setLayout(satir)
            ozet_lay.addWidget(w)
        ozet_lay.addStretch()
        ozet_lay.addWidget(_sep())
        self._yedek_lbl = QLabel("Yedek: —"); self._yedek_lbl.setStyleSheet(f"font-size:11px;color:{P['txt4']};"); self._yedek_lbl.setWordWrap(True)
        ozet_lay.addWidget(self._yedek_lbl)
        yedekle_btn = QPushButton("💾  Şimdi Yedekle"); yedekle_btn.setObjectName("ghost"); yedekle_btn.setFixedHeight(36)
        yedekle_btn.clicked.connect(self._manuel_yedekle)
        ozet_lay.addWidget(yedekle_btn)
        ust.addWidget(ozet_frame)
        ic_lay.addLayout(ust)

        # Bar chartlar
        self._bar_ilce = BarKarti("📍 İlçe Bazlı Dosya Dağılımı (Top 12)"); self._bar_ilce.setMinimumHeight(380)
        ic_lay.addWidget(self._bar_ilce)
        self._bar_per = BarKarti("👤 Personel Bazlı Zimmet (Top 12)"); self._bar_per.setMinimumHeight(360)
        ic_lay.addWidget(self._bar_per)

        # Detay tablolar
        for attr, baslik in [("_ilce_tablo","📋 İlçe Detay"), ("_personel_tablo","👤 Personel Detay")]:
            kf = QFrame(); kf.setStyleSheet(f"QFrame{{background:{P['surface']};border:1px solid {P['border']};border-radius:16px;}}")
            kl = QVBoxLayout(kf); kl.setContentsMargins(20,20,20,20); kl.setSpacing(12)
            kb = QLabel(baslik); kb.setStyleSheet(f"font-size:15px;font-weight:700;color:{P['txt']};")
            kl.addWidget(kb); kl.addWidget(_sep())
            t = _tablo_olustur(); t.setMaximumHeight(320)
            setattr(self, attr, t); kl.addWidget(t); ic_lay.addWidget(kf)

        ic_lay.addStretch()
        scroll.setWidget(ic)
        ana.addWidget(scroll)
        return sayfa

    def _manuel_yedekle(self):
        yol = veritabani_yedekle()
        if yol:
            self._yedek_bilgisini_guncelle()
            QMessageBox.information(self, "Başarılı", f"Yedek alındı:\n{yol}")
        else:
            QMessageBox.warning(self, "Hata", "Yedek alınamadı.")

    def _yedek_bilgisini_guncelle(self):
        bilgi = son_yedek_bilgisi()
        if bilgi["son"]:
            self._yedek_lbl.setText(f"Son yedek: {bilgi['son'][:19]}\n{bilgi['adet']} adet · {bilgi['boyut_kb']} KB")
        else:
            self._yedek_lbl.setText("Henüz yedek alınmamış.")

    def _sayfa_kullanicilar(self) -> QWidget:
        sayfa = QWidget()
        ana = QVBoxLayout(sayfa)
        ana.setContentsMargins(32, 28, 32, 20)
        ana.setSpacing(16)
        ana.addLayout(_bolum_baslik("Kullanıcı Yönetimi",
                                    "Sistem kullanıcıları, roller ve şifre yönetimi"))

        # Araç çubuğu
        ara_lay = QHBoxLayout(); ara_lay.setSpacing(8)
        ekle_btn = QPushButton("➕  Yeni Kullanıcı")
        ekle_btn.setFixedHeight(40)
        ekle_btn.clicked.connect(self._kullanici_ekle_dialog)
        ara_lay.addWidget(ekle_btn)
        ara_lay.addStretch()
        yenile_btn = QPushButton("↻  Yenile")
        yenile_btn.setObjectName("ghost"); yenile_btn.setFixedHeight(40)
        yenile_btn.clicked.connect(self._admin_sekmeleri_yukle)
        ara_lay.addWidget(yenile_btn)
        ana.addLayout(ara_lay)

        # Kullanıcı kartları
        scroll = QScrollArea(); scroll.setWidgetResizable(True); scroll.setFrameShape(QFrame.NoFrame)
        ic = QWidget(); ic_lay = QVBoxLayout(ic); ic_lay.setSpacing(10)
        ic_lay.setContentsMargins(0,0,0,0)

        # Kullanıcı tablosu (admin görmesi için)
        kart = QFrame()
        kart.setStyleSheet(f"""
            QFrame {{
                background:{P['surface']}; border:1px solid {P['border']};
                border-radius:16px;
            }}
        """)
        kl = QVBoxLayout(kart); kl.setContentsMargins(20,20,20,20); kl.setSpacing(12)

        ust = QHBoxLayout()
        kb = QLabel("👥  Kullanıcı Listesi")
        kb.setStyleSheet(f"font-size:15px; font-weight:700; color:{P['txt']};")
        ust.addWidget(kb); ust.addStretch()
        kl.addLayout(ust); kl.addWidget(_sep())

        self._users_table = _tablo_olustur()
        self._users_table.doubleClicked.connect(self._kullanici_duzenle)
        self._users_table.setContextMenuPolicy(Qt.CustomContextMenu)
        self._users_table.customContextMenuRequested.connect(self._kullanici_menu)
        kl.addWidget(self._users_table)

        # Aksiyon butonları
        btn_lay = QHBoxLayout(); btn_lay.setSpacing(8)
        duzenle_btn = QPushButton("✏️  Düzenle")
        duzenle_btn.setObjectName("ghost"); duzenle_btn.setFixedHeight(38)
        duzenle_btn.clicked.connect(self._kullanici_duzenle)

        aktif_btn = QPushButton("🔄  Aktif/Pasif")
        aktif_btn.setObjectName("ghost"); aktif_btn.setFixedHeight(38)
        aktif_btn.clicked.connect(self._kullanici_aktif_pasif_hizli)

        sifre_btn = QPushButton("🔑  Şifre Sıfırla")
        sifre_btn.setObjectName("ghost"); sifre_btn.setFixedHeight(38)
        sifre_btn.clicked.connect(self._kullanici_sifre_sifirla_hizli)

        btn_lay.addWidget(duzenle_btn); btn_lay.addWidget(aktif_btn)
        btn_lay.addWidget(sifre_btn); btn_lay.addStretch()
        kl.addLayout(btn_lay)

        ic_lay.addWidget(kart)
        ic_lay.addStretch()
        scroll.setWidget(ic)
        ana.addWidget(scroll)
        return sayfa

    def _kullanici_menu(self, pos):
        from PySide6.QtWidgets import QMenu
        row = self._users_table.currentRow()
        if row < 0: return
        menu = QMenu(self)
        menu.setStyleSheet(f"""
            QMenu{{background:{P['surface']};border:1px solid {P['border']};
                   border-radius:12px;padding:6px;}}
            QMenu::item{{padding:8px 20px;border-radius:8px;
                         color:{P['txt2']};font-size:13px;}}
            QMenu::item:selected{{background:{P['blue_bg']};color:{P['blue_t']};}}
        """)
        menu.addAction("✏️  Düzenle",       self._kullanici_duzenle)
        menu.addAction("🔄  Aktif/Pasif",   self._kullanici_aktif_pasif_hizli)
        menu.addAction("🔑  Şifre Sıfırla", self._kullanici_sifre_sifirla_hizli)
        menu.exec(self._users_table.viewport().mapToGlobal(pos))

    def _kullanici_aktif_pasif_hizli(self):
        row = self._users_table.currentRow()
        if row < 0: QMessageBox.warning(self, "Uyarı", "Önce bir kullanıcı seçin."); return
        tum = tum_kullanicilari_getir()
        if row >= len(tum): return
        k = tum[row]
        if k["id"] == self.kullanici["id"]:
            QMessageBox.warning(self, "Uyarı", "Kendi hesabınızı pasife alamazsınız."); return
        yeni = not k["active"]
        kullanici_durum_degistir(k["id"], yeni)
        durum = "aktif" if yeni else "pasif"
        QMessageBox.information(self, "Başarılı", f"{k['full_name']} {durum} yapıldı.")
        self._admin_sekmeleri_yukle()

    def _kullanici_sifre_sifirla_hizli(self):
        row = self._users_table.currentRow()
        if row < 0: QMessageBox.warning(self, "Uyarı", "Önce bir kullanıcı seçin."); return
        tum = tum_kullanicilari_getir()
        if row >= len(tum): return
        k = tum[row]
        yeni, ok = QInputDialog.getText(
            self, "Şifre Sıfırla",
            f"{k['full_name']} için yeni şifre:",
            QLineEdit.Password
        )
        if ok and yeni.strip():
            kullanici_sifre_sifirla(k["id"], yeni.strip())
            QMessageBox.information(self, "Başarılı", "Şifre sıfırlandı.")

    def _sayfa_loglar(self) -> QWidget:
        sayfa = QWidget()
        ana = QVBoxLayout(sayfa)
        ana.setContentsMargins(32, 28, 32, 20)
        ana.setSpacing(16)
        ana.addLayout(_bolum_baslik("Sistem Logları", "Kullanıcı giriş, işlem ve mesajlaşma kayıtları"))
        tabs = QTabWidget()
        self._login_table  = _tablo_olustur()
        self._action_table = _tablo_olustur()

        mesaj_log_w = QWidget()
        ml_lay = QVBoxLayout(mesaj_log_w); ml_lay.setContentsMargins(0,12,0,0); ml_lay.setSpacing(10)
        filtre_lay = QHBoxLayout(); filtre_lay.setSpacing(8)
        self._mesaj_log_ara = QLineEdit(); self._mesaj_log_ara.setPlaceholderText("🔍  Gönderen, alıcı veya içerik ara..."); self._mesaj_log_ara.setFixedHeight(38)
        self._mesaj_log_tip = QComboBox(); self._mesaj_log_tip.setFixedHeight(38)
        self._mesaj_log_tip.addItems(["Tümü","Özel Mesajlar (MESAJ_GONDER)","Duyurular (MESAJ_DUYURU)","Dosya Ref (MESAJ_DOSYA_REF)","Silinen (MESAJ_SİL)"])
        yb = QPushButton("↻"); yb.setObjectName("ghost"); yb.setFixedSize(38,38); yb.clicked.connect(self._mesaj_loglarini_yukle)
        filtre_lay.addWidget(self._mesaj_log_ara, stretch=1); filtre_lay.addWidget(self._mesaj_log_tip); filtre_lay.addWidget(yb)
        self._mesaj_log_table = _tablo_olustur()
        self._mesaj_log_ara.textChanged.connect(self._mesaj_loglarini_filtrele)
        self._mesaj_log_tip.currentIndexChanged.connect(self._mesaj_loglarini_filtrele)
        ml_lay.addLayout(filtre_lay); ml_lay.addWidget(self._mesaj_log_table)

        tabs.addTab(self._login_table,  "🔐 Giriş Logları")
        tabs.addTab(self._action_table, "📋 İşlem Logları")
        tabs.addTab(mesaj_log_w,        "💬 Mesaj Logları")
        ana.addWidget(tabs)
        return sayfa

    def _sayfa_uzerimdeki(self) -> QWidget:
        """Giriş yapan kullanıcının zimmetindeki dosyalar."""
        sayfa = QWidget()
        ana = QVBoxLayout(sayfa)
        ana.setContentsMargins(32, 28, 32, 20)
        ana.setSpacing(16)

        # Başlık
        ust = QHBoxLayout()
        ust.addLayout(_bolum_baslik(
            "📌  Üzerimdeki Dosyalar",
            f"Zimmetinizde bekleyen dosyalar — {self.kullanici['full_name']}"
        ))
        ust.addStretch()
        yenile = QPushButton("↻  Yenile")
        yenile.setObjectName("ghost"); yenile.setFixedHeight(36)
        yenile.clicked.connect(self._uzerimdeki_guncelle)
        ust.addWidget(yenile)
        ana.addLayout(ust)

        # Özet kartları
        kart_row = QHBoxLayout(); kart_row.setSpacing(10)
        self._uk_toplam   = KartMetrik("ZİMMETİMDEKİ", "blue")
        self._uk_gecikmis = KartMetrik("GECİKMİŞ",     "red")
        self._uk_ort_gun  = KartMetrik("ORT. BEKLEME",  "gray")
        for k in [self._uk_toplam, self._uk_gecikmis, self._uk_ort_gun]:
            kart_row.addWidget(k)
        kart_row.addStretch()
        ana.addLayout(kart_row)

        # Tablo
        kart = QFrame()
        kart.setStyleSheet(f"""
            QFrame {{
                background:{P['surface']}; border:1px solid {P['border']};
                border-radius:16px;
            }}
        """)
        kl = QVBoxLayout(kart); kl.setContentsMargins(20,20,20,20); kl.setSpacing(12)
        tbl_baslik = QLabel("Zimmetinizdeki Dosyalar")
        tbl_baslik.setStyleSheet(f"font-size:15px; font-weight:700; color:{P['txt']};")
        kl.addWidget(tbl_baslik); kl.addWidget(_sep())

        self._uzerimdeki_tablo = _tablo_olustur()
        self._uzerimdeki_tablo.setColumnCount(8)
        self._uzerimdeki_tablo.setHorizontalHeaderLabels([
            "Dosya No", "İlçe", "Şefliği", "Ada", "Parsel",
            "Teslim Tarihi", "Bekleme (g)", "Durum"
        ])
        self._uzerimdeki_tablo.doubleClicked.connect(self._uzerimdeki_gecmis)
        self._uzerimdeki_tablo.setContextMenuPolicy(Qt.CustomContextMenu)
        self._uzerimdeki_tablo.customContextMenuRequested.connect(self._uzerimdeki_menu)
        kl.addWidget(self._uzerimdeki_tablo)

        # Aksiyon butonları — role göre farklı
        btn_lay = QHBoxLayout(); btn_lay.setSpacing(8)
        rol = self.kullanici["role"]

        if rol in ["arsiv", "admin"]:
            arsive_al_btn = QPushButton("✅  Arşive Al")
            arsive_al_btn.setFixedHeight(40)
            arsive_al_btn.setStyleSheet(f"""
                QPushButton {{
                    background:{P['green']}; color:white;
                    border:none; border-radius:10px;
                    font-size:13px; font-weight:700; padding:0 16px;
                }}
                QPushButton:hover {{ background:#059669; }}
            """)
            arsive_al_btn.clicked.connect(lambda: self._uzerimdeki_arsive_al())
            btn_lay.addWidget(arsive_al_btn)

            iptal_btn = QPushButton("🔄  Gönderildi İptal")
            iptal_btn.setObjectName("ghost"); iptal_btn.setFixedHeight(40)
            iptal_btn.clicked.connect(lambda: self._uzerimdeki_gonder_iptal())
            btn_lay.addWidget(iptal_btn)
        else:
            gonder_btn = QPushButton("📤  Arşive Gönder")
            gonder_btn.setFixedHeight(40)
            gonder_btn.setStyleSheet(f"""
                QPushButton {{
                    background:{P['amber_bg']}; color:{P['amber_t']};
                    border:1px solid #FCD34D; border-radius:10px;
                    font-size:13px; font-weight:600; padding:0 16px;
                }}
                QPushButton:hover {{ background:#FEF3C7; }}
            """)
            gonder_btn.clicked.connect(lambda: self._uzerimdeki_arsive_gonder())
            btn_lay.addWidget(gonder_btn)

        gecmis_btn = QPushButton("📜  Geçmiş")
        gecmis_btn.setObjectName("ghost"); gecmis_btn.setFixedHeight(40)
        gecmis_btn.clicked.connect(lambda: self._uzerimdeki_gecmis())
        btn_lay.addWidget(gecmis_btn)
        btn_lay.addStretch()
        kl.addLayout(btn_lay)

        # Alt bilgi
        self._uzerimdeki_bilgi = QLabel("")
        self._uzerimdeki_bilgi.setStyleSheet(f"font-size:11px; color:{P['txt4']};")
        kl.addWidget(self._uzerimdeki_bilgi)
        ana.addWidget(kart)
        return sayfa

    def _uzerimdeki_guncelle(self):
        """Kullanıcının zimmetindeki dosyaları yükle."""
        dosyalar = bende_zimmetli_dosyalar(
            self.kullanici["id"], self.kullanici["full_name"]
        )
        t = self._uzerimdeki_tablo
        t.setRowCount(len(dosyalar))

        gecikmis = 0
        toplam_gun = 0
        for ri, r in enumerate(dosyalar):
            gun = int(r.get("bekleme_gun") or 0)
            durum = r.get("durum", "")
            if "GEC" in durum: gecikmis += 1
            toplam_gun += gun
            bg = P["row_red"] if "GEC" in durum else P["surface"]

            vals = [
                r.get("orijinal_dosya_no",""),
                r.get("ilce",""),
                r.get("sefligi",""),
                r.get("ada","") or "—",
                r.get("parsel","") or "—",
                r.get("teslim_tarihi","")[:10],
                str(gun),
                durum,
            ]
            for ci, v in enumerate(vals):
                item = QTableWidgetItem(v)
                item.setTextAlignment(Qt.AlignCenter)
                item.setBackground(QColor(bg))
                if ci == 6 and gun >= 10:
                    f = QFont(); f.setBold(True); item.setFont(f)
                    item.setForeground(QColor(P["red"]))
                if ci == 7 and "GEC" in durum:
                    f2 = QFont(); f2.setBold(True); item.setFont(f2)
                    item.setForeground(QColor(P["red_t"]))
                t.setItem(ri, ci, item)
            t.setRowHeight(ri, 40)

        t.resizeRowsToContents()
        # Kartları güncelle
        self._uk_toplam.guncelle(len(dosyalar))
        self._uk_gecikmis.guncelle(gecikmis)
        ort = (toplam_gun // len(dosyalar)) if dosyalar else 0
        self._uk_ort_gun.guncelle(ort, "gün")

        if not dosyalar:
            self._uzerimdeki_bilgi.setText("✅  Üzerinizde açık zimmet bulunmuyor.")
        else:
            self._uzerimdeki_bilgi.setText(
                f"💡  Çift tıklayarak hareket geçmişini görüntüleyebilirsiniz.  "
                f"Toplam {len(dosyalar)} dosya."
            )

    def _uzerimdeki_gecmis(self):
        row = self._uzerimdeki_tablo.currentRow()
        if row < 0: return
        dosyalar = bende_zimmetli_dosyalar(
            self.kullanici["id"], self.kullanici["full_name"]
        )
        if row < len(dosyalar):
            d = dosyalar[row]
            GecmisDialog(d["file_id"], d["orijinal_dosya_no"]).exec()

    def _uzerimdeki_secili(self):
        """Üzerimdekiler tablosunda seçili satırı döner."""
        row = self._uzerimdeki_tablo.currentRow()
        if row < 0: return None
        try:
            dosyalar = bende_zimmetli_dosyalar(
                self.kullanici["id"], self.kullanici["full_name"]
            )
            return dosyalar[row] if row < len(dosyalar) else None
        except Exception:
            return None

    def _uzerimdeki_arsive_gonder(self):
        """Üzerimdekiler sayfasından arşive gönder."""
        s = self._uzerimdeki_secili()
        if not s:
            QMessageBox.warning(self, "Uyarı", "Önce tablodan bir dosya seçin."); return
        if s.get("durum") not in ("ZİMMETTE", "GECİKMİŞ"):
            QMessageBox.warning(self, "Uyarı",
                "Bu dosya zimmetli değil veya zaten gönderilmiş."); return
        self._arsive_gonder_with(s)

    def _uzerimdeki_arsive_al(self):
        """Arşiv görevlisi üzerimdekilerden arşive al."""
        s = self._uzerimdeki_secili()
        if not s:
            QMessageBox.warning(self, "Uyarı", "Önce tablodan bir dosya seçin."); return
        if s.get("durum") not in ("ZİMMETTE", "GECİKMİŞ", "ARŞİVE GÖNDERİLDİ"):
            QMessageBox.warning(self, "Uyarı", "Bu dosya zimmetli değil."); return

        # ArsiveAlDialog ile aynı akışı kullan
        d = ArsiveAlDialog(s["file_id"], s["orijinal_dosya_no"], self.kullanici)
        if d.exec():
            action_log_ekle(
                self.kullanici["id"], self.kullanici["username"],
                self.kullanici["full_name"], self.kullanici["role"],
                "ARŞİVE_AL",
                f"file_id={s['file_id']} dosya={s['orijinal_dosya_no']} kaynak=uzerimdekiler"
            )
            self._uzerimdeki_guncelle()
            self.veriyi_yukle()

    def _uzerimdeki_gonder_iptal(self):
        """Arşive gönderildi işlemini iptal et."""
        s = self._uzerimdeki_secili()
        if not s:
            QMessageBox.warning(self, "Uyarı", "Önce tablodan bir dosya seçin."); return
        if s.get("durum") != "ARŞİVE GÖNDERİLDİ":
            QMessageBox.warning(self, "Uyarı",
                "Bu dosya 'Arşive Gönderildi' durumunda değil."); return
        try:
            arsive_gonder_iptal(s["file_id"])
            self._uzerimdeki_guncelle()
            self.veriyi_yukle()
        except Exception as e:
            QMessageBox.critical(self, "Hata", str(e))
        """Üzerimdekiler sayfasından arşive gönder."""
        s = self._uzerimdeki_secili()
        if not s:
            QMessageBox.warning(self, "Uyarı", "Önce tablodan bir dosya seçin."); return
        if s.get("durum") not in ("ZİMMETTE", "GECİKMİŞ"):
            QMessageBox.warning(self, "Uyarı",
                "Bu dosya zaten arşive gönderilmiş veya uygun durumda değil."); return
        # Aynı dialog — _secili_satir yerine uzerimdeki kullan
        self._arsive_gonder_with(s)

    def _uzerimdeki_menu(self, pos):
        """Üzerimdekiler sağ tık menüsü."""
        from PySide6.QtWidgets import QMenu
        s = self._uzerimdeki_secili()
        if not s: return
        menu = QMenu(self)
        menu.setStyleSheet(f"""
            QMenu{{background:{P['surface']};border:1px solid {P['border']};
                   border-radius:12px;padding:6px;}}
            QMenu::item{{padding:8px 20px;border-radius:8px;color:{P['txt2']};font-size:13px;}}
            QMenu::item:selected{{background:{P['blue_bg']};color:{P['blue_t']};}}
        """)
        menu.addAction("📜  Hareket Geçmişi", self._uzerimdeki_gecmis)
        durum = s.get("durum","")
        rol = self.kullanici["role"]
        if rol in ["arsiv","admin"]:
            menu.addSeparator()
            if durum in ("ZİMMETTE","GECİKMİŞ","ARŞİVE GÖNDERİLDİ"):
                menu.addAction("✅  Arşive Al", self._uzerimdeki_arsive_al)
            if durum == "ARŞİVE GÖNDERİLDİ":
                menu.addAction("🔄  Gönderildi İptal", self._uzerimdeki_gonder_iptal)
        elif durum in ("ZİMMETTE","GECİKMİŞ"):
            menu.addSeparator()
            menu.addAction("📤  Arşive Gönder", self._uzerimdeki_arsive_gonder)
        elif durum == "ARŞİVE GÖNDERİLDİ":
            menu.addSeparator()
            menu.addAction("🔄  Gönderildi İptal", lambda: (
                arsive_gonder_iptal(s["file_id"]),
                self._uzerimdeki_guncelle(),
                self.veriyi_yukle()
            ))
        menu.exec(self._uzerimdeki_tablo.viewport().mapToGlobal(pos))

    def _mesaj_loglarini_yukle(self):
        try:
            tumloglar = action_loglarini_getir()
            mesaj_tipleri = {"MESAJ_GONDER","MESAJ_DUYURU","MESAJ_DOSYA_REF","MESAJ_SİL","MESAJ_TOPLU_SİL","KONUŞMA_SİL"}
            self._mesaj_log_tumu = [r for r in tumloglar if r.get("action_type","") in mesaj_tipleri]
            self._mesaj_loglarini_filtrele()
        except Exception: pass

    def _mesaj_loglarini_filtrele(self):
        if not hasattr(self, '_mesaj_log_tumu'): self._mesaj_loglarini_yukle(); return
        ara = self._mesaj_log_ara.text().strip().lower()
        tip_idx = self._mesaj_log_tip.currentIndex()
        tip_filtre = {1:"MESAJ_GONDER",2:"MESAJ_DUYURU",3:"MESAJ_DOSYA_REF",4:"MESAJ_SİL"}.get(tip_idx)
        veriler = self._mesaj_log_tumu
        if tip_filtre: veriler = [r for r in veriler if r.get("action_type")==tip_filtre]
        if ara: veriler = [r for r in veriler if ara in (r.get("full_name") or "").lower() or ara in (r.get("detail") or "").lower()]
        t = self._mesaj_log_table
        t.setColumnCount(5)
        t.setHorizontalHeaderLabels(["Tarih/Saat","Kullanıcı","İşlem","Detay","Rol"])
        t.setRowCount(len(veriler))
        tip_renk = {"MESAJ_GONDER":("#EFF6FF",P["blue_t"],"💬 Mesaj"),"MESAJ_DUYURU":(P["amber_bg"],P["amber_t"],"📢 Duyuru"),
                    "MESAJ_DOSYA_REF":("#F0FDF4","#166534","📎 Dosya"),"MESAJ_SİL":(P["red_bg"],P["red_t"],"🗑 Silindi"),
                    "MESAJ_TOPLU_SİL":(P["red_bg"],P["red_t"],"🗑 Toplu"),"KONUŞMA_SİL":(P["red_bg"],P["red_t"],"🗑 Sohbet")}
        for ri,r in enumerate(veriler):
            tip = r.get("action_type","")
            bg,fg,tip_lbl = tip_renk.get(tip,(P["surface"],P["txt"],tip))
            for ci,v in enumerate([str(r.get("olusturma",""))[:19],str(r.get("full_name","")),tip_lbl,str(r.get("detail","")),str(r.get("role",""))]):
                item = QTableWidgetItem(v); item.setBackground(QColor(bg))
                if ci==2: item.setForeground(QColor(fg)); f=QFont(); f.setBold(True); item.setFont(f)
                t.setItem(ri,ci,item)
            t.setRowHeight(ri,40)
        t.resizeRowsToContents()

    def _ilce_listesi_yukle(self):
        self._ilce_cb.blockSignals(True)
        mevcut = self._ilce_cb.currentText()
        self._ilce_cb.clear(); self._ilce_cb.addItem("Tüm İlçeler")
        for i in sorted({r["ilce"] for r in self._data if r.get("ilce")}):
            self._ilce_cb.addItem(i)
        idx = self._ilce_cb.findText(mevcut)
        if idx >= 0: self._ilce_cb.setCurrentIndex(idx)
        self._ilce_cb.blockSignals(False)

    def _ozetleri_guncelle(self):
        oz = istatistik_ozet()
        self._km_toplam.guncelle(oz["toplam"])
        self._km_arsivde.guncelle(oz["arsivde"])
        self._km_zimmette.guncelle(oz["zimmette"])
        self._km_gecikmis.guncelle(oz["gecikmis"])
        gec = oz["gecikmis"]

        # Arşive gönderilmiş bekleyenler (arşiv/admin için)
        if self.kullanici["role"] in ["arsiv", "admin"]:
            try:
                bekleyenler = arsive_gonderilen_dosyalar()
            except Exception:
                bekleyenler = []
            if bekleyenler:
                self._banner.setText(
                    f"📤  {len(bekleyenler)} dosya arşive gönderildi — onay bekliyor!  "
                    f"{'  |  ⚠️  ' + str(gec) + ' gecikmiş' if gec >= 10 else ''}"
                )
                self._banner.setStyleSheet(f"""
                    background:{P['amber_bg']}; color:{P['amber_t']};
                    border:1.5px solid #FCD34D; border-radius:10px;
                    padding:12px 20px; font-size:13px; font-weight:600;
                """)
                self._banner.setVisible(True)
                if hasattr(self, '_banner2'): self._banner2.setVisible(False)
                return

        if gec >= 10:
            self._banner.setStyleSheet(f"""
                background:{P['red_bg']}; color:{P['red_t']};
                border:1.5px solid #FECACA; border-radius:10px;
                padding:12px 20px; font-size:13px; font-weight:600;
            """)
            self._banner.setText(f"⚠️  {gec} dosya 10 günden fazladır teslim edilmedi!")
            self._banner.setVisible(True)
        else:
            self._banner.setVisible(False)
        if hasattr(self, '_banner2'):
            self._banner2.setVisible(False)

    def _secili_satir(self) -> dict | None:
        row = self._table.currentRow()
        if row < 0 or row >= len(self._filtreli): return None
        return self._filtreli[row]

    def _tablo_cift_tiklama(self):
        s = self._secili_satir()
        if s: GecmisDialog(s["file_id"], s["orijinal_dosya_no"]).exec()

    def _sag_tik_menu(self, pos):
        from PySide6.QtWidgets import QMenu
        s = self._secili_satir()
        if not s: return
        menu = QMenu(self)
        menu.setStyleSheet(f"""
            QMenu{{background:{P['surface']};border:1px solid {P['border']};border-radius:12px;padding:6px;}}
            QMenu::item{{padding:8px 20px;border-radius:8px;color:{P['txt2']};font-size:13px;}}
            QMenu::item:selected{{background:{P['blue_bg']};color:{P['blue_t']};}}
            QMenu::separator{{background:{P['border']};height:1px;margin:4px 8px;}}
        """)
        menu.addAction("📜  Hareket Geçmişi", self._gecmis_goster)
        durum = s.get("durum","") if s else ""
        if self.kullanici["role"] in ["arsiv","admin"]:
            uid = self.kullanici["id"]
            uname = self.kullanici["full_name"]
            teslim_alan = (s.get("teslim_alan_personel") or "").strip().lower()
            benim = (s.get("teslim_alan_user_id") == uid or
                     teslim_alan == uname.strip().lower())
            menu.addSeparator()
            menu.addAction("➕  Zimmet Ekle",         self._zimmet_ekle)
            if durum == "ARŞİVE GÖNDERİLDİ":
                menu.addAction("✅  Arşive Al (Onayla)", self._arsive_al)
                menu.addAction("🔄  Gönderildi İptal",   self._arsive_gonder_iptal)
            else:
                menu.addAction("✅  Arşive Al",           self._arsive_al)
            if benim and durum in ("ZİMMETTE", "GECİKMİŞ"):
                menu.addAction("📤  Arşive Gönder",      self._arsive_gonder)
            menu.addAction("✏️  Dosyayı Düzenle",        self._dosya_duzenle)
            menu.addAction("🗑  Dosyayı Sil",             self._dosya_sil)
            menu.addSeparator()
        else:
            # viewer rolü — sadece kendi zimmetindeyse gönderebilir
            uid = self.kullanici["id"]
            uname = self.kullanici["full_name"]
            teslim_alan = (s.get("teslim_alan_personel") or "").strip().lower()
            benim = (s.get("teslim_alan_user_id") == uid or
                     teslim_alan == uname.strip().lower())
            if benim and durum in ("ZİMMETTE", "GECİKMİŞ"):
                menu.addSeparator()
                menu.addAction("📤  Arşive Gönder",       self._arsive_gonder)
                menu.addSeparator()
            elif benim and durum == "ARŞİVE GÖNDERİLDİ":
                menu.addSeparator()
                menu.addAction("🔄  Arşive Gönder İptal", self._arsive_gonder_iptal)
                menu.addSeparator()
        menu.addAction("🖨  Zimmet PDF",         self._zimmet_pdf)
        menu.addAction("📋  Dosya No Kopyala",   lambda: QApplication.clipboard().setText(s.get("orijinal_dosya_no","")))
        menu.exec(self._table.viewport().mapToGlobal(pos))

    def _dosya_duzenle(self):
        if self.kullanici["role"] not in ["arsiv","admin"]:
            QMessageBox.warning(self,"Yetki","Bu işlem için yetkiniz yok."); return
        s = self._secili_satir()
        if not s: QMessageBox.warning(self,"Uyarı","Önce bir dosya seçin."); return
        d = DosyaDuzenleDialog(
                s["file_id"],
                s.get("orijinal_dosya_no",""),
                s.get("sefligi",""),
                s.get("ada",""),
                s.get("parsel",""),
            )
        if d.exec():
            action_log_ekle(self.kullanici["id"],self.kullanici["username"],self.kullanici["full_name"],self.kullanici["role"],"DOSYA_DUZENLE",f"file_id={s['file_id']}")
            self.veriyi_yukle()

    def _kullanici_ekle_dialog(self):
        d = KullaniciEkleDialog()
        if d.exec(): self._admin_sekmeleri_yukle()

    def _kullanici_duzenle(self):
        row = self._users_table.currentRow()
        if row < 0: return
        tum = tum_kullanicilari_getir()
        if row < len(tum):
            d = KullaniciDuzenleDialog(tum[row])
            if d.exec(): self._admin_sekmeleri_yukle()

    def _excel_aktar(self):
        from PySide6.QtWidgets import QFileDialog
        yol, _ = QFileDialog.getSaveFileName(self,"Excel Kaydet","zimmet_raporu.xlsx","Excel (*.xlsx)")
        if not yol: return
        try:
            import openpyxl
            wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Zimmet Listesi"
            basliklar = ["Dosya No","İlçe","Şefliği","Ada","Parsel","Durum","Teslim Alan","Bekleme (gün)","Teslim Tarihi"]
            ws.append(basliklar)
            for r in self._filtreli:
                ws.append([r.get("orijinal_dosya_no",""),r.get("ilce",""),r.get("sefligi",""),
                           r.get("ada",""),r.get("parsel",""),r.get("durum",""),
                           r.get("teslim_alan_personel",""),r.get("bekleme_gun",""),r.get("teslim_tarihi","")])
            wb.save(yol)
            QMessageBox.information(self,"Başarılı",f"Excel kaydedildi:\n{yol}")
        except Exception as e:
            QMessageBox.critical(self,"Hata",str(e))

    def _istatistikleri_guncelle(self):
        oz = istatistik_ozet()
        ilce_data = ilce_bazli_istatistik()
        per_data  = personel_bazli_istatistik()
        veriler   = self._data

        # ── Donut — durum dağılımı
        self._ring_durum.set_data([
            (oz["arsivde"],  "#059669", "Arşivde"),
            (oz["zimmette"], "#2563EB", "Zimmette"),
            (oz["gecikmis"], "#DC2626", "Gecikmiş"),
        ], merkez=str(oz["toplam"]))

        # ── Donut — ilçe top5
        top5 = ilce_data[:5]
        renkler = ["#2563EB","#7C3AED","#059669","#D97706","#DC2626"]
        self._ring_ilce.set_data([
            (r["toplam"], renkler[i], r["ilce"])
            for i, r in enumerate(top5)
        ], merkez=f"{len(ilce_data)}\nilçe")

        # ── Bar chart — ilçe
        self._bar_ilce.set_data([
            (r["ilce"], r["toplam"], r["zimmette"], r["gecikmis"])
            for r in ilce_data[:12]
        ])

        # ── Bar chart — personel
        self._bar_per.set_data([
            (r["personel"], r["zimmette"], 0, r["gecikmis"])
            for r in per_data[:12]
        ])

        # ── Özet metrikler
        gunler = [r["bekleme_gun"] for r in veriler if r["bekleme_gun"] > 0]
        ort = sum(gunler) // len(gunler) if gunler else 0
        maks = max(gunler) if gunler else 0
        for key, val in [
            ("toplam",   f"{oz['toplam']:,}".replace(",",".")),
            ("arsivde",  f"{oz['arsivde']:,}".replace(",",".")),
            ("zimmette", str(oz["zimmette"])),
            ("gecikmis", str(oz["gecikmis"])),
            ("ort_gun",  f"{ort} gün"),
            ("max_gun",  f"{maks} gün"),
        ]:
            if key in self._ozet_satirlar:
                self._ozet_satirlar[key].setText(val)
                if key == "gecikmis" and oz["gecikmis"] > 0:
                    self._ozet_satirlar[key].setStyleSheet(
                        f"font-size: 13px; font-weight: 700; color: {P['red']};"
                    )

        # ── Yedek bilgisi
        self._yedek_bilgisini_guncelle()

        # ── İlçe detay tablosu
        maks_toplam = max((r["toplam"] for r in ilce_data), default=1)
        self._ilce_tablo.setColumnCount(5)
        self._ilce_tablo.setHorizontalHeaderLabels(
            ["İlçe", "Toplam", "Zimmette", "Gecikmiş", "Dağılım"]
        )
        self._ilce_tablo.setRowCount(len(ilce_data))
        for ri, satir in enumerate(ilce_data):
            vals = [
                satir.get("ilce",""),
                str(satir.get("toplam",0)),
                str(satir.get("zimmette",0)),
                str(satir.get("gecikmis",0)),
            ]
            bg = QColor(P["surface"]) if ri % 2 == 0 else QColor(P["surface2"])
            for ci, v in enumerate(vals):
                item = QTableWidgetItem(v)
                item.setTextAlignment(Qt.AlignCenter)
                item.setBackground(bg)
                if ci == 3 and int(satir.get("gecikmis",0)) > 0:
                    item.setBackground(QColor(P["row_red"]))
                    item.setForeground(QColor(P["red_t"]))
                    f = QFont(); f.setBold(True); item.setFont(f)
                self._ilce_tablo.setItem(ri, ci, item)
            bar = MiniBarWidget(satir.get("toplam",0), maks_toplam, P["blue"])
            self._ilce_tablo.setCellWidget(ri, 4, bar)
            self._ilce_tablo.setRowHeight(ri, 44)
        self._ilce_tablo.resizeRowsToContents()

        # ── Personel detay tablosu
        maks_zim = max((r["zimmette"] for r in per_data), default=1)
        self._personel_tablo.setColumnCount(5)
        self._personel_tablo.setHorizontalHeaderLabels(
            ["Personel", "Zimmette", "Gecikmiş", "En Uzun (gün)", "Yoğunluk"]
        )
        self._personel_tablo.setRowCount(len(per_data))
        for ri, satir in enumerate(per_data):
            vals = [
                satir.get("personel",""),
                str(satir.get("zimmette",0)),
                str(satir.get("gecikmis",0)),
                str(satir.get("max_gun",0)),
            ]
            bg = QColor(P["surface"]) if ri % 2 == 0 else QColor(P["surface2"])
            for ci, v in enumerate(vals):
                item = QTableWidgetItem(v)
                item.setTextAlignment(Qt.AlignCenter)
                item.setBackground(bg)
                if ci == 2 and int(satir.get("gecikmis",0)) > 0:
                    item.setBackground(QColor(P["row_red"]))
                    item.setForeground(QColor(P["red_t"]))
                    f = QFont(); f.setBold(True); item.setFont(f)
                if ci == 3 and int(satir.get("max_gun",0)) > 100:
                    item.setForeground(QColor(P["red"]))
                self._personel_tablo.setItem(ri, ci, item)
            bar = MiniBarWidget(satir.get("zimmette",0), maks_zim, P["red"])
            self._personel_tablo.setCellWidget(ri, 4, bar)
            self._personel_tablo.setRowHeight(ri, 44)
        self._personel_tablo.resizeRowsToContents()

    def _admin_sekmeleri_yukle(self):
        if self.kullanici["role"] != "admin":
            return
        for table, veriler in [
            (self._users_table,  tum_kullanicilari_getir()),
            (self._login_table,  login_loglarini_getir()),
            (self._action_table, action_loglarini_getir()),
        ]:
            self._generic_tablo_doldur(table, veriler)
        # Mesaj logları
        if hasattr(self, '_mesaj_log_table'):
            self._mesaj_loglarini_yukle()

    def _generic_tablo_doldur(self, table: QTableWidget, veriler: list[dict]):
        table.clearContents()
        if not veriler:
            table.setRowCount(0); table.setColumnCount(0); return
        kolonlar = list(veriler[0].keys())
        table.setColumnCount(len(kolonlar))
        table.setRowCount(len(veriler))
        table.setHorizontalHeaderLabels(kolonlar)
        for ri, satir in enumerate(veriler):
            bg = QColor(P["row_white"]) if ri % 2 == 0 else QColor(P["surface2"])
            for ci, kol in enumerate(kolonlar):
                v = satir.get(kol,"")
                item = QTableWidgetItem("" if v is None else str(v))
                item.setBackground(bg)
                table.setItem(ri, ci, item)
        table.resizeRowsToContents()

    # ── TABLO GÖSTERİMİ ──────────────────────────────────────
    def _tablo_goster(self, veriler: list[dict]):
        self._siralama_sutun = -1
        self._siralama_artan = True
        # Header oklarını temizle
        for i, b in enumerate(self.BSL):
            h = self._table.horizontalHeaderItem(i)
            if h:
                h.setText(b)
        self._tablo_goster_ham(veriler)

    # ── FİLTRE & ARAMA ───────────────────────────────────────
    def _filtrele(self):
        v = self._data
        ilce  = self._ilce_cb.currentText()
        durum = self._durum_cb.currentText()
        if ilce  != "Tüm İlçeler":
            v = [r for r in v if (r.get("ilce") or "") == ilce]
        if durum != "Tüm Durumlar":
            def _esles(rd):
                d = (rd or "").upper()
                du = durum.upper()
                if "GÖN" in du or "GON" in du:
                    return "GÖN" in d or "GON" in d
                if "GEC" in du: return "GEC" in d
                if "ZİM" in du or "ZIM" in du:
                    return ("ZİMMETTE" == (rd or "").strip() or
                            "ZİMMETTE" in (rd or "") or
                            ("Z" in d and "MM" in d and "GEC" not in d and "GÖN" not in d and "GON" not in d))
                return d == "ARŞİVDE" or (("ARS" in d or "ARŞ" in d) and "GÖN" not in d and "GON" not in d)
            v = [r for r in v if _esles(r.get("durum",""))]
        self._tablo_goster(v)

    def _ara(self):
        aranan = self._arama.text().strip().lower()
        if not aranan:
            self._filtrele(); return
        v = [
            r for r in self._data
            if aranan in (r.get("orijinal_dosya_no") or "").lower()
            or aranan in (r.get("ilce") or "").lower()
            or aranan in (r.get("teslim_alan_personel") or "").lower()
            or aranan in (r.get("detay_no") or "").lower()
            or aranan in (r.get("ada") or "").lower()
            or aranan in (r.get("parsel") or "").lower()
        ]
        self._tablo_goster(v)

    def _tum_kayitlar(self):
        self._arama.clear()
        self._ilce_cb.setCurrentIndex(0)
        self._durum_cb.setCurrentIndex(0)
        self._tablo_goster(self._data)

    def _aktifleri_goster(self):
        v = sorted(
            [r for r in self._data
             if "Z" in (r.get("durum","")).upper() or
                "GEC" in (r.get("durum","")).upper()],
            key=lambda x: x["bekleme_gun"], reverse=True,
        )
        self._tablo_goster(v)

    def _gecikenleri_goster(self):
        v = sorted(
            [r for r in self._data if "GEC" in (r.get("durum","")).upper()],
            key=lambda x: x["bekleme_gun"], reverse=True,
        )
        self._tablo_goster(v)

    def _sutun_sirala(self, kolon_idx: int):
        """Sütun başlığına tıklayınca sırala."""
        if self._siralama_sutun == kolon_idx:
            self._siralama_artan = not self._siralama_artan
        else:
            self._siralama_sutun = kolon_idx
            self._siralama_artan = True

        kol = self.KOL[kolon_idx] if kolon_idx < len(self.KOL) else None
        if not kol:
            return

        def _anahtar(r):
            v = r.get(kol, "")
            if v is None:
                return ("", 0)
            try:
                return ("", int(v))
            except (ValueError, TypeError):
                return (str(v).lower(), 0)

        sirali = sorted(self._filtreli, key=_anahtar,
                        reverse=not self._siralama_artan)
        self._tablo_goster_ham(sirali)

        # Header'a ok işareti
        for i in range(self._table.columnCount()):
            h = self._table.horizontalHeaderItem(i)
            if h:
                base = self.BSL[i].replace(" ↑","").replace(" ↓","")
                if i == kolon_idx:
                    h.setText(base + (" ↑" if self._siralama_artan else " ↓"))
                else:
                    h.setText(base)

    def _tablo_goster_ham(self, veriler: list[dict]):
        """_filtreli'yi güncellemeden sadece tablo gösterimini yeniler."""
        self._filtreli = veriler
        t = self._table
        t.clearContents()
        t.setRowCount(len(veriler))
        t.setColumnCount(len(self.KOL))
        t.setHorizontalHeaderLabels(self.BSL)

        for ri, satir in enumerate(veriler):
            durum_str = (satir.get("durum") or "").upper()
            if "GEC" in durum_str:
                satir_bg = QColor(P["row_red"])
            elif "GÖN" in durum_str or "GON" in durum_str:
                satir_bg = QColor("#FFFBEB")   # amber — arşive gönderildi
            elif "Z" in durum_str and "MM" in durum_str:
                satir_bg = QColor(P["row_yellow"])
            else:
                satir_bg = QColor(P["row_white"])

            for ci, kol in enumerate(self.KOL):
                v = satir.get(kol)
                item = QTableWidgetItem("" if v is None else str(v))
                item.setTextAlignment(Qt.AlignCenter)
                item.setBackground(satir_bg)
                if kol == "durum":
                    f = QFont(); f.setBold(True); item.setFont(f)
                    if "GEC" in durum_str:
                        item.setForeground(QColor(P["red_t"]))
                    elif "GÖN" in durum_str or "GON" in durum_str:
                        item.setForeground(QColor("#92400E"))  # amber koyu
                    elif "Z" in durum_str and "MM" in durum_str:
                        item.setForeground(QColor(P["blue_t"]))
                    else:
                        item.setForeground(QColor(P["green_t"]))
                if kol == "bekleme_gun":
                    try:
                        gun = int(v) if v else 0
                    except (ValueError, TypeError):
                        gun = 0
                    if gun >= 10:
                        f2 = QFont(); f2.setBold(True); item.setFont(f2)
                        item.setForeground(QColor(P["red"]))
                t.setItem(ri, ci, item)
        t.resizeRowsToContents()



    def _dosya_sil(self):
        """Admin veya arşiv görevlisi dosyayı soft-delete ile siler."""
        if self.kullanici["role"] not in ["arsiv","admin"]:
            QMessageBox.warning(self, "Yetki", "Bu işlem için yetkiniz yok."); return
        s = self._secili_satir()
        if not s:
            QMessageBox.warning(self, "Uyarı", "Önce tablodan bir dosya seçin."); return

        # Zimmetli dosyayı silmeye çalışıyorsa uyar
        if s.get("durum") in ("ZİMMETTE", "GECİKMİŞ", "ARŞİVE GÖNDERİLDİ"):
            cevap = QMessageBox.warning(
                self, "Uyarı",
                f"<b>{s['orijinal_dosya_no']}</b> dosyası hâlâ <b>{s['durum']}</b> durumunda.\n\n"
                "Zimmetli dosyayı silmek istediğinizden emin misiniz?\n"
                "(Bu işlem geri alınamaz.)",
                QMessageBox.Yes | QMessageBox.No
            )
            if cevap != QMessageBox.Yes: return

        # Onay dialogu
        d = QDialog(self)
        d.setWindowTitle("Dosyayı Sil")
        d.setFixedWidth(400)
        d.setStyleSheet(f"background:{P['surface']};")
        dl = QVBoxLayout(d); dl.setContentsMargins(28,28,28,24); dl.setSpacing(14)
        ikon = QLabel("🗑"); ikon.setAlignment(Qt.AlignCenter); ikon.setStyleSheet("font-size:36px;")
        baslik = QLabel("Dosyayı Sil")
        baslik.setAlignment(Qt.AlignCenter)
        baslik.setStyleSheet(f"font-size:16px; font-weight:700; color:{P['txt']};")
        aciklama = QLabel(
            f"<b>{s['orijinal_dosya_no']}</b> — {s.get('ilce','')}\n\n"
            f"Bu dosya sistemden kaldırılacak.\n"
            f"Hareket geçmişi kayıtlarda tutulacak."
        )
        aciklama.setAlignment(Qt.AlignCenter)
        aciklama.setStyleSheet(f"font-size:13px; color:{P['txt3']};")
        aciklama.setWordWrap(True)
        bl = QHBoxLayout(); bl.setSpacing(10)
        iptal = QPushButton("İptal"); iptal.setFixedHeight(42)
        iptal.setStyleSheet(f"""
            QPushButton {{
                background:{P['bg']}; color:{P['txt']};
                border:1.5px solid {P['border']}; border-radius:10px;
                font-size:14px; font-weight:600;
            }}
            QPushButton:hover {{ background:{P['surface']}; }}
        """)
        iptal.clicked.connect(d.reject)
        sil_ok = QPushButton("🗑  Evet, Sil"); sil_ok.setFixedHeight(42)
        sil_ok.setStyleSheet(f"""
            QPushButton {{
                background:{P['red']}; color:white;
                border:none; border-radius:10px;
                font-size:14px; font-weight:700;
            }}
            QPushButton:hover {{ background:#DC2626; }}
        """)
        sil_ok.clicked.connect(d.accept)
        bl.addWidget(iptal); bl.addWidget(sil_ok)
        dl.addWidget(ikon); dl.addWidget(baslik); dl.addWidget(aciklama); dl.addLayout(bl)

        if d.exec() == QDialog.Accepted:
            try:
                file_sil(s["file_id"])
                action_log_ekle(
                    self.kullanici["id"], self.kullanici["username"],
                    self.kullanici["full_name"], self.kullanici["role"],
                    "DOSYA_SİL",
                    f"file_id={s['file_id']} dosya={s['orijinal_dosya_no']}"
                )
                QMessageBox.information(self, "Başarılı",
                    f"{s['orijinal_dosya_no']} silindi.")
                self.veriyi_yukle()
            except Exception as e:
                QMessageBox.critical(self, "Hata", str(e))

    def _secili_satir(self) -> dict | None:
        row = self._table.currentRow()
        if row < 0 or row >= len(self._filtreli): return None
        return self._filtreli[row]

    def _tablo_cift_tiklama(self):
        if s := self._secili_satir():
            GecmisDialog(s["file_id"], s["orijinal_dosya_no"]).exec()

    # ── AKSİYONLAR ───────────────────────────────────────────
    def _gecmis_goster(self):
        s = self._secili_satir()
        if not s:
            QMessageBox.warning(self, "Uyarı", "Önce tablodan bir dosya seçin."); return
        GecmisDialog(s["file_id"], s["orijinal_dosya_no"]).exec()

    def _zimmet_ekle(self):
        if self.kullanici["role"] not in ["arsiv","admin"]:
            QMessageBox.warning(self, "Yetki", "Bu işlem için yetkiniz yok."); return
        s = self._secili_satir()
        if not s:
            QMessageBox.warning(self, "Uyarı", "Önce tablodan bir dosya seçin."); return
        d = ZimmetEkleDialog(s["file_id"], s["orijinal_dosya_no"], self.kullanici)
        if d.exec():
            action_log_ekle(self.kullanici["id"], self.kullanici["username"],
                            self.kullanici["full_name"], self.kullanici["role"],
                            "ZİMMET_EKLE",
                            f"file_id={s['file_id']} dosya={s['orijinal_dosya_no']}")
            self.veriyi_yukle(); self._admin_sekmeleri_yukle()

    def _arsive_al(self):
        if self.kullanici["role"] not in ["arsiv","admin"]:
            QMessageBox.warning(self, "Yetki", "Bu işlem için yetkiniz yok."); return
        s = self._secili_satir()
        if not s:
            QMessageBox.warning(self, "Uyarı", "Önce tablodan bir dosya seçin."); return
        d = ArsiveAlDialog(s["file_id"], s["orijinal_dosya_no"], self.kullanici)
        if d.exec():
            action_log_ekle(self.kullanici["id"], self.kullanici["username"],
                            self.kullanici["full_name"], self.kullanici["role"],
                            "ARŞİVE_AL",
                            f"file_id={s['file_id']} dosya={s['orijinal_dosya_no']}")
            self.veriyi_yukle(); self._admin_sekmeleri_yukle()

    def _arsive_gonder(self):
        s = self._secili_satir()
        if not s:
            QMessageBox.warning(self, "Uyarı", "Önce tablodan bir dosya seçin."); return
        if s.get("durum") not in ("ZİMMETTE", "GECİKMİŞ"):
            QMessageBox.warning(self, "Uyarı", "Bu dosya zimmetli değil."); return
        self._arsive_gonder_with(s)

    def _arsive_gonder_with(self, s: dict):
        try:
            gorevliler = tum_arsiv_gorevlileri()
        except Exception:
            gorevliler = []

        if not gorevliler:
            QMessageBox.warning(self, "Uyarı",
                "Sistemde aktif arşiv görevlisi bulunamadı."); return

        # Dialog
        d = QDialog(self)
        d.setWindowTitle("Arşive Gönder")
        d.setFixedWidth(440)
        d.setStyleSheet(f"background:{P['surface']};")
        dl = QVBoxLayout(d); dl.setContentsMargins(28,28,28,24); dl.setSpacing(14)

        ikon = QLabel("📤")
        ikon.setAlignment(Qt.AlignCenter); ikon.setStyleSheet("font-size:36px;")

        baslik = QLabel(f"Dosyayı Arşive Gönder")
        baslik.setAlignment(Qt.AlignCenter)
        baslik.setStyleSheet(f"font-size:16px; font-weight:700; color:{P['txt']};")

        dosya_lbl = QLabel(
            f"<b>{s['orijinal_dosya_no']}</b>  —  {s.get('ilce','')}"
        )
        dosya_lbl.setAlignment(Qt.AlignCenter)
        dosya_lbl.setStyleSheet(f"font-size:13px; color:{P['txt3']};")

        # Arşiv görevlisi seç
        ag_lbl = QLabel("ARŞİV GÖREVLİSİ SEÇ")
        ag_lbl.setStyleSheet(
            f"font-size:11px; font-weight:700; color:{P['txt4']}; letter-spacing:0.5px;"
        )
        ag_cb = QComboBox()
        ag_cb.setFixedHeight(42)
        ag_cb.setStyleSheet(f"""
            QComboBox {{
                background:{P['bg']}; border:1.5px solid {P['border']};
                border-radius:10px; padding:0 12px;
                font-size:13px; color:{P['txt']};
            }}
            QComboBox:focus {{ border-color:{P['blue']}; }}
            QComboBox::drop-down {{ border:none; width:30px; }}
        """)
        for ag in gorevliler:
            ag_cb.addItem(
                f"  {ag['full_name']}  ({ROL_ETIKET.get(ag['role'], ag['role'])})",
                userData=ag
            )

        bilgi = QLabel(
            "📩 Arşiv görevlisine otomatik bildirim mesajı gönderilecek.\n"
            "Dosya görevlinin üzerine geçecek, arşive alınca tamamlanacak."
        )
        bilgi.setStyleSheet(
            f"font-size:12px; color:{P['txt4']}; "
            f"background:{P['blue_bg']}; border-radius:8px; padding:10px;"
        )
        bilgi.setWordWrap(True)

        bl = QHBoxLayout(); bl.setSpacing(10)
        iptal = QPushButton("İptal"); iptal.setFixedHeight(42)
        iptal.setStyleSheet(f"""
            QPushButton {{
                background:{P['bg']}; color:{P['txt']};
                border:1.5px solid {P['border']}; border-radius:10px;
                font-size:14px; font-weight:600;
            }}
            QPushButton:hover {{ background:{P['surface']}; }}
        """)
        iptal.clicked.connect(d.reject)

        gonder_btn = QPushButton("📤  Gönder"); gonder_btn.setFixedHeight(42)
        gonder_btn.setStyleSheet(f"""
            QPushButton {{
                background:#D97706; color:white;
                border:none; border-radius:10px;
                font-size:14px; font-weight:700;
            }}
            QPushButton:hover {{ background:#B45309; }}
        """)
        gonder_btn.clicked.connect(d.accept)

        bl.addWidget(iptal); bl.addWidget(gonder_btn)
        dl.addWidget(ikon); dl.addWidget(baslik); dl.addWidget(dosya_lbl)
        dl.addWidget(ag_lbl); dl.addWidget(ag_cb)
        dl.addWidget(bilgi); dl.addLayout(bl)

        if d.exec() == QDialog.Accepted:
            secili_ag = ag_cb.currentData()
            if not secili_ag: return
            try:
                arsive_gonder(
                    file_id=s["file_id"],
                    gonderen_id=self.kullanici["id"],
                    gonderen_isim=self.kullanici["full_name"],
                    arsiv_gorevlisi_id=secili_ag["id"],
                    arsiv_gorevlisi_isim=secili_ag["full_name"],
                )
                self.veriyi_yukle()
                QMessageBox.information(
                    self, "Başarılı",
                    f"{s['orijinal_dosya_no']} dosyası\n"
                    f"{secili_ag['full_name']} adlı arşiv görevlisine gönderildi.\n\n"
                    f"Görevliye bildirim mesajı iletildi."
                )
            except Exception as e:
                QMessageBox.critical(self, "Hata", str(e))

    def _arsive_gonder_iptal(self):
        """Arşive gönder işlemini geri al."""
        s = self._secili_satir()
        if not s: QMessageBox.warning(self,"Uyarı","Önce bir dosya seçin."); return
        if s.get("durum") != "ARŞİVE GÖNDERİLDİ":
            QMessageBox.warning(self,"Uyarı","Bu dosya arşive gönderilmiş değil."); return
        try:
            arsive_gonder_iptal(s["file_id"])
            self.veriyi_yukle()
        except Exception as e:
            QMessageBox.critical(self,"Hata",str(e))

    def _zimmet_pdf(self):
        s = self._secili_satir()
        if not s:
            QMessageBox.warning(self, "Uyarı", "Önce tablodan bir dosya seçin."); return
        yol, _ = QFileDialog.getSaveFileName(
            self, "Zimmet PDF Kaydet",
            f"zimmet_{s['orijinal_dosya_no'].replace('/','-')}.pdf",
            "PDF (*.pdf)"
        )
        if not yol: return
        try:
            zimmet_pdf_olustur(s, yol)
            QMessageBox.information(self, "Başarılı", f"PDF oluşturuldu:\n{yol}")
        except Exception as e:
            QMessageBox.critical(self, "Hata", f"PDF hatası:\n{e}")

    def _yeni_dosya(self):
        if self.kullanici["role"] not in ["arsiv","admin"]:
            QMessageBox.warning(self, "Yetki", "Bu işlem için yetkiniz yok."); return
        d = YeniDosyaDialog(self.kullanici)
        if d.exec():
            action_log_ekle(self.kullanici["id"], self.kullanici["username"],
                            self.kullanici["full_name"], self.kullanici["role"],
                            "YENİ_DOSYA", "Yeni dosya eklendi.")
            self.veriyi_yukle(); self._admin_sekmeleri_yukle()

    def _excelden_yukle(self):
        if self.kullanici["role"] not in ["arsiv","admin"]:
            QMessageBox.warning(self, "Yetki", "Bu işlem için yetkiniz yok."); return
        if QMessageBox.question(
            self, "Onay",
            "Excel'den yükleme mevcut verilerin üzerine yazar.\nDevam edilsin mi?"
        ) != QMessageBox.Yes:
            return
        try:
            _ilk_kurulum_excelden_aktar()
            action_log_ekle(self.kullanici["id"], self.kullanici["username"],
                            self.kullanici["full_name"], self.kullanici["role"],
                            "EXCEL_YUKLE", "Excel verisi yeniden yüklendi.")
            self.veriyi_yukle(); self._admin_sekmeleri_yukle()
            QMessageBox.information(self, "Başarılı", "Excel verisi yüklendi.")
        except Exception as e:
            QMessageBox.critical(self, "Hata", f"Aktarım başarısız:\n{e}")

    def _yeni_kullanici(self):
        if self.kullanici["role"] != "admin": return
        d = YeniKullaniciDialog()
        if d.exec():
            action_log_ekle(self.kullanici["id"], self.kullanici["username"],
                            self.kullanici["full_name"], self.kullanici["role"],
                            "YENİ_KULLANICI", "Yeni kullanıcı oluşturuldu.")
            self._admin_sekmeleri_yukle()

    def _kullanici_durum(self):
        if self.kullanici["role"] != "admin": return
        row = self._users_table.currentRow()
        if row < 0:
            QMessageBox.warning(self, "Uyarı", "Önce kullanıcı seçin."); return
        headers = [self._users_table.horizontalHeaderItem(i).text()
                   for i in range(self._users_table.columnCount())]
        uid    = int(self._users_table.item(row, headers.index("id")).text())
        active = int(self._users_table.item(row, headers.index("active")).text())
        uname  = self._users_table.item(row, headers.index("username")).text()
        kullanici_durum_degistir(uid, 0 if active == 1 else 1)
        action_log_ekle(self.kullanici["id"], self.kullanici["username"],
                        self.kullanici["full_name"], self.kullanici["role"],
                        "KULLANICI_DURUM", f"user_id={uid} username={uname}")
        self._admin_sekmeleri_yukle()
        QMessageBox.information(self, "Başarılı", "Kullanıcı durumu güncellendi.")

    def _kullanici_sifre_sifirla(self):
        if self.kullanici["role"] != "admin": return
        row = self._users_table.currentRow()
        if row < 0:
            QMessageBox.warning(self, "Uyarı", "Önce kullanıcı seçin."); return
        headers = [self._users_table.horizontalHeaderItem(i).text()
                   for i in range(self._users_table.columnCount())]
        uid   = int(self._users_table.item(row, headers.index("id")).text())
        uname = self._users_table.item(row, headers.index("username")).text()
        kullanici_sifre_sifirla(uid, "12345")
        action_log_ekle(self.kullanici["id"], self.kullanici["username"],
                        self.kullanici["full_name"], self.kullanici["role"],
                        "ŞİFRE_SIFIRLA", f"user_id={uid} username={uname}")
        self._admin_sekmeleri_yukle()
        QMessageBox.information(self, "Başarılı", "Şifre 12345 olarak sıfırlandı.")

    def gorunumu_excele_aktar(self):
        if not self._filtreli:
            QMessageBox.information(self, "Bilgi", "Aktarılacak veri yok."); return
        yol, _ = QFileDialog.getSaveFileName(
            self, "Excel'e Aktar", "arsiv_rapor.xlsx", "Excel (*.xlsx)")
        if not yol: return
        try:
            pd.DataFrame(self._filtreli).to_excel(yol, index=False)
            QMessageBox.information(self, "Başarılı", f"Excel oluşturuldu:\n{yol}")
        except Exception as e:
            QMessageBox.critical(self, "Hata", str(e))


# ─────────────────────────────────────────────────────────────
# EXCEL YÜKLEME
# ─────────────────────────────────────────────────────────────
def _ilk_kurulum_excelden_aktar():
    if not DOSYA_YOLU.exists():
        raise FileNotFoundError(f"ODS bulunamadı: {DOSYA_YOLU}")

    df = pd.read_excel(DOSYA_YOLU, engine="odf", header=1)
    df = df.dropna(axis=1, how="all")
    df = df.loc[:, ~df.columns.astype(str).str.startswith("Unnamed")]
    df.columns = [str(c).strip() for c in df.columns]
    df = df.rename(columns={
        "ARŞİVE TESLİMTARİHİ":       "ARŞİVE TESLİM TARİHİ",
        "ARŞİVGÖREVLİSİ ADI SOYADI": "ARŞİV GÖREVLİSİ ADI SOYADI",
    })
    for kol in ["VERİLDİĞİ TARİH", "ARŞİVE TESLİM TARİHİ"]:
        if kol in df.columns:
            df[kol] = pd.to_datetime(df[kol], errors="coerce", dayfirst=True)
    for kol in ["DOSYA NO","ŞEFLİĞİ",
                "TESLİM ALAN PERSONELİN ADI SOYADI",
                "ARŞİV GÖREVLİSİ ADI SOYADI"]:
        if kol in df.columns:
            df[kol] = df[kol].fillna("").astype(str).str.strip()

    satirlar = []
    for _, row in df.iterrows():
        dosya_no = str(row.get("DOSYA NO","")).strip()
        if not dosya_no or dosya_no.lower() == "nan":
            continue
        def _t(val):
            if pd.isna(val): return None
            return val.strftime("%Y-%m-%d") if hasattr(val,"strftime") else str(val)
        satirlar.append({
            "dosya_no":        dosya_no,
            "sefligi":         str(row.get("ŞEFLİĞİ","")).strip(),
            "teslim_alan":     str(row.get("TESLİM ALAN PERSONELİN ADI SOYADI","")).strip(),
            "arsiv_gorevlisi": str(row.get("ARŞİV GÖREVLİSİ ADI SOYADI","")).strip(),
            "teslim_tarihi":   _t(row.get("VERİLDİĞİ TARİH")),
            "iade_tarihi":     _t(row.get("ARŞİVE TESLİM TARİHİ")),
        })
    excel_verisini_yukle(satirlar)


# ─────────────────────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────────────────────
def main():
    tablo_olustur()
    varsayilan_kullanicilari_olustur()
    mesaj_tablolari_olustur()
    online_tablolari_olustur()
    try:
        migrate_legacy_dosyalar_if_needed()
    except Exception:
        pass

    # Başlangıçta otomatik yedek
    try:
        veritabani_yedekle()
    except Exception:
        pass

    app = QApplication(sys.argv)
    app.setStyleSheet(ANA_STIL)

    # İlk açılışta Excel'den yükle
    try:
        from db import tum_files_ozet
        if not tum_files_ozet():
            _ilk_kurulum_excelden_aktar()
    except Exception:
        pass

    login = LoginDialog()
    # X ile kapatmak = çıkış, hatalı şifre = tekrar dene (dialog kapanmaz)
    if login.exec() != QDialog.Accepted or login.kullanici is None:
        sys.exit(0)

    pencere = MainWindow(login.kullanici)
    pencere.show()
    # Online durumu başlat
    try:
        presence_guncelle(login.kullanici["id"])
    except Exception:
        pass

    # Okunmamış mesaj bildirimi
    try:
        sayi = okunmamis_mesaj_sayisi(login.kullanici["id"])
        if sayi > 0:
            msg = QMessageBox(pencere)
            msg.setWindowTitle("Okunmamış Mesaj")
            msg.setIcon(QMessageBox.Information)
            msg.setText(
                f"💬 <b>{sayi} okunmamış mesajınız var.</b><br><br>"
                "Mesajları görmek için sol menüden <b>Mesajlar</b>'a tıklayın."
            )
            msg.setStandardButtons(QMessageBox.Ok)
            msg.exec()
    except Exception:
        pass
    sys.exit(app.exec())


if __name__ == "__main__":
    main()
